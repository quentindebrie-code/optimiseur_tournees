"""
Optimiseur de Tournées – Assainissement / WC Chimiques
Dépôt fixe : Impasse Gaston Phoebus, Saint Sulpice la Pointe, 81370
"""

import streamlit as st
import pandas as pd
import folium
from streamlit_folium import st_folium
import requests
from geopy.geocoders import Nominatim
import time
from io import BytesIO
import datetime
import unicodedata
import re as _re_stock

try:
    from streamlit_sortables import sort_items as _sortables_sort_items
    HAS_SORTABLES = True
except ImportError:
    HAS_SORTABLES = False
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.pagebreak import Break

# reportlab – gestion UTF-8 native, aucun problème d'encodage
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, HRFlowable, PageBreak,
                                 Image as RLImage)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────

DEPOTS = {
    "Saint-Sulpice":   "Imp. Gaston Phoebus, 81370 Saint-Sulpice-la-Pointe",
    "Villemur-sur-Tarn": "1 Avenue du Président Roosevelt 31340 Villemur-sur-Tarn",
}
DEPOT_DEPART_DEFAULT = "Saint-Sulpice"
DEPOT_RETOUR_DEFAULT = "Saint-Sulpice"
OSRM_URL      = "http://router.project-osrm.org"

ACTION_COLORS = {
    "Nettoyer":      "#AEC6E8",
    "Déposer":       "#B7E5B4",
    "Retirer":       "#F4B8C1",
    "Chargement":    "#D4B8E0",
    "Déchargement":  "#FFE0B2",
}
ACTION_BORDER_COLORS = {
    "Nettoyer":      "#1f6aa5",
    "Déposer":       "#28a745",
    "Retirer":       "#dc3545",
    "Chargement":    "#7B1FA2",
    "Déchargement":  "#E65100",
}
ACTION_MAP_COLORS = {
    "Nettoyer":     "blue",
    "Déposer":      "green",
    "Retirer":      "red",
    "Chargement":   "purple",
    "Déchargement": "orange",
}
ACTION_MAP_ICONS = {
    "Nettoyer":     "tint",
    "Déposer":      "arrow-down",
    "Retirer":      "arrow-up",
    "Chargement":   "upload",
    "Déchargement": "download",
}


# Consignes par type d'action (affichées dans le PDF)
ACTION_CONSIGNES = {
    "Nettoyer": (
        "Vidanger intégralement la cuve des eaux usées."
        "Remplir le réservoir de chasse avec de l'eau propre en ajoutant une dose de produit sanitaire bleu."
        "Remplir le réservoir d'eau propre destiné exclusivement au lave-mains."
        "Nettoyer et désinfecter l'ensemble des surfaces intérieures à l'aide d'un spray désinfectant, puis rincer à l'eau claire."
        "Contrôler et réapprovisionner les consommables : papier hygiénique, gel désinfectant et savon."
        "Vérifier le bon état général de l'équipement (porte, serrure, plancher, cuves et accessoires)."
        "Signaler toute anomalie, dégradation ou dysfonctionnement constaté sur la fiche de tournée."
    ),
    "Déposer": (
        "Positionner le WC sur l'emplacement désigné par le client, en dehors des zones de passage et de tout obstacle, sur un sol stable et plan."
        "Vérifier la stabilité et l'aplomb de l'équipement."
        "Remplir le réservoir de chasse avec de l'eau propre en ajoutant une dose de produit sanitaire bleu."
        "Remplir le réservoir d'eau propre destiné exclusivement au lave-mains."
        "Nettoyer et désinfecter l'ensemble des surfaces intérieures à l'aide d'un spray désinfectant, puis rincer à l'eau claire si nécessaire."
        "Réapprovisionner les consommables : papier hygiénique, gel désinfectant et savon."
        "Informer le client de la mise en service de l'équipement et, en cas de première installation, lui remettre les consignes d'utilisation."
        "Prendre une photo une fois installé."
    ),
    "Retirer": (
        "Vidanger intégralement toutes les cuves avant l'enlèvement, y compris lorsqu'elles sont seulement partiellement remplies."
        "Vérifier la propreté de la zone d'intervention et s'assurer qu'aucune trace ou salissure ne subsiste après le retrait de l'équipement."
        "Contrôler l'état général du WC au moment du chargement et consigner toute dégradation, anomalie ou pièce manquante sur la fiche de tournée."
    ),
}

# ─────────────────────────────────────────────────────────────────────────────
# DURÉES AUTOMATIQUES PAR ACTION (WC chimique uniquement, × quantité)
# ─────────────────────────────────────────────────────────────────────────────

WC_DUREES_PAR_ACTION = {
    "Déposer":       20,   # minutes par WC chimique
    "Retirer":       15,
    "Nettoyer":      10,
    "Chargement":    15,
    "Déchargement":   5,
}

# Barème WC handicapé (options ignorées, durée × quantité)
WC_HANDICAPE_DUREES_PAR_ACTION = {
    "Nettoyer":      25,
    "Déposer":       30,
    "Retirer":       10,
    "Chargement":    20,
    "Déchargement":  10,
}

# Produits simples (Urinoir / Lave-main) : 10 min/produit, sauf Déposer = 15 min/produit
PRODUITS_SIMPLES = {"Urinoir", "Lave-main"}
PRODUIT_SIMPLE_DUREE_STD    = 10   # minutes par produit (toutes actions hors Déposer)
PRODUIT_SIMPLE_DUREE_DEPOSE = 15   # minutes par produit pour l'action Déposer

PAUSE_DEJEUNER_MIN = 30  # pause méridienne ajoutée une fois par tournée


def _auto_duree(action: str, produit: str, quantite) -> int | None:
    """
    Retourne la durée automatique (en minutes) pour un arrêt, ou None si le
    produit ne relève d'aucun barème automatique.

    - WC chimique      : barème WC_DUREES_PAR_ACTION            × quantité
    - WC handicapé     : barème WC_HANDICAPE_DUREES_PAR_ACTION  × quantité
                         (les options sont ignorées)
    - Urinoir / Lave-main : 10 min × quantité (15 min × quantité si action Déposer)
    """
    produit_norm = str(produit).strip()
    action_norm  = str(action).strip()

    try:
        qty = max(1, int(float(quantite) or 1))
    except (ValueError, TypeError):
        qty = 1

    # Produits simples : Urinoir / Lave-main
    if produit_norm in PRODUITS_SIMPLES:
        base = (PRODUIT_SIMPLE_DUREE_DEPOSE if action_norm == "Déposer"
                else PRODUIT_SIMPLE_DUREE_STD)
        return base * qty

    if produit_norm == "WC chimique":
        bareme = WC_DUREES_PAR_ACTION
    elif produit_norm == "WC handicapé":
        bareme = WC_HANDICAPE_DUREES_PAR_ACTION
    else:
        return None

    base = bareme.get(action_norm)
    if base is None:
        return None
    return base * qty


# ═════════════════════════════════════════════════════════════════════════════
# GESTION DU PARC : COULEURS, ARTICLES, STOCKS
# ═════════════════════════════════════════════════════════════════════════════
# Modèle de données
# -----------------
# Le parc est décrit par un tableau à 4 colonnes :
#     Article | Couleur | Installés | En stock
# « Article » est la référence commerciale telle que manipulée par l'exploitation
# (ex. « WC chimique + Lave-main »). Elle est dérivée du couple Produit/Option
# saisi dans le tableau des arrêts via _article_key().
# « Installés » = unités actuellement chez les clients ; « En stock » = unités
# disponibles au dépôt. Parc total = Installés + En stock.
# ═════════════════════════════════════════════════════════════════════════════

COULEURS_PRODUITS = ["Vert", "Jaune", "Rose", "Bleu", "Orange", "Crème", "Blanc", "Gris"]

# Pastilles de couleur (fond) + couleur de texte lisible sur ce fond
COULEUR_HEX = {
    "Vert":   "#2E7D32",
    "Jaune":  "#F9A825",
    "Rose":   "#EC407A",
    "Bleu":   "#1E88E5",
    "Orange": "#EF6C00",
    "Crème":  "#D9C89E",
    "Blanc":  "#FFFFFF",
    "Gris":   "#9E9E9E",
    "":       "#E0E0E0",
}
COULEUR_TEXTE = {
    "Jaune": "#3E2723", "Crème": "#3E2723", "Blanc": "#3E2723",
    "":      "#616161",
}

# Catalogue produits proposé dans la colonne « Produit » du tableau de saisie
PRODUITS_CATALOGUE = ["WC chimique", "WC handicapé", "WC Luxe",
                      "Lave-main", "Urinoir", "WC client"]

# Matériel appartenant au client : jamais décompté du stock Deldossi
PRODUITS_HORS_STOCK = {"WC client"}

# Référentiel des articles gérés en stock (ordre d'affichage)
ARTICLES_STOCK = [
    "WC chimique + Lave-main",
    "WC chimique + Urinoir",
    "WC chimique",
    "WC handicapé",
    "WC Luxe",
    "Lave-main",
    "Urinoir",
]

STOCK_COLUMNS = ["Article", "Couleur", "Installés", "En stock"]

# ─────────────────────────────────────────────────────────────────────────────
# Mouvements de stock par action
# ─────────────────────────────────────────────────────────────────────────────
# −1 : l'unité quitte le dépôt (sortie de stock, entrée en parc installé)
# +1 : l'unité revient au dépôt (entrée en stock, sortie du parc installé)
#  0 : neutre (intervention sur place, aucun mouvement)
#
# HYPOTHÈSE À VALIDER : « Chargement » et « Déchargement » sont considérés comme
# neutres par défaut, car ils décrivent des manutentions dont le sens n'est pas
# univoque. Un interrupteur dans l'onglet Stock permet de les traiter comme des
# mouvements réels (Chargement = retour dépôt, Déchargement = dépose sur site).
# ─────────────────────────────────────────────────────────────────────────────

ACTION_MOUVEMENT_STOCK = {
    "Déposer":      -1,
    "Retirer":      +1,
    "Nettoyer":      0,
    "Chargement":    0,
    "Déchargement":  0,
}
ACTION_MOUVEMENT_STOCK_ETENDU = {
    "Déposer":      -1,
    "Retirer":      +1,
    "Nettoyer":      0,
    "Chargement":   +1,   # matériel récupéré et ramené au dépôt
    "Déchargement": -1,   # matériel déposé sur site
}


def _norm_txt(s) -> str:
    """Normalise une chaîne : minuscules, sans accents, espaces compactés."""
    s = str(s if s is not None else "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def _article_key(produit, option="") -> str:
    """Construit la référence article à partir du couple Produit / Option.

    Ex. ('WC chimique', 'Lave-main') → 'WC chimique + Lave-main'
        ('Urinoir', '')              → 'Urinoir'
    Retourne '' pour les produits hors stock (WC client) ou produit vide.
    """
    p = str(produit or "").strip()
    o = str(option or "").strip()
    if not p or p in PRODUITS_HORS_STOCK:
        return ""
    if p == "WC chimique" and o:
        return f"WC chimique + {o}"
    return p


def normalize_article(label) -> str:
    """Reconnaît une référence article écrite librement (mail, Excel exploitant).

    Tolère les variantes rencontrées sur le terrain :
        'WC lave main', 'WL PMR', 'Laves mains', 'Urinoirs', '2 WC Luxe'…
    Retourne la référence canonique, ou le libellé nettoyé si non reconnu.
    """
    raw = str(label or "").strip()
    n = _norm_txt(raw)
    if not n:
        return ""
    # Retire un éventuel préfixe numérique ("2 WC Luxe" → "wc luxe")
    n = _re_stock.sub(r"^\d+\s*[x×]?\s*", "", n)

    has_wc = bool(_re_stock.search(r"\bw[cl]\b", n))  # 'WL' = coquille fréquente pour 'WC'
    if "pmr" in n or "handicap" in n:
        return "WC handicapé"
    if "luxe" in n:
        return "WC Luxe"
    if has_wc and "lave" in n:
        return "WC chimique + Lave-main"
    if has_wc and "urinoir" in n:
        return "WC chimique + Urinoir"
    if "lave" in n and "main" in n:
        return "Lave-main"
    if "urinoir" in n:
        return "Urinoir"
    if has_wc:
        return "WC chimique"
    return raw


def normalize_couleur(label) -> str:
    """Reconnaît une couleur écrite librement ('Jaunes', 'creme', 'VERT')."""
    n = _norm_txt(label)
    if not n:
        return ""
    for c in COULEURS_PRODUITS:
        cn = _norm_txt(c)
        if n == cn or n == cn + "s" or n.startswith(cn):
            return c
    return str(label).strip().capitalize()


def _stock_default_df() -> pd.DataFrame:
    """Parc initial issu du relevé transmis par l'exploitation.

    Source : mail exploitation (WC installés / WC en stock). Ces valeurs servent
    d'amorce et sont écrasées dès qu'un fichier Excel de stock est importé.
    """
    rows = [
        # Article,                    Couleur,  Installés, En stock
        ("WC chimique + Lave-main",   "Vert",     3, 1),
        ("WC chimique + Lave-main",   "Jaune",    1, 4),
        ("WC chimique + Lave-main",   "Rose",     3, 1),
        ("WC chimique + Lave-main",   "Bleu",     1, 3),
        ("WC chimique + Lave-main",   "Orange",   0, 1),
        ("WC chimique + Urinoir",     "Jaune",    0, 2),
        ("WC chimique + Urinoir",     "Orange",   0, 2),
        ("WC handicapé",              "Vert",     0, 1),
        ("WC Luxe",                   "Crème",    0, 2),
        ("Lave-main",                 "Jaune",    0, 5),
        ("Urinoir",                   "Vert",     0, 4),
    ]
    return pd.DataFrame(rows, columns=STOCK_COLUMNS)


def _clean_stock_df(df: pd.DataFrame) -> pd.DataFrame:
    """Nettoie, normalise et agrège un tableau de stock.

    - Complète les colonnes manquantes
    - Normalise articles et couleurs
    - Convertit les quantités en entiers ≥ 0
    - Agrège les doublons (Article, Couleur)
    - Trie selon l'ordre du référentiel ARTICLES_STOCK
    """
    if df is None or len(df) == 0:
        return pd.DataFrame(columns=STOCK_COLUMNS)
    out = df.copy()
    for col in STOCK_COLUMNS:
        if col not in out.columns:
            out[col] = "" if col in ("Article", "Couleur") else 0
    out = out[STOCK_COLUMNS]
    out["Article"] = out["Article"].map(normalize_article)
    out["Couleur"] = out["Couleur"].map(normalize_couleur)
    for col in ("Installés", "En stock"):
        out[col] = (pd.to_numeric(out[col], errors="coerce")
                    .fillna(0).clip(lower=0).astype(int))
    out = out[out["Article"].astype(str).str.strip() != ""]
    out = (out.groupby(["Article", "Couleur"], as_index=False)[["Installés", "En stock"]]
              .sum())
    ordre = {a: i for i, a in enumerate(ARTICLES_STOCK)}
    ordre_c = {c: i for i, c in enumerate(COULEURS_PRODUITS)}
    out["_o"] = out["Article"].map(lambda a: ordre.get(a, 999))
    out["_c"] = out["Couleur"].map(lambda c: ordre_c.get(c, 999))
    out = (out.sort_values(["_o", "Article", "_c", "Couleur"])
              .drop(columns=["_o", "_c"]).reset_index(drop=True))
    return out


def parse_stock_excel(uploaded_file):
    """Lit un fichier Excel de stock et retourne (DataFrame | None, message).

    Détection tolérante :
    - onglet contenant 'stock' ou 'parc' (sinon premier onglet)
    - colonnes reconnues par mots-clés, quelle que soit la casse ou l'accentuation
    """
    try:
        xls = pd.ExcelFile(uploaded_file)
    except Exception as e:
        return None, f"Fichier illisible : {e}"

    sheet = None
    for s in xls.sheet_names:
        ns = _norm_txt(s)
        if "stock" in ns or "parc" in ns:
            sheet = s
            break
    sheet = sheet or xls.sheet_names[0]

    # Lecture brute : la ligne d'en-tête n'est pas nécessairement la première
    # (les classeurs exportés par l'application comportent un titre et une date).
    try:
        brut = xls.parse(sheet, header=None)
    except Exception as e:
        return None, f"Onglet « {sheet} » illisible : {e}"

    if brut is None or len(brut) == 0:
        return None, f"L'onglet « {sheet} » est vide."

    def _role_colonne(valeur):
        """Identifie le rôle d'un libellé de colonne, ou None."""
        brut = str(valeur if valeur is not None else "").strip()
        if len(brut) > 40:      # une phrase n'est pas un libellé de colonne
            return None
        n = _norm_txt(brut)
        if not n:
            return None
        if "article" in n or "produit" in n or "reference" in n or "designation" in n:
            return "Article"
        if "couleur" in n or "coloris" in n:
            return "Couleur"
        if "install" in n or "pose" in n or "chez le client" in n or "parc client" in n:
            return "Installés"
        if "stock" in n or "dispo" in n or "depot" in n:
            return "En stock"
        return None

    # Recherche de la ligne d'en-tête dans les 15 premières lignes
    header_row, mapping = None, {}
    for i in range(min(15, len(brut))):
        candidat = {}
        for j, val in enumerate(brut.iloc[i]):
            role = _role_colonne(val)
            if role and role not in candidat:
                candidat[role] = j
        if "Article" in candidat and ("Installés" in candidat or "En stock" in candidat):
            header_row, mapping = i, candidat
            break

    if header_row is None:
        apercu = ", ".join(str(v) for v in brut.iloc[0].tolist()[:6])
        return None, (
            f"En-tête introuvable dans l'onglet « {sheet} ». Le fichier doit "
            f"comporter une ligne d'en-tête avec au minimum une colonne "
            f"« Article » et une colonne « Installés » ou « En stock ». "
            f"Première ligne lue : {apercu}"
        )

    corps = brut.iloc[header_row + 1:]
    n_lignes = len(corps)

    def _col(role, defaut):
        return corps.iloc[:, mapping[role]] if role in mapping else pd.Series(
            [defaut] * n_lignes, index=corps.index)

    df = pd.DataFrame({
        "Article":   _col("Article",   ""),
        "Couleur":   _col("Couleur",   ""),
        "Installés": _col("Installés", 0),
        "En stock":  _col("En stock",  0),
    })

    # Écarte les lignes de synthèse et les notes de bas de tableau
    _exclus = {"total", "totaux", "total general", "somme", "note", "nan"}
    df = df[~df["Article"].map(lambda v: _norm_txt(v) in _exclus
                               or _norm_txt(v).startswith("note :"))]

    df = _clean_stock_df(df)
    if len(df) == 0:
        return None, f"Aucune ligne exploitable dans l'onglet « {sheet} »."
    return df, f"{len(df)} référence(s) importée(s) depuis l'onglet « {sheet} »."


def export_stock_excel(df_stock: pd.DataFrame, df_projete: pd.DataFrame | None = None,
                       maj_date=None) -> BytesIO:
    """Génère le classeur Excel du parc (onglet 'Stock' réimportable).

    - Onglet « Stock »          : état courant, colonnes attendues à la réimport
    - Onglet « Synthèse »       : matrice Article × Couleur (stock disponible)
    - Onglet « Stock projeté »  : état après exécution de la tournée (optionnel)
    """
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    bold_f   = Font(bold=True)
    center   = Alignment(horizontal="center", vertical="center")
    left     = Alignment(horizontal="left", vertical="center")
    thin     = Side(style="thin", color="CCCCCC")
    brd      = Border(left=thin, right=thin, top=thin, bottom=thin)
    warn_fill = PatternFill("solid", fgColor="F8D7DA")
    low_fill  = PatternFill("solid", fgColor="FFF3CD")

    def _write_table(ws, df, titre):
        ws.append([titre])
        tr = ws.max_row
        ws.merge_cells(f"A{tr}:E{tr}")
        c = ws.cell(tr, 1)
        c.font = Font(bold=True, size=14, color="1F4E79")
        c.alignment = center
        ws.row_dimensions[tr].height = 26
        if maj_date:
            ws.append([f"État au {maj_date}"])
            ws.merge_cells(f"A{ws.max_row}:E{ws.max_row}")
            ws.cell(ws.max_row, 1).font = Font(italic=True, size=9, color="666666")
        ws.append([])
        headers = STOCK_COLUMNS + ["Parc total"]
        ws.append(headers)
        hr = ws.max_row
        for col, h in enumerate(headers, 1):
            cc = ws.cell(hr, col)
            cc.fill = hdr_fill; cc.font = hdr_font
            cc.alignment = center; cc.border = brd
        ws.row_dimensions[hr].height = 20
        for _, row in df.iterrows():
            ws.append([row["Article"], row["Couleur"],
                       int(row["Installés"]), int(row["En stock"]),
                       int(row["Installés"]) + int(row["En stock"])])
            r = ws.max_row
            for col in range(1, 6):
                cc = ws.cell(r, col)
                cc.border = brd
                cc.alignment = left if col <= 2 else center
            if int(row["En stock"]) == 0:
                ws.cell(r, 4).fill = warn_fill
                ws.cell(r, 4).font = Font(bold=True, color="B00020")
            elif int(row["En stock"]) <= 1:
                ws.cell(r, 4).fill = low_fill
        # Ligne de total
        ws.append(["TOTAL", "", int(df["Installés"].sum()), int(df["En stock"].sum()),
                   int(df["Installés"].sum() + df["En stock"].sum())])
        r = ws.max_row
        for col in range(1, 6):
            cc = ws.cell(r, col)
            cc.font = bold_f; cc.border = brd
            cc.fill = PatternFill("solid", fgColor="DCE6F1")
            cc.alignment = left if col <= 2 else center
        for col, w in zip(range(1, 6), [30, 14, 12, 12, 12]):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws = wb.active
    ws.title = "Stock"
    _write_table(ws, df_stock, "PARC & STOCK – DELDOSSI ASSAINISSEMENT")

    # ── Onglet synthèse : matrice Article × Couleur ──
    ws2 = wb.create_sheet("Synthèse")
    couleurs = [c for c in COULEURS_PRODUITS
                if c in set(df_stock["Couleur"].astype(str))] or [""]
    ws2.append(["MATRICE STOCK DISPONIBLE – ARTICLE × COULEUR"])
    ws2.merge_cells(start_row=1, start_column=1,
                    end_row=1, end_column=max(2, len(couleurs) + 2))
    ws2.cell(1, 1).font = Font(bold=True, size=14, color="1F4E79")
    ws2.cell(1, 1).alignment = center
    ws2.append([])
    ws2.append(["Article"] + couleurs + ["Total"])
    hr = ws2.max_row
    for col in range(1, len(couleurs) + 3):
        cc = ws2.cell(hr, col)
        cc.fill = hdr_fill; cc.font = hdr_font
        cc.alignment = center; cc.border = brd
    for art in [a for a in ARTICLES_STOCK if a in set(df_stock["Article"])]:
        sub = df_stock[df_stock["Article"] == art]
        vals = []
        for c in couleurs:
            v = sub[sub["Couleur"] == c]["En stock"].sum()
            vals.append(int(v))
        ws2.append([art] + vals + [int(sum(vals))])
        r = ws2.max_row
        for col in range(1, len(couleurs) + 3):
            cc = ws2.cell(r, col)
            cc.border = brd
            cc.alignment = left if col == 1 else center
            if col > 1 and col <= len(couleurs) + 1 and cc.value == 0:
                cc.font = Font(color="BBBBBB")
    ws2.column_dimensions["A"].width = 30
    for col in range(2, len(couleurs) + 3):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

    # ── Onglet stock projeté (après tournée) ──
    if df_projete is not None and len(df_projete) > 0:
        ws3 = wb.create_sheet("Stock projeté")
        _write_table(ws3, df_projete, "PARC & STOCK PROJETÉ APRÈS TOURNÉE")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def besoins_tournee(df_stops: pd.DataFrame, inclure_manutentions: bool = False) -> pd.DataFrame:
    """Calcule les mouvements de stock générés par les arrêts saisis.

    Retourne un DataFrame : Article | Couleur | Sorties | Retours | Net
      - Sorties : unités qui quittent le dépôt (à prélever en stock)
      - Retours : unités qui reviennent au dépôt (à réintégrer en stock)
      - Net     : Retours − Sorties (variation prévisionnelle du stock)
    Les lignes sans adresse et les produits hors stock (WC client) sont ignorés.
    """
    cols = ["Article", "Couleur", "Sorties", "Retours", "Net"]
    if df_stops is None or len(df_stops) == 0:
        return pd.DataFrame(columns=cols)

    bareme = ACTION_MOUVEMENT_STOCK_ETENDU if inclure_manutentions else ACTION_MOUVEMENT_STOCK
    agg = {}
    for _, row in df_stops.iterrows():
        if str(row.get("Adresse", "") or "").strip() == "":
            continue
        art = _article_key(row.get("Produit", ""), row.get("Option", ""))
        if not art:
            continue
        sens = bareme.get(str(row.get("Action", "")).strip(), 0)
        if sens == 0:
            continue
        try:
            qty = max(1, int(float(row.get("Quantité", 1) or 1)))
        except (ValueError, TypeError):
            qty = 1
        couleur = normalize_couleur(row.get("Couleur", ""))
        key = (art, couleur)
        cur = agg.setdefault(key, {"Sorties": 0, "Retours": 0})
        if sens < 0:
            cur["Sorties"] += qty
        else:
            cur["Retours"] += qty

    if not agg:
        return pd.DataFrame(columns=cols)

    rows = [{"Article": a, "Couleur": c,
             "Sorties": v["Sorties"], "Retours": v["Retours"],
             "Net": v["Retours"] - v["Sorties"]}
            for (a, c), v in agg.items()]
    df = pd.DataFrame(rows, columns=cols)
    ordre = {a: i for i, a in enumerate(ARTICLES_STOCK)}
    ordre_c = {c: i for i, c in enumerate(COULEURS_PRODUITS)}
    df["_o"] = df["Article"].map(lambda a: ordre.get(a, 999))
    df["_c"] = df["Couleur"].map(lambda c: ordre_c.get(c, 999))
    return (df.sort_values(["_o", "Article", "_c", "Couleur"])
              .drop(columns=["_o", "_c"]).reset_index(drop=True))


def controle_disponibilite(df_stock: pd.DataFrame, df_besoins: pd.DataFrame) -> pd.DataFrame:
    """Confronte les besoins de la tournée au stock disponible.

    Retourne un DataFrame : Article | Couleur | Besoin | Disponible | Manque | Statut
    Règle couleur : si la couleur n'est pas renseignée (choix laissé au client),
    le besoin est confronté au stock **toutes couleurs confondues** de l'article.
    """
    cols = ["Article", "Couleur", "Besoin", "Disponible", "Manque", "Statut"]
    if df_besoins is None or len(df_besoins) == 0:
        return pd.DataFrame(columns=cols)

    stock = _clean_stock_df(df_stock)
    rows = []
    for _, b in df_besoins.iterrows():
        besoin = int(b["Sorties"])
        if besoin <= 0:
            continue
        art, coul = b["Article"], b["Couleur"]
        if coul:
            dispo = int(stock[(stock["Article"] == art) &
                              (stock["Couleur"] == coul)]["En stock"].sum())
            libelle_coul = coul
        else:
            dispo = int(stock[stock["Article"] == art]["En stock"].sum())
            libelle_coul = "Toutes couleurs"
        manque = max(0, besoin - dispo)
        if manque > 0:
            statut = "🔴 Rupture"
        elif dispo - besoin == 0:
            statut = "🟠 Juste suffisant"
        else:
            statut = "🟢 Disponible"
        rows.append({"Article": art, "Couleur": libelle_coul, "Besoin": besoin,
                     "Disponible": dispo, "Manque": manque, "Statut": statut})
    return pd.DataFrame(rows, columns=cols)


def projeter_stock(df_stock: pd.DataFrame, df_besoins: pd.DataFrame) -> pd.DataFrame:
    """Applique les mouvements de la tournée au parc et retourne l'état projeté.

    Les besoins sans couleur sont imputés sur la couleur la mieux dotée
    (allocation gloutonne), afin de rester cohérent avec la réalité terrain.
    """
    stock = _clean_stock_df(df_stock)
    if df_besoins is None or len(df_besoins) == 0:
        return stock

    def _idx(art, coul):
        m = stock.index[(stock["Article"] == art) & (stock["Couleur"] == coul)]
        return m[0] if len(m) else None

    for _, b in df_besoins.iterrows():
        art, coul = b["Article"], b["Couleur"]
        for qte, sens in ((int(b["Sorties"]), -1), (int(b["Retours"]), +1)):
            if qte <= 0:
                continue
            if coul:
                cibles = [(coul, qte)]
            else:
                # Allocation gloutonne : on sert d'abord la couleur la mieux dotée
                sub = stock[stock["Article"] == art].sort_values(
                    "En stock", ascending=(sens > 0))
                cibles, reste = [], qte
                for _, l in sub.iterrows():
                    if reste <= 0:
                        break
                    part = min(reste, int(l["En stock"])) if sens < 0 else reste
                    part = max(part, 0)
                    if part > 0:
                        cibles.append((l["Couleur"], part))
                        reste -= part
                if reste > 0:
                    cibles.append((sub.iloc[0]["Couleur"] if len(sub) else "", reste))
            for c, part in cibles:
                i = _idx(art, c)
                if i is None:
                    stock = pd.concat([stock, pd.DataFrame(
                        [{"Article": art, "Couleur": c, "Installés": 0, "En stock": 0}]
                    )], ignore_index=True)
                    i = _idx(art, c)
                stock.at[i, "En stock"]  = max(0, int(stock.at[i, "En stock"]) + sens * part)
                stock.at[i, "Installés"] = max(0, int(stock.at[i, "Installés"]) - sens * part)
    return _clean_stock_df(stock)


def _besoins_records_pour_export(df_stops, df_stock, inclure_manutentions=False) -> list:
    """Prépare la liste de dicts des mouvements matériel, enrichie du stock avant
    tournée et de l'éventuel manque, pour alimenter les exports Excel et PDF."""
    bes = besoins_tournee(df_stops, inclure_manutentions=inclure_manutentions)
    if len(bes) == 0:
        return []
    stock = _clean_stock_df(df_stock)
    records = []
    for _, b in bes.iterrows():
        art, coul = b["Article"], b["Couleur"]
        if coul:
            dispo = int(stock[(stock["Article"] == art) &
                              (stock["Couleur"] == coul)]["En stock"].sum())
            lbl_dispo = f"{dispo} ({coul})"
        else:
            dispo = int(stock[stock["Article"] == art]["En stock"].sum())
            lbl_dispo = f"{dispo} (toutes couleurs)"
        records.append({
            "Article":    art,
            "Couleur":    coul,
            "Sorties":    int(b["Sorties"]),
            "Retours":    int(b["Retours"]),
            "Net":        int(b["Net"]),
            "StockAvant": lbl_dispo,
            "Manque":     max(0, int(b["Sorties"]) - dispo),
        })
    return records


def _pastille_couleur(couleur, texte=None, taille="0.78em") -> str:
    """Retourne le HTML d'une pastille colorée (badge) pour une couleur produit."""
    coul = str(couleur or "").strip()
    bg   = COULEUR_HEX.get(coul, "#E0E0E0")
    fg   = COULEUR_TEXTE.get(coul, "#FFFFFF")
    lbl  = texte if texte is not None else (coul or "Non spécifiée")
    bord = ";border:1px solid #BDBDBD" if coul in ("Blanc", "Crème", "") else ""
    return (f'<span style="background:{bg};color:{fg};padding:1px 8px;'
            f'border-radius:10px;font-size:{taille};white-space:nowrap{bord}">'
            f'{lbl}</span>')


# ═════════════════════════════════════════════════════════════════════════════
# IMPORT / EXPORT DE LA TOURNÉE (saisie des arrêts)
# ═════════════════════════════════════════════════════════════════════════════
# Objectif : permettre à l'exploitation de préparer une tournée dans Excel puis
# de l'injecter dans l'application en un clic, plutôt que de tout ressaisir.
# Le classeur produit par export_stops_excel() est directement réimportable.
# ═════════════════════════════════════════════════════════════════════════════

ACTIONS_CATALOGUE = ["Nettoyer", "Déposer", "Retirer", "Chargement", "Déchargement"]
OPTIONS_CATALOGUE = ["Lave-main", "Urinoir"]

# Colonnes métier du tableau de saisie (hors colonne technique « Suppr »)
STOPS_COLUMNS = ["Action", "Produit", "Option", "Couleur", "Quantité",
                 "Nom du client", "Adresse", "Durée (min)",
                 "Pas avant", "Pas après", "Observations"]


def normalize_action(valeur) -> tuple[str, bool]:
    """Reconnaît une action écrite librement. Retourne (action, reconnue)."""
    n = _norm_txt(valeur)
    if not n:
        return "Nettoyer", False
    exact = {_norm_txt(a): a for a in ACTIONS_CATALOGUE}
    if n in exact:
        return exact[n], True
    # « dechargement » contient « chargement » : on teste le plus long d'abord
    for a in sorted(ACTIONS_CATALOGUE, key=lambda x: -len(x)):
        if _norm_txt(a) in n or n in _norm_txt(a):
            return a, True
    return "Nettoyer", False


def normalize_produit(valeur) -> tuple[str, bool]:
    """Reconnaît un produit écrit librement. Retourne (produit, reconnu)."""
    n = _norm_txt(valeur)
    if not n:
        return "WC chimique", False
    exact = {_norm_txt(p): p for p in PRODUITS_CATALOGUE}
    if n in exact:
        return exact[n], True
    if "client" in n:
        return "WC client", True
    if "pmr" in n or "handicap" in n:
        return "WC handicapé", True
    if "luxe" in n:
        return "WC Luxe", True
    if "lave" in n and "main" in n and not _re_stock.search(r"\bw[cl]\b", n):
        return "Lave-main", True
    if "urinoir" in n and not _re_stock.search(r"\bw[cl]\b", n):
        return "Urinoir", True
    if _re_stock.search(r"\bw[cl]\b", n) or "chimique" in n:
        return "WC chimique", True
    return "WC chimique", False


def normalize_option_produit(valeur) -> tuple[str, bool]:
    """Reconnaît une option (Lave-main / Urinoir). Retourne (option, reconnue)."""
    n = _norm_txt(valeur)
    if not n or n in ("-", "—", "aucune", "aucun", "nan"):
        return "", True
    if "lave" in n:
        return "Lave-main", True
    if "urinoir" in n:
        return "Urinoir", True
    return "", False


def _hhmm_from_cell(valeur) -> str:
    """Convertit une cellule Excel en 'HH:MM', ou '' si vide/illisible.

    Gère les objets time/datetime d'Excel, les fractions de journée (0,375 = 09:00)
    et les saisies texte libres ('9h30', '9:30', '0930').
    """
    if valeur is None:
        return ""
    try:
        if isinstance(valeur, float) and pd.isna(valeur):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(valeur, datetime.datetime):
        return valeur.strftime("%H:%M")
    if isinstance(valeur, datetime.time):
        return valeur.strftime("%H:%M")
    if isinstance(valeur, (int, float)) and not isinstance(valeur, bool):
        v = float(valeur)
        if 0 <= v < 1:                      # fraction de journée
            total = int(round(v * 24 * 60))
            return f"{total // 60:02d}:{total % 60:02d}"
        if 0 <= v <= 24:                    # heure entière saisie en nombre
            return f"{int(v):02d}:00"
        return ""
    s = str(valeur).strip()
    if not s or _norm_txt(s) in ("nan", "-", "—"):
        return ""
    m = _re_stock.match(r"^(\d{1,2})\s*[:hH.]\s*(\d{0,2})$", s)
    if m:
        h = int(m.group(1)); mi = int(m.group(2) or 0)
        if 0 <= h <= 23 and 0 <= mi <= 59:
            return f"{h:02d}:{mi:02d}"
    m = _re_stock.match(r"^(\d{2})(\d{2})$", s)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        if 0 <= h <= 23 and 0 <= mi <= 59:
            return f"{h:02d}:{mi:02d}"
    return ""


def _stops_template_df() -> pd.DataFrame:
    """Tableau vide au format attendu (utilisé pour le modèle vierge)."""
    return pd.DataFrame({c: pd.Series(dtype="object") for c in STOPS_COLUMNS})


def export_stops_excel(df_stops: pd.DataFrame | None = None,
                       nb_lignes_vides: int = 40,
                       tour_date=None, driver_name="") -> BytesIO:
    """Génère le classeur de saisie de tournée.

    - Onglet « Tournée » : le tableau de saisie, avec listes déroulantes
      (Action, Produit, Option, Couleur) pour éviter les fautes de frappe.
    - Onglet « Notice »  : mode d'emploi et exemple commenté.
    - Onglet « Listes »  : valeurs autorisées (masqué), support des validations.

    df_stops = None → modèle vierge ; sinon export de la saisie en cours.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left     = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin     = Side(style="thin", color="CCCCCC")
    brd      = Border(left=thin, right=thin, top=thin, bottom=thin)
    oblig    = PatternFill("solid", fgColor="FFF3CD")

    # ── Onglet « Listes » (support des menus déroulants) ──
    wl = wb.create_sheet("Listes")
    listes = {
        "A": ("Action",  ACTIONS_CATALOGUE),
        "B": ("Produit", PRODUITS_CATALOGUE),
        "C": ("Option",  OPTIONS_CATALOGUE),
        "D": ("Couleur", COULEURS_PRODUITS),
    }
    for col, (titre, vals) in listes.items():
        wl[f"{col}1"] = titre
        wl[f"{col}1"].font = Font(bold=True)
        for i, v in enumerate(vals, start=2):
            wl[f"{col}{i}"] = v
    wl.sheet_state = "hidden"

    # ── Onglet « Tournée » ──
    ws = wb.active
    ws.title = "Tournée"

    titre = "SAISIE DE TOURNÉE – DELDOSSI ASSAINISSEMENT"
    ws.append([titre])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(STOPS_COLUMNS))
    ws.cell(1, 1).font = Font(bold=True, size=15, color="1F4E79")
    ws.cell(1, 1).alignment = center
    ws.row_dimensions[1].height = 28

    sous_titre = []
    if tour_date is not None:
        sous_titre.append(f"Date : {tour_date.strftime('%d/%m/%Y')}")
    if driver_name:
        sous_titre.append(f"Chauffeur : {driver_name}")
    sous_titre.append("Seule la colonne « Adresse » est obligatoire.")
    ws.append([" · ".join(sous_titre)])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(STOPS_COLUMNS))
    ws.cell(2, 1).font = Font(italic=True, size=9, color="666666")
    ws.cell(2, 1).alignment = center
    ws.append([])

    ws.append(STOPS_COLUMNS)
    hr = ws.max_row
    for col, h in enumerate(STOPS_COLUMNS, 1):
        c = ws.cell(hr, col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = center; c.border = brd
    ws.row_dimensions[hr].height = 30

    # ── Lignes de données ──
    lignes_ecrites = 0
    if df_stops is not None and len(df_stops) > 0:
        src = df_stops.copy()
        for c in STOPS_COLUMNS:
            if c not in src.columns:
                src[c] = "" if c not in ("Quantité", "Durée (min)") else (1 if c == "Quantité" else 30)
        for _, row in src.iterrows():
            ws.append([row.get(c, "") if not pd.isna(row.get(c, "")) else ""
                       for c in STOPS_COLUMNS])
            lignes_ecrites += 1

    premiere_ligne = hr + 1
    derniere_ligne = hr + max(lignes_ecrites + nb_lignes_vides, nb_lignes_vides)

    for r in range(premiere_ligne, derniere_ligne + 1):
        for col in range(1, len(STOPS_COLUMNS) + 1):
            c = ws.cell(r, col)
            c.border = brd
            c.alignment = left if STOPS_COLUMNS[col - 1] in (
                "Nom du client", "Adresse", "Observations") else center
        ws.cell(r, STOPS_COLUMNS.index("Adresse") + 1).fill = oblig

    # ── Menus déroulants ──
    def _lettre(nom):
        return openpyxl.utils.get_column_letter(STOPS_COLUMNS.index(nom) + 1)

    validations = [
        ("Action",  f"Listes!$A$2:$A${len(ACTIONS_CATALOGUE) + 1}"),
        ("Produit", f"Listes!$B$2:$B${len(PRODUITS_CATALOGUE) + 1}"),
        ("Option",  f"Listes!$C$2:$C${len(OPTIONS_CATALOGUE) + 1}"),
        ("Couleur", f"Listes!$D$2:$D${len(COULEURS_PRODUITS) + 1}"),
    ]
    for nom, plage in validations:
        dv = DataValidation(type="list", formula1=plage, allow_blank=True,
                            showDropDown=False)
        dv.error = ("Valeur non autorisée. Sélectionnez une entrée dans la liste "
                    "déroulante.")
        dv.errorTitle = "Saisie invalide"
        ws.add_data_validation(dv)
        L = _lettre(nom)
        dv.add(f"{L}{premiere_ligne}:{L}{derniere_ligne}")

    for nom, mini, maxi in (("Quantité", 1, 20), ("Durée (min)", 1, 480)):
        dvn = DataValidation(type="whole", operator="between",
                             formula1=str(mini), formula2=str(maxi),
                             allow_blank=True)
        dvn.errorTitle = "Valeur hors limites"
        dvn.error = f"Saisissez un nombre entier entre {mini} et {maxi}."
        ws.add_data_validation(dvn)
        L = _lettre(nom)
        dvn.add(f"{L}{premiere_ligne}:{L}{derniere_ligne}")

    largeurs = {"Action": 15, "Produit": 16, "Option": 13, "Couleur": 11,
                "Quantité": 9, "Nom du client": 24, "Adresse": 46,
                "Durée (min)": 12, "Pas avant": 11, "Pas après": 11,
                "Observations": 34}
    for col, nom in enumerate(STOPS_COLUMNS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = largeurs[nom]
    ws.freeze_panes = ws.cell(premiere_ligne, 1)

    # ── Onglet « Notice » ──
    wn = wb.create_sheet("Notice")
    wn.append(["NOTICE DE SAISIE – TOURNÉE"])
    wn.merge_cells("A1:C1")
    wn.cell(1, 1).font = Font(bold=True, size=14, color="1F4E79")
    wn.append([])
    notice = [
        ("Adresse", "OBLIGATOIRE",
         "Adresse complète avec code postal et ville. Ex : Place d'Hautpoul 81600 Gaillac. "
         "Une ligne sans adresse est ignorée à l'import."),
        ("Action", "Liste déroulante",
         "Nettoyer · Déposer · Retirer · Chargement · Déchargement. Par défaut : Nettoyer."),
        ("Produit", "Liste déroulante",
         "WC chimique · WC handicapé · WC Luxe · Lave-main · Urinoir · WC client. "
         "Par défaut : WC chimique."),
        ("Option", "Liste déroulante",
         "Lave-main ou Urinoir. Ne s'applique qu'au WC chimique : la valeur est "
         "automatiquement vidée pour les autres produits."),
        ("Couleur", "Facultatif",
         "À renseigner uniquement si le client exige une couleur précise. "
         "Laissée vide, la disponibilité est contrôlée toutes couleurs confondues."),
        ("Quantité", "Entier 1 à 20", "Nombre d'unités concernées. Par défaut : 1."),
        ("Nom du client", "Facultatif", "Nom du client ou du site d'intervention."),
        ("Durée (min)", "Entier 1 à 480",
         "Durée d'intervention sur place. Par défaut : 30. Le bouton « Précalculer "
         "les durées » de l'application recalcule cette colonne au barème."),
        ("Pas avant", "Format HH:MM",
         "Heure d'arrivée au plus tôt. Ex : 09:00. Laisser vide si aucune contrainte."),
        ("Pas après", "Format HH:MM",
         "Heure d'arrivée au plus tard. Ex : 11:30. Laisser vide si aucune contrainte."),
        ("Observations", "Facultatif",
         "Code portail, contact sur place, consignes particulières…"),
    ]
    wn.append(["Colonne", "Statut", "Explication"])
    for col in range(1, 4):
        c = wn.cell(wn.max_row, col)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = brd
    for nom, statut, expl in notice:
        wn.append([nom, statut, expl])
        r = wn.max_row
        for col in range(1, 4):
            wn.cell(r, col).border = brd
            wn.cell(r, col).alignment = left
        wn.cell(r, 1).font = Font(bold=True)
        if statut == "OBLIGATOIRE":
            wn.cell(r, 2).fill = oblig
            wn.cell(r, 2).font = Font(bold=True, color="B26A00")
    wn.append([])
    wn.append(["EXEMPLE"])
    wn.cell(wn.max_row, 1).font = Font(bold=True, size=12, color="1F4E79")
    wn.append(["Déposer | WC chimique | Lave-main | Vert | 2 | Mairie de Gaillac | "
               "Place d'Hautpoul 81600 Gaillac | 40 | 09:00 | (vide) | Code portail 1234"])
    wn.merge_cells(start_row=wn.max_row, start_column=1,
                   end_row=wn.max_row, end_column=3)
    wn.cell(wn.max_row, 1).alignment = left
    wn.cell(wn.max_row, 1).font = Font(italic=True, size=9, color="555555")
    wn.append([])
    wn.append(["L'ordre des lignes n'a aucune importance : l'application recalcule "
               "l'itinéraire optimal."])
    wn.merge_cells(start_row=wn.max_row, start_column=1,
                   end_row=wn.max_row, end_column=3)
    wn.cell(wn.max_row, 1).font = Font(italic=True, size=9, color="555555")
    for col, w in zip(range(1, 4), [20, 18, 88]):
        wn.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_stops_excel(uploaded_file):
    """Lit un classeur de tournée et retourne (DataFrame | None, message, avertissements).

    Détection tolérante de l'onglet, de la ligne d'en-tête et des colonnes.
    Les valeurs non reconnues sont ramenées à une valeur par défaut valide et
    signalées dans la liste d'avertissements plutôt que de bloquer l'import.
    """
    avertissements = []
    try:
        xls = pd.ExcelFile(uploaded_file)
    except Exception as e:
        return None, f"Fichier illisible : {e}", avertissements

    def _role(valeur):
        # Un libellé de colonne est court. Ce garde-fou évite qu'une phrase
        # d'introduction contenant un mot-clé soit prise pour un en-tête.
        brut = str(valeur if valeur is not None else "").strip()
        if len(brut) > 40:
            return None
        n = _norm_txt(brut)
        if not n:
            return None
        if "adresse" in n or n in ("lieu", "site"):
            return "Adresse"
        if "action" in n or "prestation" in n or "intervention" in n:
            return "Action"
        if "option" in n:
            return "Option"
        if "couleur" in n or "coloris" in n:
            return "Couleur"
        if "produit" in n or "materiel" in n or "equipement" in n or "article" in n:
            return "Produit"
        if n.startswith("qte") or "quantite" in n or n == "nb" or "nombre" in n:
            return "Quantité"
        if "client" in n or "nom" in n or "societe" in n:
            return "Nom du client"
        if "duree" in n or "temps" in n:
            return "Durée (min)"
        if "avant" in n or "au plus tot" in n or "ouverture" in n:
            return "Pas avant"
        if "apres" in n or "au plus tard" in n or "fermeture" in n:
            return "Pas après"
        if "observation" in n or "remarque" in n or "commentaire" in n or "note" in n:
            return "Observations"
        return None

    # Onglet prioritaire : celui dont le nom évoque une tournée
    ordre_onglets = sorted(
        xls.sheet_names,
        key=lambda s: 0 if ("tourn" in _norm_txt(s) or "arret" in _norm_txt(s)
                            or "saisie" in _norm_txt(s)) else 1)

    trouve = None
    for sheet in ordre_onglets:
        if _norm_txt(sheet) in ("listes", "notice"):
            continue
        try:
            brut = xls.parse(sheet, header=None)
        except Exception:
            continue
        if brut is None or len(brut) == 0:
            continue
        # On retient la ligne qui identifie le plus de colonnes, et non la
        # première correspondance : un titre ou une consigne peut contenir
        # fortuitement un mot-clé.
        meilleur = None
        for i in range(min(15, len(brut))):
            candidat = {}
            for j, val in enumerate(brut.iloc[i]):
                r = _role(val)
                if r and r not in candidat:
                    candidat[r] = j
            if "Adresse" in candidat and (meilleur is None
                                          or len(candidat) > len(meilleur[1])):
                meilleur = (i, candidat)
        if meilleur:
            trouve = (sheet, brut, meilleur[0], meilleur[1])
            break

    if not trouve:
        return None, (
            "Colonne « Adresse » introuvable. Le classeur doit comporter une ligne "
            "d'en-tête contenant au minimum une colonne « Adresse ». "
            "Utilisez le modèle vierge téléchargeable pour garantir la compatibilité."
        ), avertissements

    sheet, brut, header_row, mapping = trouve
    corps = brut.iloc[header_row + 1:]

    def _cell(row, nom, defaut=""):
        if nom not in mapping:
            return defaut
        v = row.iloc[mapping[nom]]
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return defaut
        except (TypeError, ValueError):
            pass
        return v

    def _entier(v, defaut, mini, maxi):
        try:
            n = int(round(float(str(v).replace(",", ".").strip())))
        except (ValueError, TypeError):
            return defaut
        return max(mini, min(maxi, n))

    lignes = []
    act_inconnues, prod_inconnus, opt_inconnues, heures_ko = set(), set(), set(), set()

    for _, row in corps.iterrows():
        adresse = str(_cell(row, "Adresse", "")).strip()
        if not adresse or _norm_txt(adresse) in ("nan", "total", "-", "—"):
            continue

        action_brut = _cell(row, "Action", "")
        action, ok_a = normalize_action(action_brut)
        if not ok_a and str(action_brut).strip():
            act_inconnues.add(str(action_brut).strip())

        produit_brut = _cell(row, "Produit", "")
        produit, ok_p = normalize_produit(produit_brut)
        if not ok_p and str(produit_brut).strip():
            prod_inconnus.add(str(produit_brut).strip())

        option_brut = _cell(row, "Option", "")
        option, ok_o = normalize_option_produit(option_brut)
        if not ok_o and str(option_brut).strip():
            opt_inconnues.add(str(option_brut).strip())

        couleur = normalize_couleur(_cell(row, "Couleur", ""))
        if couleur and couleur not in COULEURS_PRODUITS:
            couleur = ""

        pav_brut, pap_brut = _cell(row, "Pas avant", ""), _cell(row, "Pas après", "")
        pav, pap = _hhmm_from_cell(pav_brut), _hhmm_from_cell(pap_brut)
        for brut_v, converti in ((pav_brut, pav), (pap_brut, pap)):
            if str(brut_v).strip() and not converti and _norm_txt(brut_v) != "nan":
                heures_ko.add(str(brut_v).strip())

        lignes.append({
            "Suppr":         False,
            "Action":        action,
            "Produit":       produit,
            "Option":        option,
            "Couleur":       couleur,
            "Quantité":      _entier(_cell(row, "Quantité", 1), 1, 1, 20),
            "Nom du client": str(_cell(row, "Nom du client", "")).strip(),
            "Adresse":       adresse,
            "Durée (min)":   _entier(_cell(row, "Durée (min)", 30), 30, 1, 480),
            "Pas avant":     pav,
            "Pas après":     pap,
            "Observations":  str(_cell(row, "Observations", "")).strip(),
        })

    if not lignes:
        return None, (f"Aucun arrêt exploitable dans l'onglet « {sheet} » : "
                      f"toutes les lignes sont sans adresse."), avertissements

    df = pd.DataFrame(lignes, columns=["Suppr"] + STOPS_COLUMNS)
    df = _normalize_option(df)

    manquantes = [c for c in STOPS_COLUMNS if c not in mapping]
    if manquantes:
        avertissements.append(
            "Colonnes absentes du fichier, valeurs par défaut appliquées : "
            + ", ".join(manquantes))
    if act_inconnues:
        avertissements.append(
            "Action non reconnue (remplacée par « Nettoyer ») : "
            + ", ".join(sorted(act_inconnues)))
    if prod_inconnus:
        avertissements.append(
            "Produit non reconnu (remplacé par « WC chimique ») : "
            + ", ".join(sorted(prod_inconnus)))
    if opt_inconnues:
        avertissements.append(
            "Option non reconnue (vidée) : " + ", ".join(sorted(opt_inconnues)))
    if heures_ko:
        avertissements.append(
            "Horaire illisible (ignoré, format attendu HH:MM) : "
            + ", ".join(sorted(heures_ko)))

    return df, f"{len(df)} arrêt(s) importé(s) depuis l'onglet « {sheet} ».", avertissements


# ─────────────────────────────────────────────────────────────────────────────
# NOMENCLATURE DES PIÈCES À CONTRÔLER (état des lieux WC chimique)
# Regroupées par zone d'inspection physique pour un contrôle méthodique.
# Adaptez librement les libellés / l'ordre / les zones selon votre catalogue.
# ─────────────────────────────────────────────────────────────────────────────

PIECES_ETAT_LIEUX = [
    ("Porte & fermeture", [
        ("PR034", "Poignée de porte (couleur noire)"),
        ("PR036", "Support du verrou"),
    ]),
    ("Cuvette & sanitaire", [
        ("CO002", "Lunette + cuvette des WC (couleur noire)"),
        ("PR008", "Cylindre porte-papier hygiénique"),
    ]),
    ("Lave-mains & robinetterie", [
        ("CO013", "Distributeur de savon"),
        ("CO090", "Coussin de la pompe à pied"),
        ("PR023", "Embout du robinet"),
        ("PR042", "Mousseur pour robinet"),
        ("PR041", "Embout et filtre pour mousseur"),
        ("PR043", "Joint robinet"),
    ]),
    ("Structure & mobilité", [
        ("CO001", "Roue 200 DBU N/G 20x58"),
    ]),
]

ETAT_LIEUX_INTRO = (
    "Fiche à compléter par l'opérateur lors du contrôle de l'équipement. "
    "Pour chaque pièce, cochez « Bon état » si elle est présente et fonctionnelle. "
    "En cas de pièce usée, défectueuse ou manquante, cochez « À remplacer », "
    "indiquez la quantité à commander et précisez la nature du défaut dans la "
    "colonne Observations. La fiche datée et signée atteste de l'état constaté "
    "au moment de l'intervention."
)

ETAT_LIEUX_INTRO_GENERAL = (
    "Fiche unique de contrôle des pièces pour l'ensemble de la tournée. "
    "À compléter par l'opérateur : cochez « Bon état » pour chaque pièce présente "
    "et fonctionnelle, « À remplacer » en cas d'usure, de défaut ou d'absence, en "
    "indiquant la quantité à commander. Précisez le point d'intervention concerné "
    "dans la colonne Observations si nécessaire. La fiche datée et signée atteste "
    "de l'état constaté lors de la tournée."
)


def _normalize_option(df):
    """Vide la colonne « Option » pour tout produit autre que « WC chimique ».

    L'option (Urinoir / Lave-main) n'a de sens que pour le WC chimique. Cette
    normalisation garantit qu'un Urinoir ou un Lave-main (ou un WC handicapé)
    n'a jamais d'option renseignée, dans l'affichage comme dans les exports.
    """
    if "Produit" in df.columns and "Option" in df.columns:
        mask = df["Produit"].astype(str).str.strip() != "WC chimique"
        df.loc[mask, "Option"] = ""
    return df


def _etat_lieux_prestation_parts(stop):
    """Retourne (produit_label, option_label) pour l'en-tête d'une fiche.

    Ex. WC chimique + option Lave-main, qté 2 → ('2 × WC chimique', '2 × Lave-main').
    Sans option → ('1 × Urinoir', '—').
    """
    try:
        q = int(stop.get("qty_num", 1) or 1)
    except (ValueError, TypeError):
        q = 1
    produit = str(stop.get("produit", "") or "").strip()
    option  = str(stop.get("option", "") or "").strip()
    couleur = str(stop.get("couleur", "") or "").strip()
    produit_label = f"{q} × {produit}" if produit else "—"
    if produit and couleur:
        produit_label += f"  (couleur : {couleur})"
    option_label  = f"{q} × {option}"  if option  else "—"
    return produit_label, option_label


# ─────────────────────────────────────────────────────────────────────────────
# CASES À COCHER DESSINÉES (PDF)
# Le caractère Unicode « ☐ » n'existe pas dans la police Helvetica de reportlab :
# il est rendu comme un carré plein noir (glyphe .notdef). On dessine donc de
# vraies cases vides (contour seul) via un mini-tableau, ce qui garantit un rendu
# correct quelle que soit la police.
# ─────────────────────────────────────────────────────────────────────────────

def _pdf_checkbox(size=3.4*mm):
    """Retourne une case à cocher vide (contour seul) sous forme de flowable."""
    cb = Table([[""]], colWidths=[size], rowHeights=[size])
    cb.setStyle(TableStyle([
        ("BOX",           (0, 0), (-1, -1), 0.6, colors.HexColor("#555555")),
        ("BACKGROUND",    (0, 0), (-1, -1), colors.white),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
    ]))
    return cb


_OPT_LABEL_STYLE = ParagraphStyle("opt_label", fontName="Helvetica",
                                  fontSize=9, leading=11)


def _options_inline(options):
    """Ligne d'options « [ ] libellé » avec cases dessinées (contour seul).

    options : liste de tuples (libellé, largeur_mm_du_libellé).
    Retourne un mini-tableau à insérer comme contenu de cellule.
    """
    cells, widths = [], []
    for label, w in options:
        cells.append(_pdf_checkbox())
        widths.append(5*mm)
        cells.append(Paragraph(label, _OPT_LABEL_STYLE))
        widths.append(w*mm)
    t = Table([cells], colWidths=widths)
    t.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 1),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 2),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return t

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG PAGE
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Optimiseur de Tournées", page_icon="",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
    [data-testid="stMetric"] { background:#f8f9fa; border-radius:10px; padding:12px; }
    [data-testid="stMetricValue"] { font-size:1.6rem !important; color:#1f4e79; }
    .stop-card { border-radius:8px; padding:8px 12px; margin:4px 0; }
    .depot-card { background:#fff3cd; border-left:4px solid #ffc107; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────

def _init_df():
    return pd.DataFrame({
        "Suppr":         [False,         False,         False],
        "Action":        ["Nettoyer",    "Nettoyer",    "Nettoyer"],
        "Produit":       ["WC chimique", "WC chimique", "WC chimique"],
        "Option":        ["Lave-main",   "Lave-main",   "Lave-main"],
        "Couleur":       ["",            "",            ""],
        "Quantité":      [1,             1,             1],
        "Nom du client": ["",            "",            ""],
        "Adresse":       ["",            "",            ""],
        "Durée (min)":   [30,            30,            30],
        "Pas avant":     ["",            "",            ""],
        "Pas après":     ["",            "",            ""],
        "Observations":  ["",            "",            ""],
    })

if "df_stops"          not in st.session_state: st.session_state.df_stops          = _init_df()
if "heure_min_depart"  not in st.session_state: st.session_state.heure_min_depart  = datetime.time(7, 0)
if "result"            not in st.session_state: st.session_state.result            = None
if "tour_date"         not in st.session_state: st.session_state.tour_date         = datetime.date.today()
if "driver"            not in st.session_state: st.session_state.driver            = ""
if "depot_depart_key" not in st.session_state: st.session_state.depot_depart_key = DEPOT_DEPART_DEFAULT
if "depot_retour_key" not in st.session_state: st.session_state.depot_retour_key = DEPOT_RETOUR_DEFAULT
if "eviter_peri"       not in st.session_state: st.session_state.eviter_peri       = False
if "pause_dejeuner"    not in st.session_state: st.session_state.pause_dejeuner    = True
if "etat_lieux_par_arret" not in st.session_state: st.session_state.etat_lieux_par_arret = True
if "manual_order"      not in st.session_state: st.session_state.manual_order      = None
if "manual_result"     not in st.session_state: st.session_state.manual_result     = None
if "manual_sort_nonce" not in st.session_state: st.session_state.manual_sort_nonce = 0


# ─────────────────────────────────────────────────────────────────────────────
# RÉSULTAT ACTIF — source unique de vérité pour l'affichage ET les exports
# ─────────────────────────────────────────────────────────────────────────────
# Règle métier : dès que l'utilisateur a réorganisé manuellement la tournée dans
# l'onglet « Tournée optimisée », c'est CET ordre qui fait foi partout (carte,
# cartes d'arrêts, récapitulatif latéral, Excel, PDF, projection de stock).
# Toute lecture directe de st.session_state.result à des fins d'affichage ou
# d'export est un bug : passer impérativement par active_result().

def active_result():
    """Retourne le résultat de tournée qui fait foi.

    Priorité : ordre manuel de l'utilisateur > ordre calculé par l'optimiseur.
    Retourne None si aucune tournée n'a encore été optimisée.

    Le dict manuel est produit par _recalc_manual_route() sous la forme
    ``{**result, ...}`` : il expose donc exactement les mêmes clés que le
    résultat d'optimisation et reste un drop-in pour export_excel/export_pdf.
    """
    manual = st.session_state.get("manual_result")
    if manual is not None and manual.get("stops_ordered") is not None:
        return manual
    return st.session_state.get("result")


def is_manual_order() -> bool:
    """True si la tournée affichée/exportée résulte d'une réorganisation manuelle."""
    manual = st.session_state.get("manual_result")
    return manual is not None and manual.get("stops_ordered") is not None


def reset_manual_order() -> None:
    """Annule toute personnalisation manuelle (retour à l'ordre optimisé).

    À appeler dès qu'une nouvelle optimisation est lancée ou que les données
    d'entrée changent : conserver un ordre manuel obsolète exporterait des
    arrêts qui n'existent plus.
    """
    st.session_state.manual_order  = None
    st.session_state.manual_result = None
    # streamlit-sortables conserve l'ordre glissé dans l'état du widget tant que
    # la clé ne change pas : sans ce nonce, le bouton « Ordre optimisé » et une
    # nouvelle optimisation seraient annulés au rerun suivant par l'ancien ordre
    # renvoyé par le composant. Changer la clé force un remontage propre.
    st.session_state.manual_sort_nonce = st.session_state.get("manual_sort_nonce", 0) + 1

# ── Parc & stocks ────────────────────────────────────────────────────────────
# df_stock          : état courant du parc (Article / Couleur / Installés / En stock)
# stock_source      : origine des données affichée à l'utilisateur
# stock_import_ok   : True si un fichier Excel a été importé dans cette session
# stock_manutentions: prise en compte de Chargement/Déchargement dans les mouvements
if "df_stock"           not in st.session_state: st.session_state.df_stock           = _stock_default_df()
if "stock_source"       not in st.session_state: st.session_state.stock_source       = "Valeurs de référence (relevé exploitation)"
if "stock_import_ok"    not in st.session_state: st.session_state.stock_import_ok    = False
if "stock_import_name"  not in st.session_state: st.session_state.stock_import_name  = None
if "stock_manutentions" not in st.session_state: st.session_state.stock_manutentions = False
if "stock_seuil_alerte" not in st.session_state: st.session_state.stock_seuil_alerte = 1
if "stock_last_file_id" not in st.session_state: st.session_state.stock_last_file_id = None

# ── Import de tournée ────────────────────────────────────────────────────────
if "stops_last_file_id" not in st.session_state: st.session_state.stops_last_file_id = None
if "stops_import_name"  not in st.session_state: st.session_state.stops_import_name  = None
if "stops_import_msg"   not in st.session_state: st.session_state.stops_import_msg   = None
if "stops_import_warn"  not in st.session_state: st.session_state.stops_import_warn  = []

# ─────────────────────────────────────────────────────────────────────────────
# GEOCODAGE – API Adresse gouv.fr + fallback Nominatim
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def geocode(address: str):
    """Retourne (lat, lon) ou None."""
    try:
        r = requests.get("https://api-adresse.data.gouv.fr/search/",
                         params={"q": address, "limit": 1}, timeout=10)
        features = r.json().get("features", [])
        if features:
            lon, lat = features[0]["geometry"]["coordinates"]
            return (lat, lon)
    except Exception:
        pass
    try:
        geo = Nominatim(user_agent="tournee_optimizer_v4")
        loc = geo.geocode(address + ", France", timeout=10)
        if loc:
            return (loc.latitude, loc.longitude)
        time.sleep(1.1)
    except Exception:
        pass
    return None

# ─────────────────────────────────────────────────────────────────────────────
# ROUTAGE OSRM
# ─────────────────────────────────────────────────────────────────────────────

def _parse_hhmm(s):
    """Convertit 'HH:MM' en minutes depuis minuit, ou None si vide/invalide."""
    if not s or not str(s).strip():
        return None
    s = str(s).strip()
    try:
        h, m = s.split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return None


def _fmt_min(minutes):
    """Convertit des minutes depuis minuit en chaîne HH:MM."""
    if minutes is None:
        return ""
    h = int(minutes) // 60
    m = int(minutes) % 60
    return f"{h:02d}:{m:02d}"



def _compute_optimal_departure(tour, matrix, time_windows):
    """
    Calcule l'heure de départ optimale du dépôt.

    Logique :
    - On calcule pour chaque arrêt le temps cumulé depuis le dépôt
      (trajet + durées d'intervention des arrêts précédents).
    - Pour les contraintes "Pas après" (latest) : le départ ne peut pas
      dépasser latest_i - cumul_i  → on prend le minimum de ces valeurs.
    - Pour les contraintes "Pas avant" (earliest) : on peut partir plus tôt
      pour éviter l'attente → on prend le minimum entre le départ calculé
      depuis "Pas après" et earliest_i - cumul_i.
    - Si aucune contrainte : retourne None (pas de contrainte de départ).
    """
    n = len(tour)
    # Calcul des temps cumulés (trajet pur + interventions) depuis le dépôt
    cumul = 0.0
    cumuls = []          # cumuls[k] = temps cumulé pour atteindre tour[k] (k>=1)
    for k in range(1, n):
        prev = tour[k - 1]
        curr = tour[k]
        cumul += (matrix[prev][curr] or 0) / 60
        cumuls.append(cumul)
        # Ajouter la durée d'intervention de l'arrêt courant (sauf le dernier retour)
        dur = time_windows[curr].get("duration", 0) or 0
        cumul += dur

    candidates = []
    for k, idx in enumerate(tour[1:]):
        tw = time_windows[idx]
        cumul_k = cumuls[k]
        if tw.get("latest") is not None:
            # Dernier départ pour arriver à temps : latest - cumul
            candidates.append(tw["latest"] - cumul_k)
        if tw.get("earliest") is not None:
            # Départ idéal pour arriver exactement à l'ouverture (sans attente)
            candidates.append(tw["earliest"] - cumul_k)

    if not candidates:
        return None  # Aucune contrainte

    # Départ le plus tardif qui respecte toutes les contraintes
    optimal = min(candidates)
    # Arrondir à la minute inférieure
    return max(0, int(optimal))


import re as _re

def _parse_addr_ville_rue(address: str):
    """
    Extrait (ville, rue) depuis une adresse française.
    Format typique : "Numéro Rue, CP Ville" ou "Rue, CP Ville".
    Retourne (ville, rue) sous forme de chaînes.
    """
    addr = str(address).strip()
    # Chercher le code postal (5 chiffres)
    m = _re.search(r'\b(\d{5})\s+(.+?)(?:,|$)', addr)
    if m:
        ville = m.group(2).strip().rstrip(',').strip()
        rue   = addr[:m.start()].strip().rstrip(',').strip()
    else:
        parts = [p.strip() for p in addr.split(',')]
        ville = parts[-1] if parts else addr
        rue   = ', '.join(parts[:-1]) if len(parts) > 1 else ''
    return ville, rue


def _sortable_label(s: dict, pos: int, disp_stop: dict) -> str:
    """
    Construit le label affiché dans le drag-and-drop 'Ordre des arrêts'.
    Format : [orig_idx] Ville · Rue · Client · HH:MM (⚠️)
    """
    orig_idx = pos   # passé explicitement par l'appelant
    ville, rue       = _parse_addr_ville_rue(s.get("address", ""))
    client           = str(s.get("client") or "").strip()
    arr_str          = _fmt_min(disp_stop.get("arrival_min")) or ""
    violated_flag    = " ⚠️" if disp_stop.get("violated") else ""
    arr_part         = f" · {arr_str}" if arr_str else ""
    # Assemblage Ville + Rue + Client (on exclut les parties vides)
    name_parts = [p for p in [ville[:28], rue[:28], client[:22]] if p]
    display_name = " · ".join(name_parts) if name_parts else s.get("address", "")[:40]
    return f"[{orig_idx}] {display_name}{arr_part}{violated_flag}"


def _find_pause_position(stops_ordered: list) -> int:
    """
    Retourne l'indice (0-based) de l'arrêt APRÈS lequel placer la pause déjeuner.
    Heuristique : premier arrêt dont le départ estimé dépasse 12h00 (720 min).
    Si tous les arrêts précèdent midi, on insère au milieu de la tournée.
    """
    NOON = 720  # 12:00
    for i, s in enumerate(stops_ordered):
        dep = s.get("departure_min")
        if dep is not None and dep >= NOON:
            return max(0, i)       # juste avant cet arrêt (après le précédent)
    return max(0, len(stops_ordered) // 2)


def _pause_slot(stops_ordered: list, pause_idx: int) -> tuple[str, str]:
    """
    Retourne (heure_debut, heure_fin) de la pause déjeuner au format HH:MM.
    La pause commence au départ du stop[pause_idx].
    """
    s = stops_ordered[pause_idx] if 0 <= pause_idx < len(stops_ordered) else {}
    start = s.get("departure_min") or s.get("arrival_min")
    if start is None:
        return ("", "")
    end = start + PAUSE_DEJEUNER_MIN
    return (_fmt_min(start), _fmt_min(end))


def _qty_label(row):
    """
    Construit la description lisible de la commande à partir des colonnes
    Produit / Option / Quantité.
    Ex : "3 × WC chimique + 3 × Urinoir"
         "2 × Lave-main"
    """
    produit = str(row.get("Produit", "") or "").strip()
    option  = str(row.get("Option",  "") or "").strip()
    couleur = str(row.get("Couleur", "") or "").strip()
    try:
        qty = int(row.get("Quantité", 1) or 1)
    except (ValueError, TypeError):
        qty = 1

    if not produit:
        return f"{qty} × ?"

    label = f"{qty} × {produit}"
    # Option uniquement pour WC chimique et si renseignée
    if produit == "WC chimique" and option:
        label += f" + {qty} × {option}"
    # Couleur exigée par le client (critère facultatif)
    if couleur:
        label += f" — couleur {couleur}"
    return label

def _compute_arrivals(tour, matrix, depart_min, time_windows, pause_after_idx=None):
    """
    Calcule l'heure d'arrivée réelle à chaque arrêt (en minutes depuis minuit),
    en tenant compte des fenêtres temporelles et des durées d'intervention.
    time_windows[i] contient aussi "duration" (durée intervention en minutes).

    pause_after_idx : indice 0-based dans la liste des arrêts (hors dépôt) après
    lequel injecter la pause déjeuner (PAUSE_DEJEUNER_MIN). Quand ce paramètre est
    fourni, tous les arrêts situés APRÈS la pause voient leur heure décalée d'autant,
    ce qui garantit la cohérence complète du planning.
    Si pause_after_idx vaut None, aucune pause n'est injectée.

    Retourne une liste de dicts par arrêt (hors dépôt) :
      arrival_min, wait_min, departure_min, tw_early, tw_late, violated, duration_min
    """
    results      = []
    current_time = depart_min   # minutes depuis minuit
    current_idx  = tour[0]      # = 0 (dépôt)
    stop_count   = 0            # indice 0-based de l'arrêt courant dans results

    for step, next_idx in enumerate(tour[1:], 1):
        travel_min    = (matrix[current_idx][next_idx] or 0) / 60
        arrival_min   = current_time + travel_min
        tw            = time_windows[next_idx]
        earliest      = tw["earliest"]
        latest        = tw["latest"]
        duration_min  = tw.get("duration", 0) or 0
        wait_min      = 0

        # Si on arrive trop tôt → on attend jusqu'à l'ouverture
        if earliest is not None and arrival_min < earliest:
            wait_min    = earliest - arrival_min
            arrival_min = earliest

        # Violation : arrive après la limite latest
        violated = (latest is not None and arrival_min > latest)

        # Départ de l'arrêt = arrivée + durée intervention
        departure_min = arrival_min + duration_min

        results.append({
            "arrival_min":   arrival_min,
            "wait_min":      wait_min,
            "departure_min": departure_min,
            "tw_early":      earliest,
            "tw_late":       latest,
            "violated":      violated,
            "duration_min":  duration_min,
        })

        current_time = departure_min   # on repart après l'intervention

        # ── Pause déjeuner : on l'injecte APRÈS l'arrêt indiqué ──
        # Le chauffeur ne repart vers l'arrêt suivant qu'après la fin de la pause.
        if pause_after_idx is not None and stop_count == pause_after_idx:
            current_time += PAUSE_DEJEUNER_MIN

        stop_count  += 1
        current_idx  = next_idx

    return results


def _tour_cost_tw(tour, matrix, depart_min, time_windows):
    """
    Coût total tenant compte des fenêtres temporelles.
    Temps de trajet + attentes + pénalité forte pour violations.
    """
    PENALTY = 1e6   # pénalité par minute de dépassement
    arrivals = _compute_arrivals(tour, matrix, depart_min, time_windows)
    # Durée totale jusqu'au dernier arrêt + retour dépôt
    if not arrivals:
        return 0
    last_departure = arrivals[-1]["departure_min"]
    back_min = (matrix[tour[-1]][tour[0]] or 0) / 60
    total_time = (last_departure + back_min) - depart_min

    # Pénalités pour violations
    penalty = 0
    for a in arrivals:
        if a["violated"] and a["tw_late"] is not None:
            penalty += (a["arrival_min"] - a["tw_late"]) * PENALTY

    return total_time + penalty


def _two_opt(tour, matrix, depart_min=None, time_windows=None):
    """Amelioration 2-opt avec prise en compte optionnelle des fenêtres temporelles."""
    use_tw = depart_min is not None and time_windows is not None
    def cost(t):
        if use_tw:
            return _tour_cost_tw(t, matrix, depart_min, time_windows)
        total = sum(matrix[t[i]][t[i+1]] for i in range(len(t)-1))
        return total + matrix[t[-1]][t[0]]

    best     = tour[:]
    improved = True
    while improved:
        improved = False
        for i in range(1, len(best) - 1):
            for j in range(i + 1, len(best)):
                candidate = best[:i] + best[i:j+1][::-1] + best[j+1:]
                if cost(candidate) < cost(best):
                    best     = candidate
                    improved = True
    return best


# ─────────────────────────────────────────────────────────────────────────────
# HEURISTIQUE PÉRIPHÉRIQUE TOULOUSAIN
# ─────────────────────────────────────────────────────────────────────────────

# Centre géographique de Toulouse et rayon de la rocade
_TOULOUSE_CENTER  = (43.6047, 1.4442)
_ROCADE_RAYON_KM  = 10.0   # seuil : < 10 km du centre = intra-rocade


def _haversine_km(coord_a, coord_b):
    """Distance orthodromique en km entre deux points (lat, lon)."""
    import math
    lat1, lon1 = math.radians(coord_a[0]), math.radians(coord_a[1])
    lat2, lon2 = math.radians(coord_b[0]), math.radians(coord_b[1])
    dlat, dlon = lat2 - lat1, lon2 - lon1
    a = math.sin(dlat / 2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2)**2
    return 6371 * 2 * math.asin(math.sqrt(a))


def _est_intra_rocade(coord):
    """
    Retourne True si le point est à l'intérieur de la rocade toulousaine
    (distance au centre < _ROCADE_RAYON_KM).
    """
    return _haversine_km(coord, _TOULOUSE_CENTER) < _ROCADE_RAYON_KM


# Bornes horaires des heures de pointe (minutes depuis minuit)
_POINTE_MATIN_DEBUT  = 7  * 60        # 07:00
_POINTE_MATIN_FIN    = 9  * 60        # 09:00
_POINTE_SOIR_DEBUT   = 17 * 60        # 17:00
_POINTE_SOIR_FIN     = 19 * 60        # 19:00
_FIN_POINTE_MATIN    = _POINTE_MATIN_FIN        # les arrêts intra-rocade ne commencent pas avant
_DEBUT_POINTE_SOIR   = _POINTE_SOIR_DEBUT - 30  # les arrêts intra-rocade se terminent avant


def _injecte_contraintes_peri(time_windows, coords_list, depart_min):
    """
    Quand le mode périphérique est activé et qu'on est en heure de pointe,
    injecte des contraintes horaires automatiques sur les arrêts intra-rocade
    pour que le chauffeur les traite en dehors des bouchons.

    Logique :
    - Pointe matin (07h–09h) : arrêts intra-rocade → earliest = 09:00
      (le chauffeur fait d'abord les arrêts extérieurs, entre en ville après 9h)
    - Pointe soir  (17h–19h) : arrêts intra-rocade → latest  = 16:30
      (le chauffeur termine les arrêts en centre-ville avant les bouchons du soir)

    Les contraintes existantes saisies par l'utilisateur sont respectées :
    - earliest injecté = max(earliest_utilisateur, 09:00)
    - latest  injecté  = min(latest_utilisateur,   16:30)

    Retourne (time_windows_modifiées, liste_des_indices_arrêts_affectés).
    coords_list[0] = dépôt, coords_list[k] = arrêt k.
    """
    pointe_matin = _POINTE_MATIN_DEBUT <= depart_min <= _POINTE_MATIN_FIN
    pointe_soir  = _POINTE_SOIR_DEBUT  <= depart_min <= _POINTE_SOIR_FIN

    if not (pointe_matin or pointe_soir):
        return time_windows, []   # hors pointe : rien à injecter

    tw_modif  = [tw.copy() for tw in time_windows]   # copie pour ne pas altérer l'original
    affectes  = []

    for k, coord in enumerate(coords_list[1:], start=1):   # index 0 = dépôt, ignoré
        if not _est_intra_rocade(coord):
            continue   # arrêt hors rocade : pas de contrainte injectée

        tw = tw_modif[k]
        modifie = False

        if pointe_matin:
            # Pas avant 09:00 (ou la contrainte utilisateur si elle est plus tardive)
            seuil = _FIN_POINTE_MATIN
            if tw["earliest"] is None or tw["earliest"] < seuil:
                tw["earliest"] = seuil
                modifie = True

        if pointe_soir:
            # Pas après 16:30 (ou la contrainte utilisateur si elle est plus tôt)
            seuil = _DEBUT_POINTE_SOIR
            if tw["latest"] is None or tw["latest"] > seuil:
                tw["latest"] = seuil
                modifie = True

        if modifie:
            affectes.append(k)

    return tw_modif, affectes


def osrm_trip(coords_latlon, time_windows=None, depart_min=None):
    """
    Optimisation TSP depot fixe avec fenêtres temporelles optionnelles.
    1. Matrice de durees OSRM (/table)
    2. Nearest-neighbor depuis le depot (index 0)
    3. Amelioration 2-opt (avec TW si fournis)
    4. Route finale (/route) pour geometrie et distances reelles
    """
    n         = len(coords_latlon)
    coord_str = ";".join(f"{lon},{lat}" for lat, lon in coords_latlon)

    # Etape 1 : matrice de durees
    try:
        r    = requests.get(f"{OSRM_URL}/table/v1/driving/{coord_str}",
                            params={"annotations": "duration"}, timeout=30)
        data = r.json()
    except Exception as e:
        st.error(f"Erreur reseau OSRM table : {e}")
        return None
    if data.get("code") != "Ok":
        st.error(f"OSRM table error : {data.get('code')}")
        return None

    matrix      = data["durations"]   # liste n x n de durees en secondes
    matrix_real = matrix              # alias explicite — durées réelles OSRM pour les ETAs

    # Etape 3 : nearest-neighbor depuis le depot
    unvisited = list(range(1, n))
    tour      = [0]
    current   = 0
    while unvisited:
        nearest = min(unvisited, key=lambda j: matrix[current][j] or float("inf"))
        tour.append(nearest)
        unvisited.remove(nearest)
        current = nearest

    # Etape 4 : amelioration 2-opt (avec TW si fournis)
    tour = _two_opt(tour, matrix, depart_min=depart_min, time_windows=time_windows)

    # Etape 5 : route reelle pour geometrie et distances
    tour_with_return = tour + [tour[0]]
    route_coords     = [coords_latlon[i] for i in tour_with_return]
    route_coord_str  = ";".join(f"{lon},{lat}" for lat, lon in route_coords)

    try:
        r2    = requests.get(f"{OSRM_URL}/route/v1/driving/{route_coord_str}",
                             params={"overview": "full", "geometries": "geojson"},
                             timeout=20)
        rdata = r2.json()
    except Exception as e:
        st.error(f"Erreur reseau OSRM route : {e}")
        return None
    if rdata.get("code") != "Ok":
        st.error(f"OSRM route error : {rdata.get('code')}")
        return None

    route = rdata["routes"][0]
    geom  = [[lat, lon] for lon, lat in route["geometry"]["coordinates"]]

    return {
        "order":        tour,
        "distance_km":  route["distance"] / 1000,
        "duration_min": route["duration"] / 60,
        "geometry":     geom,
        "matrix":       matrix,       # = matrix_real (durées OSRM réelles)
        "matrix_real":  matrix_real,  # alias conservé pour la compatibilité
    }



def osrm_route_distance(coords_latlon):
    coords    = coords_latlon + [coords_latlon[0]]
    coord_str = ";".join(f"{lon},{lat}" for lat, lon in coords)
    try:
        r    = requests.get(f"{OSRM_URL}/route/v1/driving/{coord_str}",
                            params=dict(overview="false"), timeout=20)
        data = r.json()
    except Exception:
        return None
    if data.get("code") != "Ok":
        return None
    route = data["routes"][0]
    return {"distance_km":  route["distance"] / 1000,
            "duration_min": route["duration"] / 60}


def _recalc_manual_route(result, new_order):
    """
    Recalcule distance, durée, ETAs et géométrie pour un ordre manuel d'arrêts.

    Paramètres :
        result    : st.session_state.result (résultat de l'optimisation de référence)
        new_order : liste d'orig_idx (entiers 0-based) dans result["stops_ordered"]
                    ex. [2, 0, 1] = l'arrêt 2 en premier, puis 0, puis 1

    Retourne un dict au même format que result, avec les champs mis à jour.
    """
    stops              = result["stops_ordered"]
    stop_matrix        = result["stop_matrix"]        # (n+1)×(n+1), indice 0 = dépôt
    depart_min         = result["depart_min"]
    depot_coords       = result["depot_coords"]
    depot_retour_coords = result.get("depot_retour_coords") or result["depot_coords"]
    fuel_conso         = result.get("fuel_conso", 12.20)
    fuel_price_val     = result.get("fuel_price",  1.85)
    pause_enabled      = result.get("pause_dejeuner", True)

    n = len(new_order)

    # ── Fenêtres temporelles dans le nouvel ordre ──
    tw_new = [{"earliest": None, "latest": None, "duration": 0}]  # index 0 = dépôt
    for orig_idx in new_order:
        s = stops[orig_idx]
        tw_new.append({
            "earliest": s.get("tw_early"),
            "latest":   s.get("tw_late"),
            "duration": s.get("duration_min", 30) or 30,
        })

    # ── Sous-matrice pour _compute_arrivals ──
    # stop_matrix[0]       = dépôt
    # stop_matrix[i+1][j+1] = stops[i] → stops[j]
    matrix_order = [0] + [oi + 1 for oi in new_order]
    sub_matrix = [
        [(stop_matrix[matrix_order[i]][matrix_order[j]] or 0) for j in range(n + 1)]
        for i in range(n + 1)
    ]

    # ── ETAs ──
    # Passe 1 : sans pause → localise le passage de midi
    arrivals_prelim = _compute_arrivals(list(range(n + 1)), sub_matrix, depart_min, tw_new)
    pause_idx_manual = _find_pause_position(
        [{"departure_min": a.get("departure_min")} for a in arrivals_prelim])
    # Passe 2 : avec pause injectée (si activée) → ETAs cohérentes après la pause
    arrivals = _compute_arrivals(
        list(range(n + 1)), sub_matrix, depart_min, tw_new,
        pause_after_idx=(pause_idx_manual if pause_enabled else None))

    new_stops = []
    for rank, (orig_idx, arr) in enumerate(zip(new_order, arrivals), 1):
        s = stops[orig_idx].copy()
        s["order_num"]     = rank
        s["arrival_min"]   = arr.get("arrival_min")
        s["departure_min"] = arr.get("departure_min")
        s["wait_min"]      = arr.get("wait_min", 0)
        s["violated"]      = arr.get("violated", False)
        new_stops.append(s)

    # ── Géométrie + distance/durée réelles via OSRM /route ──
    route_coords = [depot_coords]
    for orig_idx in new_order:
        s = stops[orig_idx]
        if s.get("lat") is not None and s.get("lon") is not None:
            route_coords.append((s["lat"], s["lon"]))

    geometry = result["geometry"]
    dist_km  = result["distance_km"]
    dur_min  = result["duration_min"]
    # False si OSRM n'a pas répondu : distance/durée/géométrie restent celles de
    # l'ordre optimisé alors que l'ordre des arrêts, lui, est bien celui choisi.
    # On le signale à l'utilisateur plutôt que d'exporter un chiffre trompeur.
    recalc_ok = True

    if len(route_coords) > 1:
        _dest     = depot_retour_coords
        all_coords = route_coords + [_dest]
        coord_str  = ";".join(f"{lon},{lat}" for lat, lon in all_coords)
        try:
            resp  = requests.get(
                f"{OSRM_URL}/route/v1/driving/{coord_str}",
                params={"overview": "full", "geometries": "geojson"},
                timeout=20,
            )
            rdata = resp.json()
            if rdata.get("code") == "Ok":
                route   = rdata["routes"][0]
                dist_km = route["distance"] / 1000
                # durée trajet pur + pause éventuelle (cohérent avec l'optimisation)
                dur_min = route["duration"] / 60 + (PAUSE_DEJEUNER_MIN if pause_enabled else 0)
                geometry = [[lat, lon] for lon, lat in route["geometry"]["coordinates"]]
            else:
                recalc_ok = False
        except Exception:
            recalc_ok = False  # conserve geometry/dist/dur du résultat précédent

    fuel_l    = dist_km * fuel_conso / 100
    fuel_cost = fuel_l * fuel_price_val

    # ── Temps de retour au dépôt ──
    if new_stops:
        last_dep    = (new_stops[-1].get("departure_min")
                       or new_stops[-1].get("arrival_min")
                       or depart_min)
        last_mi     = new_order[-1] + 1   # indice dans stop_matrix
        travel_back = (stop_matrix[last_mi][0] or 0) / 60
        # Pause déjà balisée dans last_dep si elle précède le dernier arrêt
        if pause_enabled and pause_idx_manual >= len(new_stops) - 1:
            return_min = last_dep + PAUSE_DEJEUNER_MIN + travel_back
        else:
            return_min = last_dep + travel_back
    else:
        return_min = depart_min

    return {
        **result,                         # conserve tous les champs de référence
        "stops_ordered":  new_stops,
        "distance_km":    dist_km,
        "duration_min":   dur_min,
        "fuel_liters":    fuel_l,
        "fuel_cost":      fuel_cost,
        "geometry":       geometry,
        "return_min":     return_min,
        "is_manual":      True,
        # Ordre manuel effectivement appliqué (indices dans result["stops_ordered"])
        # → conservé pour audit / débogage et pour l'entête des exports.
        "manual_order":   list(new_order),
        "recalc_ok":      recalc_ok,
        # Les colonnes "Temps gagné" et "km sauvés" restent celles de l'optimisation
        "km_saved":       result.get("km_saved", 0),
        "time_saved_min": result.get("time_saved_min", 0),
    }


# ─────────────────────────────────────────────────────────────────────────────

def build_map(depot_coords, stops_ordered, geometry, depot_depart_addr="Dépôt départ", depot_retour_addr=None, depot_retour_coords=None):
    m = folium.Map(location=depot_coords, zoom_start=11, tiles="CartoDB positron")
    if geometry:
        folium.PolyLine(geometry, color="#1f4e79", weight=4, opacity=0.85).add_to(m)
    folium.Marker(depot_coords,
                  popup=folium.Popup(f"<b>Dépôt départ</b><br>{depot_depart_addr}", max_width=260),
                  tooltip="Dépôt – Départ",
                  icon=folium.Icon(color="orange", icon="home", prefix="fa")).add_to(m)
    # Marqueur retour dépôt si différent du départ
    if depot_retour_coords and depot_retour_coords != depot_coords:
        folium.Marker(depot_retour_coords,
                      popup=folium.Popup(f"<b>Dépôt retour</b><br>{depot_retour_addr}", max_width=260),
                      tooltip="Dépôt – Retour",
                      icon=folium.Icon(color="darkred", icon="flag", prefix="fa")).add_to(m)

    for stop in stops_ordered:
        if stop["lat"] is None:
            continue
        color = ACTION_MAP_COLORS.get(stop["action"], "gray")
        icon  = ACTION_MAP_ICONS.get(stop["action"], "map-marker")
        popup_html = (f"<b>Arrêt {stop['order_num']}</b><br>"
                      f"<b>Action :</b> {stop['action']}<br>"
                      f"<b>Qté :</b> {stop['quantity']}<br>"
                      + (f"<b>Client :</b> {stop['client']}<br>" if stop['client'] else "")
                      + f"<b>Adresse :</b> {stop['address']}")
        folium.Marker([stop["lat"], stop["lon"]],
                      popup=folium.Popup(popup_html, max_width=260),
                      tooltip=f"{stop['order_num']}. {stop['action']} – {stop['address']}",
                      icon=folium.Icon(color=color, icon=icon, prefix="fa")).add_to(m)
        bg = {"blue": "#337ab7", "green": "#28a745", "red": "#dc3545"}.get(color, "#555")
        folium.Marker([stop["lat"], stop["lon"]],
                      icon=folium.DivIcon(
                          html=(f'<div style="font-size:11px;font-weight:bold;color:#fff;'
                                f'background:{bg};border-radius:50%;width:20px;height:20px;'
                                f'display:flex;align-items:center;justify-content:center;'
                                f'margin-top:-42px;margin-left:14px;">{stop["order_num"]}</div>'),
                          icon_size=(20, 20), icon_anchor=(0, 0))).add_to(m)
    return m


def generate_map_image(depot_coords, stops_ordered, geometry, depot_retour_coords=None):
    """
    Génère une image PNG de la carte via Pillow (PIL).
    Disponible partout, aucune dépendance système supplémentaire.
    """
    from PIL import Image, ImageDraw

    W, H   = 900, 600
    MARGIN = 60
    BG     = (240, 239, 231)   # fond beige carte
    GRID   = (200, 200, 200)
    ROUTE  = (31, 78, 121)

    ACTION_RGB = {
        "Nettoyer":     (31,  106, 165),
        "Déposer":      (40,  167,  69),
        "Retirer":      (220,  53,  69),
        "Chargement":   (123,  31, 162),
        "Déchargement": (230,  81,   0),
    }

    img  = Image.new("RGB", (W, H), BG)
    draw = ImageDraw.Draw(img)

    # Points pour normalisation
    all_lats = [depot_coords[0]] + [s["lat"] for s in stops_ordered if s["lat"]]
    all_lons = [depot_coords[1]] + [s["lon"] for s in stops_ordered if s["lon"]]
    min_lat, max_lat = min(all_lats), max(all_lats)
    min_lon, max_lon = min(all_lons), max(all_lons)
    span_lat = max_lat - min_lat or 0.01
    span_lon = max_lon - min_lon or 0.01
    pad_lat, pad_lon = span_lat * 0.14, span_lon * 0.14
    min_lat -= pad_lat; max_lat += pad_lat
    min_lon -= pad_lon; max_lon += pad_lon
    span_lat = max_lat - min_lat
    span_lon = max_lon - min_lon

    def proj(lat, lon):
        x = int(MARGIN + (lon - min_lon) / span_lon * (W - 2 * MARGIN))
        y = int(H - MARGIN - (lat - min_lat) / span_lat * (H - 2 * MARGIN))
        return (x, y)

    # Grille
    for i in range(6):
        xi = MARGIN + int(i * (W - 2*MARGIN) / 5)
        yi = MARGIN + int(i * (H - 2*MARGIN) / 5)
        draw.line([(xi, MARGIN), (xi, H-MARGIN)], fill=GRID, width=1)
        draw.line([(MARGIN, yi), (W-MARGIN, yi)], fill=GRID, width=1)

    # Bordure
    draw.rectangle([MARGIN-1, MARGIN-1, W-MARGIN+1, H-MARGIN+1],
                   outline=(150, 150, 150), width=1)

    # Tracé route
    if geometry and len(geometry) > 1:
        pts = [proj(lat, lon) for lat, lon in geometry]
        for i in range(len(pts) - 1):
            # Ombre épaisse
            draw.line([pts[i], pts[i+1]], fill=(180, 200, 220), width=6)
        for i in range(len(pts) - 1):
            draw.line([pts[i], pts[i+1]], fill=ROUTE, width=3)

    # Dépôt
    dx, dy = proj(depot_coords[0], depot_coords[1])
    r = 12
    draw.ellipse([dx-r, dy-r, dx+r, dy+r], fill=(255, 193, 7), outline=(51, 51, 51), width=2)
    draw.text((dx, dy), "D", fill=(51, 51, 51), anchor="mm")
    draw.text((dx + r + 4, dy - 5), "Départ", fill=(51, 51, 51))

    # Dépôt retour (si différent du départ)
    if depot_retour_coords and depot_retour_coords != depot_coords:
        rx2, ry2 = proj(depot_retour_coords[0], depot_retour_coords[1])
        draw.ellipse([rx2-r, ry2-r, rx2+r, ry2+r], fill=(220, 53, 69), outline=(51, 51, 51), width=2)
        draw.text((rx2, ry2), "R", fill=(255, 255, 255), anchor="mm")
        draw.text((rx2 + r + 4, ry2 - 5), "Retour", fill=(51, 51, 51))

    # Arrêts
    for stop in stops_ordered:
        if stop["lat"] is None:
            continue
        sx, sy  = proj(stop["lat"], stop["lon"])
        rgb     = ACTION_RGB.get(stop["action"], (85, 85, 85))
        r_stop  = 11
        # Ombre
        draw.ellipse([sx-r_stop+1, sy-r_stop+1, sx+r_stop+1, sy+r_stop+1],
                     fill=(180, 180, 180))
        draw.ellipse([sx-r_stop, sy-r_stop, sx+r_stop, sy+r_stop],
                     fill=rgb, outline=(255, 255, 255), width=2)
        draw.text((sx, sy), str(stop["order_num"]), fill=(255, 255, 255), anchor="mm")
        # Étiquette
        label = stop["address"][:32] + ("…" if len(stop["address"]) > 32 else "")
        lx, ly = sx + r_stop + 4, sy - 7
        # Fond blanc semi-transparent
        bbox = draw.textbbox((lx, ly), label)
        draw.rectangle([bbox[0]-2, bbox[1]-1, bbox[2]+2, bbox[3]+1],
                       fill=(255, 255, 255, 200))
        draw.text((lx, ly), label, fill=(30, 30, 30))

    # Légende en bas
    lx, ly = MARGIN, H - MARGIN + 14
    draw.text((lx, ly), "●  Dépôt", fill=(51, 51, 51))
    lx += 80
    seen = []
    for stop in stops_ordered:
        a = stop["action"]
        if a not in seen:
            seen.append(a)
            rgb = ACTION_RGB.get(a, (85, 85, 85))
            draw.ellipse([lx, ly+2, lx+10, ly+12], fill=rgb)
            draw.text((lx+14, ly), a, fill=(51, 51, 51))
            lx += len(a) * 7 + 24

    # Titre
    title = "Tournée optimisée"
    tw = draw.textlength(title)
    draw.text(((W - tw) // 2, 14), title, fill=(31, 78, 121))

    buf = BytesIO()
    img.save(buf, format="PNG", optimize=True)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT — FEUILLE « ÉTAT DES LIEUX » (Excel + PDF)
# ─────────────────────────────────────────────────────────────────────────────

def _write_etat_lieux_block(ws, tour_date, driver_name,
                            client_name="", address="", action_label="",
                            produit_label="", option_label="", general=False,
                            nb_arrets=None, depot_addr=""):
    """Écrit UNE fiche 'état des lieux' à partir de la ligne courante de ws.

    Utilise systématiquement ws.append + ws.max_row (indices relatifs) afin que
    plusieurs fiches puissent être empilées dans la même feuille.
    Si general=True, écrit la fiche généraliste (en-tête tournée) au lieu d'une
    fiche par arrêt.
    Retourne l'indice de la dernière ligne écrite.
    """
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    zone_fill = PatternFill("solid", fgColor="DCE6F1")
    zone_font = Font(bold=True, color="1F4E79", size=10)
    bold_f    = Font(bold=True)
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",  vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="CCCCCC")
    brd       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Titre ──
    if general:
        titre = f"ÉTAT DES LIEUX – CONTRÔLE DES PIÈCES (FICHE TOURNÉE) – {tour_date.strftime('%d/%m/%Y')}"
    else:
        titre = f"ÉTAT DES LIEUX – CONTRÔLE DES PIÈCES – {tour_date.strftime('%d/%m/%Y')}"
    ws.append([titre])
    tr = ws.max_row
    ws.merge_cells(f"A{tr}:F{tr}")
    c = ws.cell(tr, 1)
    c.font      = Font(bold=True, size=15, color="1F4E79")
    c.alignment = center
    ws.row_dimensions[tr].height = 28

    # ── Bloc d'entête ──
    if general:
        nb_txt = (f"{nb_arrets} arrêt(s)" if nb_arrets is not None else "—")
        entete_rows = [
            ("Opérateur",                driver_name or "—"),
            ("Date de la tournée",       tour_date.strftime('%d/%m/%Y')),
            ("Dépôt de départ",          depot_addr or "—"),
            ("Points d'intervention",    nb_txt),
            ("Client / site",            "(à compléter sur le terrain)"),
        ]
    else:
        entete_rows = [
            ("Opérateur",                driver_name or "—"),
            ("Client / site",            client_name or "—"),
            ("Adresse d'intervention",   address or "—"),
            ("Nature de l'intervention", action_label or "—"),
            ("Produit",                  produit_label or "—"),
            ("Option",                   option_label or "—"),
        ]
    for label, val in entete_rows:
        ws.append([label + " :", val])
        r = ws.max_row
        ws.cell(r, 1).font      = bold_f
        ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"B{r}:F{r}")
        ws.cell(r, 2).alignment = left

    # ── Consigne ──
    ws.append([ETAT_LIEUX_INTRO_GENERAL if general else ETAT_LIEUX_INTRO])
    r = ws.max_row
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(r, 1)
    c.font      = Font(italic=True, size=9, color="555555")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 46
    ws.append([])

    # ── En-têtes du tableau ──
    headers = ["N° Article", "Désignation", "☑ Bon état", "☒ À remplacer",
               "Qté à commander", "Observations"]
    ws.append(headers)
    hr = ws.max_row
    for col, h in enumerate(headers, 1):
        c = ws.cell(hr, col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = center; c.border = brd
    ws.row_dimensions[hr].height = 22

    # ── Lignes par zone ──
    for zone, pieces in PIECES_ETAT_LIEUX:
        ws.append([zone, "", "", "", "", ""])
        zr = ws.max_row
        ws.merge_cells(f"A{zr}:F{zr}")
        zc = ws.cell(zr, 1)
        zc.fill = zone_fill; zc.font = zone_font
        zc.alignment = Alignment(horizontal="left", vertical="center")
        zc.border = brd
        ws.row_dimensions[zr].height = 18

        for ref, desig in pieces:
            ws.append([ref, desig, "☐", "☐", "", ""])
            pr = ws.max_row
            for col in range(1, 7):
                c = ws.cell(pr, col)
                c.border = brd
                if col in (3, 4):        # cases à cocher
                    c.alignment = center
                    c.font = Font(size=13)
                elif col == 5:           # quantité
                    c.alignment = center
                elif col == 2:           # désignation
                    c.alignment = left
                elif col == 1:           # référence
                    c.alignment = center
                    c.font = bold_f
                else:                    # observations
                    c.alignment = left
            ws.row_dimensions[pr].height = 20

    # ── Bloc synthèse ──
    ws.append([])
    ws.append(["SYNTHÈSE DU CONTRÔLE"])
    sr = ws.max_row
    ws.merge_cells(f"A{sr}:F{sr}")
    c = ws.cell(sr, 1)
    c.fill = hdr_fill; c.font = hdr_font
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[sr].height = 20

    for label in [
        "État général de l'équipement (Conforme / Réserves / Non conforme) :",
        "Nombre de pièces à remplacer :",
        "Équipement remis en service (Oui / Non) :",
    ]:
        ws.append([label])
        r = ws.max_row
        ws.merge_cells(f"A{r}:D{r}")
        ws.cell(r, 1).alignment = left
        ws.cell(r, 1).font = bold_f
        for col in range(1, 7):
            ws.cell(r, col).border = brd
        ws.merge_cells(f"E{r}:F{r}")
        ws.row_dimensions[r].height = 20

    # ── Signature ──
    ws.append([])
    ws.append(["Date du contrôle :", tour_date.strftime('%d/%m/%Y'),
               "", "Signature de l'opérateur :"])
    r = ws.max_row
    ws.cell(r, 1).font = bold_f
    ws.cell(r, 4).font = bold_f
    ws.row_dimensions[r].height = 40
    for col in (1, 2, 4):
        ws.cell(r, col).alignment = Alignment(horizontal="left", vertical="top")

    return ws.max_row


def add_etat_lieux_sheet(wb, tour_date, driver_name, stops=None,
                         par_arret=True, depot_addr=""):
    """Ajoute une feuille 'État des lieux' aux exports Excel.

    - par_arret=True  : UNE fiche par arrêt, pré-remplie (client, adresse, nature
      de l'intervention), séparées par un saut de page (1 fiche = 1 page).
    - par_arret=False : UNE seule fiche généraliste pour toute la tournée.
    Si aucun arrêt n'est fourni, une fiche généraliste est générée.
    """
    ws = wb.create_sheet("État des lieux")

    # Largeurs de colonnes (une seule fois pour toute la feuille)
    for col, width in zip(range(1, 7), [16, 42, 13, 15, 16, 40]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Mise en page impression : portrait, ajusté à la largeur d'une page
    try:
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    except Exception:
        pass  # la mise en page impression est un confort, non bloquant

    stops = stops or []

    # Mode généraliste (choisi par l'utilisateur ou aucun arrêt) → une seule fiche
    if not par_arret or not stops:
        _write_etat_lieux_block(
            ws, tour_date, driver_name,
            general=True, nb_arrets=len(stops), depot_addr=depot_addr,
        )
        return ws

    # Mode par arrêt → une fiche par arrêt, saut de page entre chaque
    n = len(stops)
    for i, stop in enumerate(stops):
        _produit_label, _option_label = _etat_lieux_prestation_parts(stop)
        last_row = _write_etat_lieux_block(
            ws, tour_date, driver_name,
            client_name=stop.get("client", ""),
            address=stop.get("address", ""),
            action_label=stop.get("action", ""),
            produit_label=_produit_label,
            option_label=_option_label,
        )
        # Saut de page après chaque fiche sauf la dernière
        if i < n - 1:
            ws.row_breaks.append(Break(id=last_row))
            ws.append([])   # ligne de respiration avant la fiche suivante

    return ws


def build_etat_lieux_flowables(styles, tour_date, driver_name,
                               client_name="", address="", action_label="",
                               produit_label="", option_label="",
                               fiche_index=None, fiche_total=None,
                               general=False, nb_arrets=None, depot_addr=""):
    """Retourne la liste de flowables reportlab pour UNE fiche État des lieux.

    - Mode par arrêt (défaut) : l'en-tête reprend le client, l'adresse
      d'intervention et la nature de l'intervention.
    - Mode généraliste (general=True) : fiche unique pour toute la tournée
      (en-tête tournée : opérateur, date, dépôt, nombre de points d'intervention).
    Les cases à cocher sont dessinées (contour seul).
    """
    story = []

    title_style = ParagraphStyle("el_title", parent=styles["Heading1"],
                                 fontSize=16, textColor=colors.HexColor("#1F4E79"),
                                 alignment=TA_CENTER, spaceAfter=2)
    subtitle_style = ParagraphStyle("el_subtitle", parent=styles["Normal"],
                                    fontSize=9, textColor=colors.grey,
                                    alignment=TA_CENTER, spaceAfter=6)
    intro_style = ParagraphStyle("el_intro", parent=styles["Normal"],
                                 fontSize=9, textColor=colors.HexColor("#555555"),
                                 leading=12, spaceAfter=8)
    section_style = ParagraphStyle("el_section", parent=styles["Heading2"],
                                   fontSize=12, textColor=colors.HexColor("#1F4E79"),
                                   spaceBefore=6, spaceAfter=4)
    cell_style = ParagraphStyle("el_cell", parent=styles["Normal"],
                                fontSize=9, leading=11)
    cell_bold  = ParagraphStyle("el_cell_bold", parent=styles["Normal"],
                                fontSize=9, leading=11, fontName="Helvetica-Bold")
    lbl_style  = ParagraphStyle("el_lbl", parent=styles["Normal"],
                                fontSize=9, leading=11, fontName="Helvetica-Bold")
    hdr_style  = ParagraphStyle("el_hdr", parent=styles["Normal"],
                                fontSize=8.5, fontName="Helvetica-Bold",
                                textColor=colors.white, alignment=1)
    zone_style = ParagraphStyle("el_zone", parent=styles["Normal"],
                                fontSize=9.5, fontName="Helvetica-Bold",
                                textColor=colors.HexColor("#1F4E79"))

    def P(txt, style=None):
        return Paragraph(str(txt) if txt is not None else "", style or cell_style)

    # ── Nouvelle page ──
    story.append(PageBreak())

    # ── Titre + sous-titre ──
    if general:
        story.append(Paragraph(
            "ÉTAT DES LIEUX – CONTRÔLE DES PIÈCES (FICHE TOURNÉE)", title_style))
        nb_txt = f" — {nb_arrets} arrêt(s)" if nb_arrets is not None else ""
        story.append(Paragraph(
            f"Tournée du {tour_date.strftime('%d/%m/%Y')}{nb_txt}", subtitle_style))
    else:
        story.append(Paragraph("ÉTAT DES LIEUX – CONTRÔLE DES PIÈCES", title_style))
        if fiche_index is not None and fiche_total is not None:
            story.append(Paragraph(
                f"Fiche {fiche_index} / {fiche_total} — {tour_date.strftime('%d/%m/%Y')}",
                subtitle_style))
        else:
            story.append(Paragraph(tour_date.strftime('%d/%m/%Y'), subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1,
                            color=colors.HexColor("#1F4E79"), spaceAfter=6))

    # ── Bloc d'entête ──
    if general:
        nb_lbl = f"{nb_arrets} arrêt(s)" if nb_arrets is not None else "—"
        client_para = Paragraph(
            "<b>Client / site :</b>&nbsp;&nbsp;<font color='#888888'>"
            "(à compléter sur le terrain)</font>", cell_style)
        depot_para = Paragraph(
            f"<b>Dépôt de départ :</b>&nbsp;&nbsp;{depot_addr or '—'}", cell_style)
        entete_data = [
            [P("Opérateur :", lbl_style),  P(driver_name or "—"),
             P("Date :", lbl_style),       P(tour_date.strftime('%d/%m/%Y'))],
            [P("Points d'intervention :", lbl_style), P(nb_lbl),
             P("", lbl_style),             P("")],
            [depot_para, P(""), P(""), P("")],
            [client_para, P(""), P(""), P("")],
        ]
        entete_table = Table(entete_data, colWidths=[42*mm, 55*mm, 42*mm, 55*mm])
        entete_table.setStyle(TableStyle([
            ("FONTSIZE",      (0, 0), (-1, -1), 9),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("SPAN",          (1, 1), (3, 1)),   # points d'intervention (valeur large)
            ("SPAN",          (0, 2), (3, 2)),   # dépôt sur toute la largeur
            ("SPAN",          (0, 3), (3, 3)),   # client sur toute la largeur
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOX",           (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("INNERGRID",     (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E5E5")),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ]))
        story.append(entete_table)
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(ETAT_LIEUX_INTRO_GENERAL, intro_style))
    else:
        option_para = Paragraph(
            f"<b>Option :</b>&nbsp;&nbsp;{option_label or '—'}", cell_style)
        addr_para = Paragraph(
            f"<b>Adresse d'intervention :</b>&nbsp;&nbsp;{address or '—'}", cell_style)
        entete_data = [
            [P("Opérateur :", lbl_style),     P(driver_name or "—"),
             P("Nature de l'intervention :", lbl_style), P(action_label or "—")],
            [P("Client / site :", lbl_style), P(client_name or "—"),
             P("Produit :", lbl_style),       P(produit_label or "—")],
            [option_para, P(""), P(""), P("")],
            [addr_para, P(""), P(""), P("")],
        ]
        entete_table = Table(entete_data, colWidths=[38*mm, 59*mm, 42*mm, 55*mm])
        entete_table.setStyle(TableStyle([
            ("FONTSIZE",      (0, 0), (-1, -1), 9),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("SPAN",          (0, 2), (3, 2)),   # option sur toute la largeur
            ("SPAN",          (0, 3), (3, 3)),   # adresse sur toute la largeur
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOX",           (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("INNERGRID",     (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E5E5")),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ]))
        story.append(entete_table)
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(ETAT_LIEUX_INTRO, intro_style))

    # ── Tableau de contrôle ──
    table_data = [[P(h, hdr_style) for h in
                   ["N° Article", "Désignation", "Bon\nétat", "À\nremplacer",
                    "Qté à\ncommander", "Observations"]]]
    row_styles = []

    for zone, pieces in PIECES_ETAT_LIEUX:
        table_data.append([P(zone, zone_style), P(""), P(""), P(""), P(""), P("")])
        zi = len(table_data) - 1
        row_styles.append(("BACKGROUND", (0, zi), (-1, zi), colors.HexColor("#DCE6F1")))
        row_styles.append(("SPAN", (0, zi), (-1, zi)))
        for ref, desig in pieces:
            table_data.append([
                P(ref, cell_bold), P(desig),
                _pdf_checkbox(), _pdf_checkbox(), P(""), P(""),
            ])

    CW = [20*mm, 62*mm, 16*mm, 20*mm, 22*mm, 54*mm]  # total = 194 mm (marges 8 mm)
    ctrl_table = Table(table_data, colWidths=CW, repeatRows=1)
    base = [
        ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("ALIGN",        (0, 0), (-1, 0), "CENTER"),
        ("ALIGN",        (2, 1), (4, -1), "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("LEFTPADDING",  (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]
    ctrl_table.setStyle(TableStyle(base + row_styles))
    story.append(ctrl_table)
    story.append(Spacer(1, 6*mm))

    # ── Synthèse du contrôle ──
    story.append(HRFlowable(width="100%", thickness=1,
                            color=colors.HexColor("#1F4E79"), spaceBefore=2, spaceAfter=6))
    story.append(Paragraph("Synthèse du contrôle", section_style))

    synth_rows = [
        [P("État général de l'équipement :", cell_bold),
         _options_inline([("Conforme", 20),
                          ("Conforme avec réserves", 44),
                          ("Non conforme", 26)])],
        [P("Nombre de pièces à remplacer :", cell_bold), P("")],
        [P("Équipement remis en service :", cell_bold),
         _options_inline([("Oui", 16), ("Non", 16)])],
    ]
    synth_table = Table(synth_rows, colWidths=[62*mm, 132*mm])
    synth_table.setStyle(TableStyle([
        ("FONTSIZE",      (0, 0), (-1, -1), 9),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
    ]))
    story.append(synth_table)
    story.append(Spacer(1, 8*mm))

    # ── Signature ──
    sign_rows = [[
        Paragraph("Date du contrôle : " + tour_date.strftime('%d/%m/%Y'), cell_bold),
        Paragraph("Signature de l'opérateur :", cell_bold),
    ]]
    sign_table = Table(sign_rows, colWidths=[97*mm, 97*mm], rowHeights=[26*mm])
    sign_table.setStyle(TableStyle([
        ("VALIGN",      (0, 0), (-1, -1), "TOP"),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(sign_table)

    return story


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def export_excel(result, tour_date, driver_name, fuel_price_per_l):
    pause_enabled = result.get("pause_dejeuner", True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feuille de tournée"

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    bold_f   = Font(bold=True)
    center   = Alignment(horizontal="center", vertical="center")
    left     = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin     = Side(style="thin", color="CCCCCC")
    brd      = Border(left=thin, right=thin, top=thin, bottom=thin)
    action_fills = {
        "Nettoyer":     PatternFill("solid", fgColor="AEC6E8"),
        "Déposer":      PatternFill("solid", fgColor="B7E5B4"),
        "Retirer":      PatternFill("solid", fgColor="F4B8C1"),
        "Chargement":   PatternFill("solid", fgColor="D4B8E0"),
        "Déchargement": PatternFill("solid", fgColor="FFE0B2"),
    }

    ws.merge_cells("A1:O1")
    c = ws["A1"]
    c.value     = f"FEUILLE DE TOURNÉE – {tour_date.strftime('%d/%m/%Y')}"
    c.font      = Font(bold=True, size=15, color="1F4E79")
    c.alignment = center
    ws.row_dimensions[1].height = 28

    # L'entête doit indiquer sans ambiguïté quel ordre a été exporté : c'est la
    # seule preuve, sur le document papier remis au chauffeur, que la
    # réorganisation manuelle a bien été prise en compte.
    _is_manual_export = bool(result.get("is_manual"))
    _recap_rows = [
        ("Chauffeur",           driver_name or "—"),
        ("Dépôt de départ",  result.get("depot_depart_addr", "")),
        ("Dépôt de retour",  result.get("depot_retour_addr", "")),
        ("Ordre des arrêts",
         "Ordre personnalisé manuellement (prioritaire sur l'optimisation)"
         if _is_manual_export else "Ordre optimisé automatiquement"),
        ("Distance totale",     f"{result['distance_km']:.1f} km"),
        ("Durée estimée",       f"{int(result['duration_min']//60)}h{int(result['duration_min']%60):02d}"),
        ("Carburant estimé",    f"{result['fuel_liters']:.1f} L  ({result['fuel_cost']:.2f} €)"),
        ("Pause déjeuner",      f"{PAUSE_DEJEUNER_MIN} min incluse" if pause_enabled else "Non comptée"),
        ("Gain optimisation",   f"{result['km_saved']:.1f} km – {int(result['time_saved_min'])} min"),
    ]
    if _is_manual_export and not result.get("recalc_ok", True):
        _recap_rows.append((
            "⚠️ Kilométrage",
            "Service de calcul d'itinéraire indisponible : distance et durée "
            "sont celles de l'ordre optimisé. L'ordre des arrêts ci-dessous "
            "est bien l'ordre manuel."))
    for label, val in _recap_rows:
        r = ws.max_row + 1
        ws.cell(r, 1).value = label + " :"
        ws.cell(r, 1).font  = bold_f
        ws.cell(r, 2).value = val
        ws.merge_cells(f"B{r}:O{r}")
        ws.cell(r, 2).alignment = left
    ws.append([])

    headers = ["Ordre", "Action", "Produit", "Option", "Couleur", "Qté", "Nom du client", "Adresse", "Durée (min)", "Pas avant", "Pas après", "Arrivée", "Départ", "Observations", "✓ Fait"]
    ws.append(headers)
    hr = ws.max_row
    for col, h in enumerate(headers, 1):
        c = ws.cell(hr, col)
        c.value = h; c.fill = hdr_fill; c.font = hdr_font
        c.alignment = center; c.border = brd
    ws.row_dimensions[hr].height = 20

    pause_fill = PatternFill("solid", fgColor="FFF9C4")
    pause_font = Font(bold=True, color="795548")
    pause_idx  = _find_pause_position(result["stops_ordered"])

    for si, stop in enumerate(result["stops_ordered"]):
        ws.append([stop["order_num"], stop["action"],
                   stop.get("produit", ""), stop.get("option", ""),
                   stop.get("couleur", "") or "—",
                   stop.get("qty_num", ""),
                   stop["client"] or "", stop["address"],
                   stop.get("duration_min", ""),
                   _fmt_min(stop.get("tw_early")),
                   _fmt_min(stop.get("tw_late")),
                   _fmt_min(stop.get("arrival_min")),
                   _fmt_min(stop.get("departure_min")),
                   stop.get("observations", "") or "",
                   ""])
        r = ws.max_row
        for col in range(1, 16):
            c = ws.cell(r, col)
            c.border = brd
            c.alignment = center if col not in (8, 14) else left
            if col == 12 and stop.get("violated"):
                c.font = Font(bold=True, color="DC3545")
        # Mise en évidence de la couleur exigée par le client
        if stop.get("couleur"):
            ws.cell(r, 5).font = Font(bold=True, color="1F4E79")
        ws.cell(r, 2).fill = action_fills.get(stop["action"],
                                               PatternFill("solid", fgColor="F0F0F0"))
        ws.row_dimensions[r].height = 18

        # ── Ligne pause déjeuner (après stop[pause_idx]) — seulement si activée ──
        if pause_enabled and si == pause_idx:
            _ps, _pe = _pause_slot(result["stops_ordered"], pause_idx)
            slot_txt = f"{_ps} – {_pe}" if _ps else f"{PAUSE_DEJEUNER_MIN} min"
            ws.append(["", f"🍽️ PAUSE DÉJEUNER — {PAUSE_DEJEUNER_MIN} min",
                        "", "", "", "", "", "",
                        PAUSE_DEJEUNER_MIN,
                        "", "", _ps, _pe,
                        f"Créneau : {slot_txt}", ""])
            pr = ws.max_row
            ws.merge_cells(f"B{pr}:G{pr}")
            for col in range(1, 16):
                c = ws.cell(pr, col)
                c.fill   = pause_fill
                c.font   = pause_font
                c.border = brd
                c.alignment = center if col not in (14,) else left
            ws.row_dimensions[pr].height = 18

    ws.append(["↩", "Retour dépôt", "", "", "", "", "", result.get("depot_retour_addr",""), "", "", "", "", "", "", ""])
    r = ws.max_row
    for col in range(1, 16):
        c = ws.cell(r, col)
        c.fill = PatternFill("solid", fgColor="FFF3CD")
        c.font = bold_f; c.border = brd
        c.alignment = center if col not in (8, 14) else left

    for col, width in zip(range(1, 16), [8, 14, 16, 14, 10, 6, 22, 44, 12, 12, 12, 12, 12, 30, 8]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # ── Feuille « Récap matériel » : ce qu'il faut charger / rapporter ──
    besoins_rec = result.get("besoins_records") or []
    if besoins_rec:
        wsm = wb.create_sheet("Récap matériel")
        wsm.append([f"RÉCAP MATÉRIEL – TOURNÉE DU {tour_date.strftime('%d/%m/%Y')}"])
        wsm.merge_cells("A1:F1")
        wsm.cell(1, 1).font = Font(bold=True, size=14, color="1F4E79")
        wsm.cell(1, 1).alignment = center
        wsm.row_dimensions[1].height = 26
        wsm.append([])
        mh = ["Article", "Couleur", "À charger au dépôt", "À rapporter au dépôt",
              "Variation nette", "Stock avant tournée"]
        wsm.append(mh)
        hr2 = wsm.max_row
        for col, h in enumerate(mh, 1):
            cc = wsm.cell(hr2, col)
            cc.fill = hdr_fill; cc.font = hdr_font
            cc.alignment = center; cc.border = brd
        wsm.row_dimensions[hr2].height = 22
        for b in besoins_rec:
            wsm.append([b.get("Article", ""), b.get("Couleur", "") or "Indifférente",
                        int(b.get("Sorties", 0)), int(b.get("Retours", 0)),
                        int(b.get("Net", 0)), b.get("StockAvant", "")])
            rr = wsm.max_row
            for col in range(1, 7):
                cc = wsm.cell(rr, col)
                cc.border = brd
                cc.alignment = left if col <= 2 else center
            if b.get("Manque", 0):
                for col in range(1, 7):
                    wsm.cell(rr, col).fill = PatternFill("solid", fgColor="F8D7DA")
                wsm.cell(rr, 3).font = Font(bold=True, color="B00020")
        for col, w in zip(range(1, 7), [30, 16, 20, 22, 16, 20]):
            wsm.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
        wsm.append([])
        wsm.append(["Note : « À charger » correspond aux actions Déposer "
                    "(et Déchargement si les manutentions sont comptabilisées). "
                    "« À rapporter » correspond aux actions Retirer."])
        wsm.merge_cells(start_row=wsm.max_row, start_column=1,
                        end_row=wsm.max_row, end_column=6)
        wsm.cell(wsm.max_row, 1).font = Font(italic=True, size=9, color="666666")
        wsm.cell(wsm.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")

    # ── Feuille « État des lieux » (une par arrêt ou une fiche généraliste) ──
    add_etat_lieux_sheet(
        wb, tour_date, driver_name, result["stops_ordered"],
        par_arret=result.get("etat_lieux_par_arret", True),
        depot_addr=result.get("depot_depart_addr", ""),
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────────────────────────────────────
# EXPORT PDF – reportlab (UTF-8 natif, zéro problème d'encodage)
# ─────────────────────────────────────────────────────────────────────────────

def export_pdf(result, tour_date, driver_name):
    pause_enabled = result.get("pause_dejeuner", True)
    buf    = BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                               leftMargin=8*mm, rightMargin=8*mm,
                               topMargin=10*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                 fontSize=16, textColor=colors.HexColor("#1F4E79"),
                                 alignment=TA_CENTER, spaceAfter=4)
    small_style = ParagraphStyle("small", parent=styles["Normal"],
                                 fontSize=8, textColor=colors.grey,
                                 alignment=TA_CENTER)
    cell_style  = ParagraphStyle("cell", parent=styles["Normal"],
                                 fontSize=8, leading=10, wordWrap="CJK")
    cell_bold   = ParagraphStyle("cell_bold", parent=styles["Normal"],
                                 fontSize=8, leading=10, fontName="Helvetica-Bold")
    cell_white  = ParagraphStyle("cell_white", parent=styles["Normal"],
                                 fontSize=8, leading=10, fontName="Helvetica-Bold",
                                 textColor=colors.white)
    check_style = ParagraphStyle("check", parent=styles["Normal"],
                                 fontSize=9, leading=13, leftIndent=4)
    obs_style          = ParagraphStyle("obs_title", parent=styles["Normal"],
                                 fontSize=11, fontName="Helvetica-Bold",
                                 textColor=colors.HexColor("#1F4E79"), spaceAfter=2)
    consigne_title_style = ParagraphStyle("consigne_title", parent=styles["Heading2"],
                                          fontSize=12, textColor=colors.HexColor("#1F4E79"),
                                          spaceAfter=4)
    action_label_style   = ParagraphStyle("action_label", parent=styles["Normal"],
                                          fontSize=10, fontName="Helvetica-Bold",
                                          textColor=colors.white, spaceAfter=2)
    consigne_text_style  = ParagraphStyle("consigne_text", parent=styles["Normal"],
                                          fontSize=9, leftIndent=4, spaceAfter=8)

    story = []

    # Titre
    story.append(Paragraph(
        f"FEUILLE DE TOURNÉE – {tour_date.strftime('%d/%m/%Y')}", title_style))
    story.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor("#1F4E79"), spaceAfter=6))

    # Récapitulatif
    _is_manual_export = bool(result.get("is_manual"))
    recap_data = [
        ["Chauffeur :",             driver_name or "—"],
        ["Dépôt de départ :", result.get("depot_depart_addr", "")],
        ["Dépôt de retour :",  result.get("depot_retour_addr", "")],
        ["Ordre des arrêts :",
         ("Ordre personnalisé manuellement (prioritaire sur l'optimisation)"
          if _is_manual_export else "Ordre optimisé automatiquement")],
        ["Distance totale :",       f"{result['distance_km']:.1f} km"],
        ["Durée de trajet :",
         f"{int(result['duration_min']//60)}h{int(result['duration_min']%60):02d}"],
        ["Carburant estimé :",
         f"{result['fuel_liters']:.1f} L  ({result['fuel_cost']:.2f} \u20ac)"],
        ["Pause déjeuner :",
         (f"{PAUSE_DEJEUNER_MIN} min incluse" if pause_enabled else "Non comptée")],
        ["Économie optimisation :",
         f"\u2212{result['km_saved']:.1f} km  /  \u2212{int(result['time_saved_min'])} min"],
    ]
    if _is_manual_export and not result.get("recalc_ok", True):
        recap_data.append([
            "Kilométrage :",
            "Service d'itinéraire indisponible — distance et durée sont celles "
            "de l'ordre optimisé. L'ordre des arrêts ci-dessous est l'ordre manuel."])
    recap_table = Table(recap_data, colWidths=[55*mm, 120*mm])
    recap_table.setStyle(TableStyle([
        ("FONTNAME",  (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",  (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE",  (0, 0), (-1, -1), 10),
        ("VALIGN",    (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1),
         [colors.white, colors.HexColor("#F5F5F5")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
    ]))
    story.append(recap_table)
    story.append(Spacer(1, 8*mm))

    # Tableau des arrêts
    action_bg = {
        "Nettoyer":     colors.HexColor("#AEC6E8"),
        "Déposer":      colors.HexColor("#B7E5B4"),
        "Retirer":      colors.HexColor("#F4B8C1"),
        "Chargement":   colors.HexColor("#D4B8E0"),
        "Déchargement": colors.HexColor("#FFE0B2"),
    }

    def P(txt, style=None):
        """Wrapper Paragraph pour word-wrap dans les cellules."""
        return Paragraph(str(txt) if txt is not None else "", style or cell_style)

    # En-têtes avec Paragraph pour alignement uniforme
    hdr_style = ParagraphStyle("hdr", parent=styles["Normal"],
                                fontSize=8, fontName="Helvetica-Bold",
                                textColor=colors.white, alignment=1)
    table_data = [[P(h, hdr_style) for h in
                   ["N°", "Action", "Produit", "Option", "Coul.", "Qté",
                    "Client", "Adresse", "Durée", "Pav.", "Pap.", "Arr.", "Dép.", "Obs."]]]
    row_styles = []

    depot_row_style = ParagraphStyle("dep_row", parent=styles["Normal"],
                                      fontSize=8, fontName="Helvetica-Bold", leading=10)

    # Ligne dépôt départ
    table_data.append([
        P(""), P("🏭 Dépôt départ", depot_row_style), P(""), P(""), P(""), P(""),
        P(""), P(result.get("depot_depart_addr",""), depot_row_style),
        P(""), P(""), P(""),
        P(_fmt_min(result.get("depart_min")) or "", depot_row_style), P(""), P("")
    ])
    row_styles += [("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#FFF3CD"))]

    pdf_pause_idx = _find_pause_position(result["stops_ordered"])

    for i, stop in enumerate(result["stops_ordered"]):
        arr_str = _fmt_min(stop.get("arrival_min")) or ""
        arr_para_style = cell_style
        if stop.get("violated"):
            arr_para_style = ParagraphStyle("viol", parent=styles["Normal"],
                                             fontSize=8, leading=10,
                                             textColor=colors.HexColor("#DC3545"),
                                             fontName="Helvetica-Bold")
        action_para_style = ParagraphStyle(f"act_{i}", parent=styles["Normal"],
                                            fontSize=8, leading=10,
                                            fontName="Helvetica-Bold",
                                            textColor=colors.HexColor("#1F4E79"))
        couleur_stop = stop.get("couleur", "") or ""
        couleur_style = (ParagraphStyle(f"coul_{i}", parent=styles["Normal"],
                                        fontSize=8, leading=10,
                                        fontName="Helvetica-Bold",
                                        textColor=colors.HexColor("#1F4E79"))
                         if couleur_stop else cell_style)
        table_data.append([
            P(str(stop["order_num"])),
            P(stop["action"], action_para_style),
            P(stop.get("produit", "")),
            P(stop.get("option", "")),
            P(couleur_stop or "—", couleur_style),
            P(str(stop.get("qty_num", ""))),
            P(stop["client"] or ""),
            P(stop["address"]),
            P(str(stop.get("duration_min", "") or "")),
            P(_fmt_min(stop.get("tw_early")) or ""),
            P(_fmt_min(stop.get("tw_late")) or ""),
            P(arr_str, arr_para_style),
            P(_fmt_min(stop.get("departure_min")) or ""),
            P(stop.get("observations", "") or ""),
        ])
        ri = len(table_data) - 1
        bg = action_bg.get(stop["action"], colors.HexColor("#F0F0F0"))
        row_styles.append(("BACKGROUND", (1, ri), (1, ri), bg))

        # ── Ligne pause déjeuner (insérée après pdf_pause_idx) — si activée ──
        if pause_enabled and i == pdf_pause_idx:
            _ps, _pe = _pause_slot(result["stops_ordered"], pdf_pause_idx)
            slot_txt = f"{_ps} – {_pe}" if _ps else f"{PAUSE_DEJEUNER_MIN} min"
            pause_style = ParagraphStyle("pause_lbl", parent=styles["Normal"],
                                          fontSize=8, leading=10,
                                          fontName="Helvetica-Bold",
                                          textColor=colors.HexColor("#795548"))
            table_data.append([
                P(""),
                P(f"🍽️ Pause déjeuner", pause_style),
                P(""), P(""), P(""), P(""), P(""), P(""),
                P(str(PAUSE_DEJEUNER_MIN), pause_style),
                P(""), P(""),
                P(_ps, pause_style),
                P(_pe, pause_style),
                P(f"Créneau : {slot_txt}", pause_style),
            ])
            pi = len(table_data) - 1
            row_styles.append(("BACKGROUND", (0, pi), (-1, pi),
                                colors.HexColor("#FFF9C4")))
            row_styles.append(("LINEABOVE",  (0, pi), (-1, pi),
                                1, colors.HexColor("#F9A825")))
            row_styles.append(("LINEBELOW",  (0, pi), (-1, pi),
                                1, colors.HexColor("#F9A825")))

    # Ligne dépôt retour
    table_data.append([
        P(""), P("🏁 Dépôt retour", depot_row_style), P(""), P(""), P(""), P(""),
        P(""), P(result.get("depot_retour_addr",""), depot_row_style),
        P(""), P(""), P(""), P(""),
        P(_fmt_min(result.get("return_min")) or "", depot_row_style), P("")
    ])
    last = len(table_data) - 1
    row_styles += [("BACKGROUND", (0, last), (-1, last), colors.HexColor("#FFF3CD"))]

    # Largeur utile A4 avec marges 8mm : 194mm
    # N°=7, Action=21, Produit=20, Option=14, Coul.=14, Qté=7, Client=16,
    # Adresse=28, Durée=9, Pav.=10, Pap.=10, Arr.=10, Dép.=10, Obs.=18 → 194mm
    CW = [7*mm, 21*mm, 20*mm, 14*mm, 14*mm, 7*mm, 16*mm, 28*mm,
          9*mm, 10*mm, 10*mm, 10*mm, 10*mm, 18*mm]
    stops_table = Table(table_data, colWidths=CW, repeatRows=1)
    base_style = [
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("ALIGN",         (7, 1), (7, -1), "LEFT"),
        ("ALIGN",         (1, 1), (3, -1), "LEFT"),
        ("ALIGN",         (13, 1), (13, -1), "LEFT"),
        ("NOSPLIT",        (1, 1), (4, -1)),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -2), [colors.white, colors.HexColor("#F9F9F9")]),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 3),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 3),
    ]
    stops_table.setStyle(TableStyle(base_style + row_styles))
    story.append(stops_table)
    story.append(Spacer(1, 6*mm))

    # ── Récapitulatif matériel (chargement / retour dépôt) ──
    besoins_rec = result.get("besoins_records") or []
    if besoins_rec:
        story.append(HRFlowable(width="100%", thickness=1,
                                 color=colors.HexColor("#1F4E79"),
                                 spaceBefore=2, spaceAfter=6))
        mat_title_style = ParagraphStyle("mat_title", parent=styles["Heading2"],
                                          fontSize=12,
                                          textColor=colors.HexColor("#1F4E79"),
                                          spaceAfter=4)
        story.append(Paragraph("Matériel à charger au départ du dépôt", mat_title_style))
        mat_data = [[P(h, hdr_style) for h in
                     ["Article", "Couleur", "À charger", "À rapporter",
                      "Variation", "Stock avant tournée"]]]
        mat_styles = []
        for bi, b in enumerate(besoins_rec, 1):
            mat_data.append([
                P(b.get("Article", "")),
                P(b.get("Couleur", "") or "Indifférente"),
                P(str(b.get("Sorties", 0))),
                P(str(b.get("Retours", 0))),
                P(f"{b.get('Net', 0):+d}"),
                P(str(b.get("StockAvant", ""))),
            ])
            if b.get("Manque", 0):
                mat_styles.append(("BACKGROUND", (0, bi), (-1, bi),
                                    colors.HexColor("#F8D7DA")))
        mat_table = Table(mat_data,
                          colWidths=[52*mm, 26*mm, 22*mm, 24*mm, 22*mm, 48*mm],
                          repeatRows=1)
        mat_table.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
            ("ALIGN",          (0, 1), (1, -1), "LEFT"),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F9F9F9")]),
            ("GRID",           (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",     (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 3),
        ] + mat_styles))
        story.append(mat_table)
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(
            "Les lignes surlignées en rouge signalent un stock insuffisant au "
            "moment de la préparation de la tournée.",
            ParagraphStyle("mat_note", parent=styles["Normal"], fontSize=8,
                            textColor=colors.HexColor("#666666"))))
        story.append(Spacer(1, 4*mm))

    # ── Carte de la tournée ──
    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor("#1F4E79"), spaceBefore=4, spaceAfter=6))
    map_title_style = ParagraphStyle("map_title", parent=styles["Heading2"],
                                      fontSize=12, textColor=colors.HexColor("#1F4E79"),
                                      spaceAfter=4)
    story.append(Paragraph("Carte de la tournée", map_title_style))
    try:
        map_bytes = generate_map_image(
            result["depot_coords"], result["stops_ordered"], result["geometry"]
        )
        img_buf = BytesIO(map_bytes)
        rl_img  = RLImage(img_buf, width=175*mm, height=120*mm)
        story.append(rl_img)
    except Exception as e:
        story.append(Paragraph(f"(Carte non disponible : {e})", styles["Normal"]))

    # ── Consignes par action ──
    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor("#1F4E79"), spaceBefore=4, spaceAfter=6))
    story.append(Paragraph("Consignes par type d'intervention", consigne_title_style))

    action_header_colors = {
        "Nettoyer":     "#1f6aa5",
        "Déposer":      "#28a745",
        "Retirer":      "#dc3545",
        "Chargement":   "#7B1FA2",
        "Déchargement": "#E65100",
    }
    seen = []
    for stop in result["stops_ordered"]:
        if stop["action"] not in seen:
            seen.append(stop["action"])
    for action, consigne_text in ACTION_CONSIGNES.items():
        if action not in seen:
            continue
        hcol = colors.HexColor(action_header_colors.get(action, "#555555"))
        consigne_data = [[Paragraph(f"■  {action}", action_label_style)]]
        consigne_table = Table(consigne_data, colWidths=[194*mm])
        consigne_table.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), hcol),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ]))
        story.append(consigne_table)
        story.append(Paragraph(consigne_text, consigne_text_style))

    # ── Checklist par action ──
    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor("#1F4E79"), spaceBefore=4, spaceAfter=6))
    story.append(Paragraph("Checklist par intervention", consigne_title_style))


    # ── Annotations / Observations ──
    story.append(Spacer(1, 4*mm))
    story.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor("#1F4E79"), spaceBefore=4, spaceAfter=6))
    story.append(Paragraph("Annotations / Observations", obs_style))
    story.append(Paragraph(
        "À remplir par le chauffeur durant ou à l'issue de la tournée :",
        ParagraphStyle("obs_sub", parent=styles["Normal"], fontSize=8,
                       textColor=colors.grey, spaceAfter=4)
    ))
    obs_rows = [[""] for _ in range(8)]
    obs_table = Table(obs_rows, colWidths=[194*mm], rowHeights=[10*mm]*8)
    obs_table.setStyle(TableStyle([
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("BACKGROUND",    (0, 0), (-1, -1), colors.white),
        ("TOPPADDING",    (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(obs_table)
    story.append(Spacer(1, 4*mm))

    # ── Pages « État des lieux » : une par arrêt ou une fiche généraliste ──
    _el_stops    = result.get("stops_ordered", [])
    _el_par_arret = result.get("etat_lieux_par_arret", True)
    if _el_par_arret and _el_stops:
        _el_total = len(_el_stops)
        for _el_i, _el_stop in enumerate(_el_stops, 1):
            _pl, _ol = _etat_lieux_prestation_parts(_el_stop)
            story += build_etat_lieux_flowables(
                styles, tour_date, driver_name,
                client_name=_el_stop.get("client", ""),
                address=_el_stop.get("address", ""),
                action_label=_el_stop.get("action", ""),
                produit_label=_pl, option_label=_ol,
                fiche_index=_el_i, fiche_total=_el_total,
            )
    else:
        # Fiche généraliste unique pour toute la tournée
        story += build_etat_lieux_flowables(
            styles, tour_date, driver_name,
            general=True, nb_arrets=len(_el_stops),
            depot_addr=result.get("depot_depart_addr", ""),
        )

    # ── Footer ──
    story.append(Spacer(1, 4*mm))
    story.append(HRFlowable(width="100%", thickness=0.5,
                             color=colors.grey, spaceBefore=4))
    story.append(Paragraph(
        f"Généré le {datetime.datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        " — Optimiseur de Tournées", small_style))

    doc.build(story)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("## 🚛 Optimiseur de Tournées")
    st.caption("Assainissement · WC Chimiques")
    st.divider()

    # ═════════════════════════════════════════════════════════════════════════
    # IMPORT DU FICHIER DE STOCK (à faire à chaque ouverture de l'application)
    # ═════════════════════════════════════════════════════════════════════════
    st.subheader("📦 Stock du parc")

    up_stock = st.file_uploader(
        "Fichier Excel du stock",
        type=["xlsx", "xlsm"],
        key="uploader_stock",
        help=(
            "À importer **à chaque ouverture** de l'application : Streamlit ne "
            "conserve aucune donnée entre deux sessions.\n\n"
            "Le fichier doit contenir un onglet **Stock** avec les colonnes "
            "**Article**, **Couleur**, **Installés**, **En stock**. "
            "Le modèle téléchargeable depuis l'onglet *Stock & disponibilité* "
            "est directement réimportable."
        ),
    )

    if up_stock is not None:
        file_id = f"{up_stock.name}-{up_stock.size}"
        if st.session_state.stock_last_file_id != file_id:
            df_imp, msg = parse_stock_excel(up_stock)
            if df_imp is not None:
                st.session_state.df_stock           = df_imp
                st.session_state.stock_source       = f"Import : {up_stock.name}"
                st.session_state.stock_import_ok    = True
                st.session_state.stock_import_name  = up_stock.name
                st.session_state.stock_last_file_id = file_id
                st.success(f"✅ {msg}")
                st.rerun()
            else:
                st.error(f"❌ {msg}")

    _nb_ref  = len(st.session_state.df_stock)
    _tot_stk = int(st.session_state.df_stock["En stock"].sum())
    _tot_ins = int(st.session_state.df_stock["Installés"].sum())

    if st.session_state.stock_import_ok:
        st.success(
            f"📥 **{st.session_state.stock_import_name}**  \n"
            f"{_nb_ref} référence(s) · {_tot_stk} en stock · {_tot_ins} installé(s)"
        )
    else:
        st.warning(
            f"⚠️ **Aucun fichier importé.** Les valeurs de référence du relevé "
            f"exploitation sont utilisées ({_tot_stk} en stock · {_tot_ins} installé(s)). "
            f"Importez le fichier du jour pour travailler sur des données à jour."
        )

    st.divider()

    st.subheader("📅 Tournée")
    st.session_state.tour_date = st.date_input(
        "Date", value=st.session_state.tour_date, format="DD/MM/YYYY")
    st.session_state.driver = st.text_input(
        "👤 Chauffeur", value=st.session_state.driver,
        placeholder="ex : Jean Dupont")
    st.divider()
    st.subheader("⚖️ Règles de travail")
    st.session_state.heure_min_depart = st.time_input(
        "🕖 Départ au plus tôt",
        value=st.session_state.heure_min_depart,
        step=300,
        help=(
            "Heure minimale de départ du dépôt. "
            "En France, le Code du travail impose un repos quotidien de 11h consécutives "
            "(art. L3131-1) et le transport routier interdit le travail avant 5h00 "
            "dans la plupart des conventions collectives. "
            "Valeur recommandée : 07h00."
        )
    )
    st.caption(
        "ℹ️ *Code du travail — art. L3131-1 :* "
        "repos quotidien minimum de **11h consécutives**. "
        "Convention collective transport : départ rarement avant **06h00**."
    )

    st.divider()
    st.subheader("🚛 Véhicule")
    fuel_conso = st.number_input(
        "Consommation (L/100 km)", min_value=5.0, max_value=30.0,
        value=12.20, step=0.5,
        help="Petit camion avec remorque : environ 14–16 L/100 km")
    fuel_price = st.number_input(
        "Prix carburant (€/L)", min_value=1.0, max_value=3.5,
        value=1.85, step=0.01)

    st.divider()
    st.subheader("🏭 Dépôts")
    depot_options = list(DEPOTS.keys())

    st.session_state.depot_depart_key = st.selectbox(
        "📍 Dépôt de départ",
        options=depot_options,
        index=depot_options.index(st.session_state.depot_depart_key)
              if st.session_state.depot_depart_key in depot_options else 0,
    )
    st.caption(f"📌 {DEPOTS[st.session_state.depot_depart_key]}")

    st.session_state.depot_retour_key = st.selectbox(
        "🏁 Dépôt de retour",
        options=depot_options,
        index=depot_options.index(st.session_state.depot_retour_key)
              if st.session_state.depot_retour_key in depot_options else 0,
    )
    st.caption(f"📌 {DEPOTS[st.session_state.depot_retour_key]}")

    st.divider()
    st.subheader("🚦 Trafic")

    def _on_peri_change():
        """Efface le résultat précédent pour forcer une ré-optimisation."""
        st.session_state.result = None
        # Sans cela, un ordre manuel obsolète resterait actif et continuerait
        # d'être exporté alors que la tournée de référence a été invalidée.
        reset_manual_order()

    st.session_state.eviter_peri = st.toggle(
        "Périphérique toulousain",
        value=st.session_state.eviter_peri,
        on_change=_on_peri_change,
        help=(
            "En heure de pointe (07h–09h ou 17h–19h), reporte automatiquement "
            "les arrêts situés à l'intérieur de la rocade toulousaine :\n\n"
            "• Pointe matin → arrêts intra-rocade planifiés après 09:00\n"
            "• Pointe soir  → arrêts intra-rocade planifiés avant 16:30\n\n"
            "Le chauffeur traite d'abord les arrêts extérieurs pendant les bouchons, "
            "puis entre en ville une fois la circulation fluide."
        ),
    )
    if st.session_state.eviter_peri:
        hm    = st.session_state.heure_min_depart
        _ref  = hm.hour * 60 + hm.minute
        _mat  = _POINTE_MATIN_DEBUT <= _ref <= _POINTE_MATIN_FIN
        _soir = _POINTE_SOIR_DEBUT  <= _ref <= _POINTE_SOIR_FIN
        if _mat:
            st.warning(
                f"⚠️ Départ à **{hm.strftime('%H:%M')}** — pointe matin.\n\n"
                "Les arrêts intra-rocade seront repoussés après **09:00**."
            )
        elif _soir:
            st.warning(
                f"⚠️ Départ à **{hm.strftime('%H:%M')}** — pointe soir.\n\n"
                "Les arrêts intra-rocade seront planifiés avant **16:30**."
            )
        else:
            st.info(
                f"ℹ️ Départ à **{hm.strftime('%H:%M')}** — hors heure de pointe. "
                "Aucun report automatique ne sera appliqué."
            )

    st.divider()
    st.subheader("🍽️ Pause déjeuner")

    def _on_pause_change():
        """Efface le résultat précédent pour forcer un recalcul du planning."""
        st.session_state.result = None
        reset_manual_order()

    st.session_state.pause_dejeuner = st.toggle(
        f"Inclure la pause déjeuner ({PAUSE_DEJEUNER_MIN} min)",
        value=st.session_state.pause_dejeuner,
        on_change=_on_pause_change,
        help=(
            "Active ou désactive la pause méridienne de "
            f"{PAUSE_DEJEUNER_MIN} minutes.\n\n"
            "• **Activée** : une pause est insérée autour de midi ; elle décale "
            "les heures d'arrivée des arrêts suivants, l'heure de retour au dépôt "
            "et la durée totale. Elle apparaît dans le PDF et l'Excel.\n"
            "• **Désactivée** : aucune pause n'est comptée ; le planning et les "
            "exports sont calculés sans les 30 minutes."
        ),
    )
    if st.session_state.pause_dejeuner:
        st.caption(
            f"🍽️ Une **pause de {PAUSE_DEJEUNER_MIN} min** sera insérée autour de "
            "midi et reportée dans la tournée et les exports."
        )
    else:
        st.caption(
            "⛔ **Aucune pause déjeuner** ne sera comptée dans la tournée "
            "ni dans les documents exportés."
        )

    st.divider()
    st.subheader("📋 Fiches état des lieux")

    st.session_state.etat_lieux_par_arret = st.toggle(
        "Une fiche par arrêt",
        value=st.session_state.etat_lieux_par_arret,
        help=(
            "Contrôle le nombre de fiches d'état des lieux ajoutées aux exports "
            "PDF et Excel :\n\n"
            "• **Activé** : une fiche par arrêt, pré-remplie avec le client, "
            "l'adresse et la nature de l'intervention (contrôle par équipement).\n"
            "• **Désactivé** : une **seule fiche généraliste** pour toute la "
            "tournée (en-tête tournée + champ à compléter sur le terrain)."
        ),
    )
    if st.session_state.etat_lieux_par_arret:
        st.caption(
            "🧾 **Une fiche par arrêt** sera générée dans le PDF et l'Excel "
            "(pré-remplie : client, adresse, nature de l'intervention)."
        )
    else:
        st.caption(
            "🗂️ **Une seule fiche généraliste** couvrant toute la tournée sera "
            "générée dans le PDF et l'Excel."
        )

    if active_result():
        st.divider()
        # Les chiffres affichés ici doivent correspondre à ce qui sera exporté :
        # on lit donc le résultat actif (ordre manuel prioritaire), pas le
        # résultat brut de l'optimiseur.
        r = active_result()
        pause_txt = (f"🍽️ Pause : {PAUSE_DEJEUNER_MIN} min incluse"
                     if r.get("pause_dejeuner", True) else "⛔ Sans pause déjeuner")
        _entete = ("✏️ Tournée en ordre manuel"
                   if is_manual_order() else "✅ Dernière optimisation")
        st.success(
            f"{_entete}\n\n"
            f"**{r['distance_km']:.1f} km** · "
            f"**{int(r['duration_min']//60)}h{int(r['duration_min']%60):02d}**\n\n"
            f"⛽ {r['fuel_liters']:.1f} L ({r['fuel_cost']:.2f} €)\n\n"
            f"⏱ Gain : {int(r['time_saved_min'])} min / {r['km_saved']:.1f} km\n\n"
            f"{pause_txt}")

# ─────────────────────────────────────────────────────────────────────────────
# ONGLETS
# ─────────────────────────────────────────────────────────────────────────────

st.title("Optimisation Tournées WC chimiques - Deldossi Assainissement")
st.caption(f"🏭 Départ : **{st.session_state.depot_depart_key}** · Retour : **{st.session_state.depot_retour_key}**")

tab_saisie, tab_optim, tab_stock, tab_export = st.tabs([
    "📋  Saisie des arrêts",
    "🗺️  Tournée optimisée",
    "📦  Stock & disponibilité",
    "📥  Export",
])

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 1 — SAISIE
# ══════════════════════════════════════════════════════════════════════════════

with tab_saisie:
    # ── Helpers (définis avant toute utilisation) ─────────────────────────────
    def _flush_editor():
        """Applique les édits en cours du tableau dans df_stops."""
        key = "editor_stops"
        if key not in st.session_state:
            return
        state = st.session_state[key]
        df    = st.session_state.df_stops.copy()
        # Éditions de cellules
        for row_idx, cols in (state.get("edited_rows") or {}).items():
            for col, val in cols.items():
                df.at[int(row_idx), col] = val
        # Lignes ajoutées via le "+" natif du data_editor
        for row in (state.get("added_rows") or []):
            df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
        # Lignes supprimées via la corbeille native
        deleted = sorted(state.get("deleted_rows") or [], reverse=True)
        for idx in deleted:
            df = df.drop(index=idx).reset_index(drop=True)
        # Vide l'option pour les produits sans option (Urinoir / Lave-main / WC handicapé)
        df = _normalize_option(df)
        st.session_state.df_stops = df

    def _current_df():
        """Retourne le df avec les éditions non encore flushées."""
        key   = "editor_stops"
        df    = st.session_state.df_stops.copy()
        if key not in st.session_state:
            return _normalize_option(df)
        state = st.session_state[key]
        for row_idx, cols in (state.get("edited_rows") or {}).items():
            for col, val in cols.items():
                df.at[int(row_idx), col] = val
        for row in (state.get("added_rows") or []):
            df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
        deleted = sorted(state.get("deleted_rows") or [], reverse=True)
        for idx in deleted:
            df = df.drop(index=idx).reset_index(drop=True)
        return _normalize_option(df)

    # ══════════════════════════════════════════════════════════════════════════
    # BLOC 1 — IMPORT DE LA TOURNÉE DEPUIS EXCEL
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("### 1\ufe0f\u20e3  Import de la tournée")
    st.caption(
        "Préparez la tournée dans Excel puis injectez-la ici, ou saisissez-la "
        "directement dans le tableau ci-dessous. L'ordre des lignes n'a aucune "
        "importance : l'itinéraire est recalculé."
    )

    col_up, col_mode = st.columns([3, 2], gap="medium")

    with col_up:
        up_stops = st.file_uploader(
            "Fichier Excel de la tournée",
            type=["xlsx", "xlsm"],
            key="uploader_stops",
            help=(
                "Le classeur doit comporter une ligne d'en-tête avec au minimum une "
                "colonne **Adresse**. Les autres colonnes (Action, Produit, Option, "
                "Couleur, Quantité, Durée, Pas avant, Pas après, Observations) sont "
                "facultatives et prennent une valeur par défaut si absentes.\n\n"
                "Le **modèle vierge** ci-contre garantit la compatibilité et intègre "
                "des listes déroulantes."
            ),
        )

    with col_mode:
        _mode_import = st.radio(
            "Mode d'import",
            ["Remplacer la saisie", "Ajouter aux arrêts existants"],
            key="stops_import_mode",
            help=(
                "**Remplacer** : le tableau est réinitialisé avec le contenu du fichier.\n\n"
                "**Ajouter** : les arrêts du fichier sont ajoutés à la suite de ceux "
                "déjà saisis (utile pour fusionner plusieurs secteurs)."
            ),
        )

    if up_stops is not None:
        _file_id_stops = f"{up_stops.name}-{up_stops.size}-{_mode_import}"
        if st.session_state.stops_last_file_id != _file_id_stops:
            _df_imp, _msg_imp, _warn_imp = parse_stops_excel(up_stops)
            if _df_imp is not None:
                if _mode_import.startswith("Ajouter"):
                    _base = st.session_state.df_stops.copy()
                    # On écarte les lignes vides du tableau d'accueil
                    _base = _base[_base["Adresse"].fillna("").astype(str).str.strip() != ""]
                    _df_final = pd.concat([_base, _df_imp], ignore_index=True)
                else:
                    _df_final = _df_imp
                if "Couleur" not in _df_final.columns:
                    _df_final["Couleur"] = ""
                st.session_state.df_stops           = _df_final.reset_index(drop=True)
                st.session_state.stops_last_file_id = _file_id_stops
                st.session_state.stops_import_name  = up_stops.name
                st.session_state.stops_import_msg   = _msg_imp
                st.session_state.stops_import_warn  = _warn_imp
                st.session_state.result             = None
                reset_manual_order()
                if "editor_stops" in st.session_state:
                    del st.session_state["editor_stops"]
                st.rerun()
            else:
                st.error(f"❌ {_msg_imp}")

    if st.session_state.stops_import_msg:
        st.success(
            f"📥 **{st.session_state.stops_import_name}** — "
            f"{st.session_state.stops_import_msg}"
        )
        if st.session_state.stops_import_warn:
            with st.expander(
                    f"⚠️ {len(st.session_state.stops_import_warn)} point(s) de "
                    "vigilance sur l'import"):
                for _w in st.session_state.stops_import_warn:
                    st.markdown(f"- {_w}")
                st.caption(
                    "Ces corrections ont été appliquées automatiquement pour ne pas "
                    "bloquer l'import. Vérifiez les lignes concernées dans le tableau "
                    "ci-dessous."
                )

    col_m1, col_m2 = st.columns(2)
    with col_m1:
        st.download_button(
            "📄 Télécharger le modèle vierge",
            data=export_stops_excel(None, tour_date=st.session_state.tour_date,
                                    driver_name=st.session_state.driver),
            file_name="Modele_saisie_tournee.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Classeur prêt à remplir : listes déroulantes sur Action, Produit, "
                 "Option et Couleur, contrôles de saisie sur Quantité et Durée, "
                 "et onglet Notice.")
    with col_m2:
        _df_export_stops = _current_df() if "editor_stops" in st.session_state \
            else st.session_state.df_stops
        _df_export_stops = _df_export_stops[
            _df_export_stops["Adresse"].fillna("").astype(str).str.strip() != ""]
        st.download_button(
            "⬇️ Exporter la saisie en cours",
            data=export_stops_excel(_df_export_stops,
                                    tour_date=st.session_state.tour_date,
                                    driver_name=st.session_state.driver),
            file_name=f"Tournee_saisie_{st.session_state.tour_date.strftime('%d-%m-%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            disabled=(len(_df_export_stops) == 0),
            help="Sauvegarde le tableau actuel dans un classeur réimportable. "
                 "Indispensable pour retrouver la tournée lors d'une prochaine session.")

    st.divider()

    # ══════════════════════════════════════════════════════════════════════════
    # BLOC 2 — TABLEAU DE SAISIE (+ actions sur le côté)
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("### 2\ufe0f\u20e3  Tableau de saisie des arrêts")
    st.caption(
        f"Départ : **{st.session_state.depot_depart_key}** · "
        f"Retour : **{st.session_state.depot_retour_key}** · "
        f"l'ordre de saisie n'a pas d'importance."
    )

    col_table, col_side = st.columns([4, 1.3], gap="medium")

    with col_table:
        # Légende des actions (inline, sans colonnes imbriquées)
        chips = " ".join(
            f'<span style="background:{c};padding:2px 8px;border-radius:4px;'
            f'font-size:0.8em;margin-right:4px">■ {a}</span>'
            for a, c in ACTION_COLORS.items()
        )
        st.markdown(chips, unsafe_allow_html=True)
        st.markdown("")

        # Colonne de sélection pour la suppression (migration + typage booléen)
        if "Suppr" not in st.session_state.df_stops.columns:
            st.session_state.df_stops.insert(0, "Suppr", False)
        st.session_state.df_stops["Suppr"] = (
            st.session_state.df_stops["Suppr"].fillna(False).astype(bool))

        # Migration : ajout de la colonne Couleur sur les sessions antérieures
        if "Couleur" not in st.session_state.df_stops.columns:
            st.session_state.df_stops["Couleur"] = ""
        st.session_state.df_stops["Couleur"] = (
            st.session_state.df_stops["Couleur"].fillna("").astype(str))

        st.data_editor(
            st.session_state.df_stops,
            use_container_width=True,
            num_rows="dynamic",
            column_order=["Suppr", "Action", "Produit", "Option", "Couleur",
                          "Quantité", "Nom du client", "Adresse", "Durée (min)",
                          "Pas avant", "Pas après", "Observations"],
            column_config={
                "Suppr": st.column_config.CheckboxColumn(
                    "🗑️", width="small", default=False,
                    help="Cochez les lignes à supprimer, puis cliquez sur "
                         "« Supprimer la sélection » à droite."),
                "Action": st.column_config.SelectboxColumn(
                    "Action", required=True, width="small",
                    options=["Nettoyer", "Déposer", "Retirer", "Chargement", "Déchargement"]),
                "Produit": st.column_config.SelectboxColumn(
                    "Produit", required=True, width="medium",
                    options=PRODUITS_CATALOGUE),
                "Option": st.column_config.SelectboxColumn(
                    "Option (WC chim. uniquement)", width="medium",
                    options=["", "Lave-main", "Urinoir"],
                    help="Réservée au WC chimique. Pour un Urinoir, un Lave-main ou "
                         "un WC handicapé, l'option est automatiquement vidée."),
                "Couleur": st.column_config.SelectboxColumn(
                    "🎨 Couleur", width="small",
                    options=[""] + COULEURS_PRODUITS,
                    help="Critère facultatif : à renseigner uniquement si le client "
                         "exige une couleur précise. Laissée vide, la disponibilité "
                         "est contrôlée toutes couleurs confondues."),
                "Quantité": st.column_config.NumberColumn(
                    "Qté", required=True, width="small",
                    min_value=1, max_value=20, step=1, default=1,
                    help="Nombre d'unités. Pour WC chimique : 1 option par WC."),
                "Nom du client": st.column_config.TextColumn(
                    "Nom du client", width="medium"),
                "Adresse": st.column_config.TextColumn(
                    "Adresse complète (rue, ville, CP)", width="large",
                    help="Ex : Place d'Hautpoul 81600 Gaillac"),
                "Durée (min)": st.column_config.NumberColumn(
                    "⏱ Durée (min)", width="small", min_value=1, max_value=480,
                    step=5, default=30,
                    help="Durée de l'intervention sur place en minutes (ex: 30)"),
                "Pas avant": st.column_config.TextColumn(
                    "⏰ Pas avant", width="small",
                    help="Arriver au plus tôt à cette heure (format HH:MM). Ex : 09:00"),
                "Pas après": st.column_config.TextColumn(
                    "⏰ Pas après", width="small",
                    help="Arriver au plus tard à cette heure (format HH:MM). Ex : 11:30"),
                "Observations": st.column_config.TextColumn(
                    "📝 Observations", width="large",
                    help="Notes ou remarques particulières pour cet arrêt (ex : code portail, contact sur place…)"),
            },
            hide_index=False,
            key="editor_stops",
        )
        # df_stops n'est réécrit que par _flush_editor() lors d'une action.

        live_df    = _current_df()
        valid_rows = live_df[live_df["Adresse"].fillna("").str.strip() != ""]
        n_valid    = len(valid_rows)
        st.caption(f"📍 **{n_valid}** arrêt(s) avec adresse renseignée")

        # Validation Produit / Option
        if "Produit" in live_df.columns and "Option" in live_df.columns:
            wc_rows_no_opt = valid_rows[
                (valid_rows["Produit"] == "WC chimique") &
                (valid_rows["Option"].fillna("").str.strip() == "")
            ]
            if len(wc_rows_no_opt) > 0:
                st.warning(
                    f"\U0001f6bd **{len(wc_rows_no_opt)} arr\u00eat(s) avec WC chimique sans option.** "
                    "Chaque WC chimique doit \u00eatre accompagn\u00e9 d'un **Urinoir** ou d'un "
                    "**Lave-main** (Code du travail, art. R4228-7). "
                    "S\u00e9lectionnez une option dans la colonne *Option*."
                )
            st.caption(
                "ℹ️ La colonne *Option* n'est utilisée que pour le **WC chimique**. "
                "Pour un **Urinoir**, un **Lave-main** ou un **WC handicapé**, "
                "l'option est automatiquement vidée."
            )

        # ── Contrôle temps réel de la disponibilité en stock ──────────────────
        _besoins_live = besoins_tournee(
            live_df, inclure_manutentions=st.session_state.stock_manutentions)
        _dispo_live = controle_disponibilite(st.session_state.df_stock, _besoins_live)

        if len(_dispo_live) == 0:
            _nb_retours = int(_besoins_live["Retours"].sum()) if len(_besoins_live) else 0
            if _nb_retours > 0:
                st.info(
                    f"📥 Cette tournée génère uniquement des **retours au dépôt** "
                    f"({_nb_retours} unité(s) réintégrée(s) en stock). "
                    f"Aucun prélèvement de stock n'est nécessaire."
                )
            else:
                st.caption(
                    "📦 Aucun mouvement de matériel n'est généré par cette saisie "
                    "(seules les actions *Déposer* et *Retirer* — et les manutentions "
                    "si l'option est activée — affectent le parc)."
                )
        else:
            _manquants = _dispo_live[_dispo_live["Manque"] > 0]
            _justes    = _dispo_live[(_dispo_live["Manque"] == 0) &
                                     (_dispo_live["Disponible"] == _dispo_live["Besoin"])]
            if len(_manquants) > 0:
                _lignes = "  \n".join(
                    f"- **{r['Article']}** · {r['Couleur']} → besoin **{r['Besoin']}**, "
                    f"disponible **{r['Disponible']}** → **{r['Manque']} manquant(s)**"
                    for _, r in _manquants.iterrows()
                )
                st.error(
                    "🔴 **Stock insuffisant pour cette tournée**  \n" + _lignes +
                    "  \n\n➡️ Ajustez les quantités, les couleurs, ou consultez "
                    "l'onglet **📦 Stock & disponibilité** pour arbitrer."
                )
            elif len(_justes) > 0:
                st.warning(
                    f"🟠 **Stock juste suffisant** sur {len(_justes)} référence(s) : "
                    "la tournée videra complètement ces lignes de stock. "
                    "Aucune marge en cas d'aléa terrain."
                )
            else:
                _tot_sorties = int(_dispo_live["Besoin"].sum())
                st.success(
                    f"🟢 **Stock suffisant** — {_tot_sorties} unité(s) à sortir du dépôt, "
                    f"toutes disponibles. Détail dans l'onglet "
                    f"**📦 Stock & disponibilité**."
                )

    with col_side:
        st.markdown("**Actions**")
        if st.button("➕ Ajouter un arrêt", use_container_width=True):
            _flush_editor()
            new_row = pd.DataFrame({
                "Suppr": [False],
                "Action": ["Nettoyer"], "Produit": ["WC chimique"],
                "Option": ["Lave-main"], "Couleur": [""], "Quantité": [1],
                "Nom du client": [""], "Adresse": [""],
                "Durée (min)": [30], "Pas avant": [""], "Pas après": [""],
                "Observations": [""],
            })
            st.session_state.df_stops = pd.concat(
                [st.session_state.df_stops, new_row], ignore_index=True)
            if "editor_stops" in st.session_state:
                del st.session_state["editor_stops"]
            st.rerun()

        if st.button("➖ Supprimer le dernier", use_container_width=True):
            _flush_editor()
            if len(st.session_state.df_stops) > 1:
                st.session_state.df_stops = (
                    st.session_state.df_stops.iloc[:-1].reset_index(drop=True))
            if "editor_stops" in st.session_state:
                del st.session_state["editor_stops"]
            st.rerun()

        if st.button("🗑️ Tout vider", use_container_width=True, type="secondary"):
            st.session_state.df_stops = _init_df()
            st.session_state.result   = None
            reset_manual_order()
            if "editor_stops" in st.session_state:
                del st.session_state["editor_stops"]
            st.rerun()

        # ── Suppression des lignes cochées dans le tableau ──
        st.divider()
        sel_count = 0
        if "Suppr" in live_df.columns:
            sel_count = int(live_df["Suppr"].fillna(False).astype(bool).sum())
        btn_label = ("🗑️ Supprimer la sélection"
                     + (f" ({sel_count})" if sel_count else ""))
        if st.button(btn_label,
                     use_container_width=True,
                     type="secondary",
                     disabled=(sel_count == 0),
                     help="Cochez la colonne 🗑️ des lignes à supprimer dans le "
                          "tableau, puis cliquez ici."):
            df_view = _current_df()
            if "Suppr" in df_view.columns:
                keep_mask = ~df_view["Suppr"].fillna(False).astype(bool)
            else:
                keep_mask = pd.Series([True] * len(df_view), index=df_view.index)
            if int(keep_mask.sum()) == 0:
                st.warning("⚠️ Impossible de supprimer toutes les lignes "
                           "(au moins une doit rester).")
            else:
                keep = df_view[keep_mask].copy()
                keep["Suppr"] = False
                st.session_state.df_stops = keep.reset_index(drop=True)
                if "editor_stops" in st.session_state:
                    del st.session_state["editor_stops"]
                st.rerun()
        if sel_count == 0:
            st.caption("Cochez la colonne 🗑️ dans le tableau pour sélectionner "
                       "des lignes à supprimer.")

    # ══════════════════════════════════════════════════════════════════════════
    # BLOC 3 — PRÉCALCUL DES DURÉES
    # ══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("### 3\ufe0f\u20e3  Précalcul des durées")
    col_auto, col_info = st.columns([2, 3])
    with col_auto:
        if st.button("⏱️ Précalculer les durées (tous produits au barème)",
                     use_container_width=True,
                     help=(
                         "Remplit automatiquement la colonne **Durée (min)** "
                         "selon les barèmes ci-dessous (durée × quantité) :\n\n"
                         "**WC chimique**\n"
                         "• Nettoyer : 10 min/WC\n"
                         "• Déposer : 15 min/WC\n"
                         "• Retirer : 20 min/WC\n"
                         "• Chargement : 15 min/WC\n"
                         "• Déchargement : 5 min/WC\n\n"
                         "**WC handicapé** (options ignorées)\n"
                         "• Nettoyer : 20 min/WC\n"
                         "• Déposer : 30 min/WC\n"
                         "• Retirer : 40 min/WC\n"
                         "• Chargement : 30 min/WC\n"
                         "• Déchargement : 10 min/WC\n\n"
                         "**Urinoir / Lave-main**\n"
                         "• Déposer : 15 min/produit\n"
                         "• Toutes les autres actions : 10 min/produit\n\n"
                         "Vous pouvez ensuite ajuster manuellement."
                     )):
            _flush_editor()
            df_tmp = st.session_state.df_stops.copy()
            nb_maj = 0
            nb_wcc = 0   # WC chimique mis à jour
            nb_wch = 0   # WC handicapé mis à jour
            nb_smp = 0   # Urinoir / Lave-main mis à jour
            for idx, row in df_tmp.iterrows():
                duree_auto = _auto_duree(
                    row.get("Action", ""),
                    row.get("Produit", ""),
                    row.get("Quantité", 1),
                )
                if duree_auto is not None:
                    df_tmp.at[idx, "Durée (min)"] = duree_auto
                    nb_maj += 1
                    produit = str(row.get("Produit", "")).strip()
                    if produit == "WC handicapé":
                        nb_wch += 1
                    elif produit in PRODUITS_SIMPLES:
                        nb_smp += 1
                    else:
                        nb_wcc += 1
            st.session_state.df_stops = df_tmp
            if "editor_stops" in st.session_state:
                del st.session_state["editor_stops"]
            if nb_maj > 0:
                details = []
                if nb_wcc > 0:
                    details.append(f"**{nb_wcc}** WC chimique")
                if nb_wch > 0:
                    details.append(f"**{nb_wch}** WC handicapé")
                if nb_smp > 0:
                    details.append(f"**{nb_smp}** Urinoir/Lave-main")
                st.success(f"✅ Durées précalculées pour {' · '.join(details)}.")
            else:
                st.info("ℹ️ Aucun arrêt au barème automatique trouvé à mettre à jour.")
            st.rerun()
    with col_info:
        st.caption(
            "💡 Les durées sont calculées automatiquement (× quantité) pour les "
            "WC chimiques, les WC handicapés et les Urinoir/Lave-main. "
            "La colonne **Durée (min)** reste modifiable manuellement après précalcul."
        )

    # ══════════════════════════════════════════════════════════════════════════
    # BLOC 4 — OPTIMISATION DE LA TOURNÉE
    # ══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("### 4\ufe0f\u20e3  Optimisation de la tournée")

    if st.button("🚀 Optimiser la tournée", type="primary",
                 use_container_width=True, disabled=(n_valid < 1)):
        _flush_editor()
        valid_stops = st.session_state.df_stops[
            st.session_state.df_stops["Adresse"].fillna("").str.strip() != ""
        ].reset_index(drop=True)

        with st.spinner("🔍 Géocodage des adresses en cours…"):
            # Géocodage des dépôts
            depot_depart_addr = DEPOTS[st.session_state.depot_depart_key]
            depot_retour_addr = DEPOTS[st.session_state.depot_retour_key]
            if not depot_depart_addr:
                st.error("❌ Veuillez renseigner le **Dépôt de départ** dans la barre latérale.")
                st.stop()
            depot_coords = geocode(depot_depart_addr)
            if not depot_coords:
                st.error(f"❌ Impossible de géocoder le dépôt de départ : {depot_depart_addr}")
                st.stop()
            if depot_retour_addr and depot_retour_addr != depot_depart_addr:
                depot_retour_coords = geocode(depot_retour_addr)
                if not depot_retour_coords:
                    st.warning(f"⚠️ Dépôt de retour introuvable, le dépôt de départ sera utilisé.")
                    depot_retour_coords = depot_coords
            else:
                depot_retour_coords = depot_coords

            pb           = st.progress(0, text="Géocodage des arrêts…")
            geo          = {}
            all_addrs    = valid_stops["Adresse"].tolist()
            for i, addr in enumerate(all_addrs):
                geo[addr] = geocode(addr)
                pb.progress((i + 1) / len(all_addrs),
                             text=f"Géocodage : {addr[:50]}…")
            pb.empty()

        failed = [a for a, c in geo.items() if c is None]
        if failed:
            st.warning("⚠️ Ces adresses n'ont pas pu être géocodées :\n"
                       + "\n".join(f"- {a}" for a in failed))

        coords_list = [depot_coords]
        for i, row in valid_stops.iterrows():
            c = geo.get(row["Adresse"])
            if c:
                coords_list.append(c)

        if len(coords_list) < 2:
            st.error("❌ Aucune adresse valide après géocodage.")
            st.stop()

        # Fenêtres temporelles (index 0 = dépôt, pas de contrainte)
        time_windows = [{"earliest": None, "latest": None, "duration": 0}]  # index 0 = dépôt
        for _, row in valid_stops.iterrows():
            if geo.get(row["Adresse"]):
                try:
                    dur = int(float(row.get("Durée (min)", 30) or 30))
                except (ValueError, TypeError):
                    dur = 30
                time_windows.append({
                    "earliest": _parse_hhmm(row.get("Pas avant", "")),
                    "latest":   _parse_hhmm(row.get("Pas après", "")),
                    "duration": dur,
                })

        # ── Mode périphérique ──
        # Si activé et heure de pointe, on injecte des contraintes horaires sur les
        # arrêts intra-rocade AVANT de lancer l'optimiseur. Le moteur TW fera le reste :
        # il planifiera les arrêts extérieurs pendant les bouchons, et les arrêts
        # en centre-ville une fois la circulation fluide.
        eviter_peri  = st.session_state.eviter_peri
        pause_enabled = st.session_state.pause_dejeuner
        legal_min    = (st.session_state.heure_min_depart.hour * 60
                        + st.session_state.heure_min_depart.minute)
        peri_affectes = []   # indices des arrêts dont les TW ont été injectées automatiquement
        if eviter_peri:
            time_windows, peri_affectes = _injecte_contraintes_peri(
                time_windows, coords_list, legal_min
            )

        with st.spinner("🗺️ Calcul de l'itinéraire optimisé…"):
            # Première passe sans heure de départ (ordre pur)
            trip = osrm_trip(coords_list, time_windows=time_windows, depart_min=None)
            if not trip:
                st.stop()
            orig = osrm_route_distance(coords_list)

        # Borne légale / réglementaire configurée dans la sidebar
        # (déjà calculée avant le spinner — on conserve la variable telle quelle)

        # Calcul automatique de l'heure de départ optimale (purement mathématique)
        # On utilise matrix_real (durées OSRM non pénalisées) pour que les seuils
        # de départ soient basés sur les temps de trajet réels.
        depart_min_opt = _compute_optimal_departure(trip["order"], trip["matrix_real"], time_windows)

        if depart_min_opt is not None:
            depart_min_raw = depart_min_opt   # résultat pur sans contrainte légale

            if depart_min_raw < legal_min:
                # Le calcul pur tomberait avant l'heure légale → on clamp
                depart_min = legal_min
                st.warning(
                    f"⚠️ Le calcul théorique suggérait un départ à **{_fmt_min(depart_min_raw)}**, "
                    f"ce qui ne respecte pas l'heure de départ au plus tôt fixée à "
                    f"**{_fmt_min(legal_min)}**. "
                    f"Le départ est donc fixé à **{_fmt_min(depart_min)}**. "
                    f"Certaines contraintes client marquées 'Pas après' risquent de ne pas "
                    f"pouvoir être respectées — vérifiez les arrêts en rouge ci-dessous."
                )
            else:
                depart_min = depart_min_raw
                st.success(
                    f"🕖 **Heure de départ recommandée : {_fmt_min(depart_min)}** "
                    f"— calculée automatiquement pour respecter toutes les contraintes "
                    f"avec le minimum d'attente (départ légal : {_fmt_min(legal_min)})."
                )
        else:
            # Aucune contrainte client : on part à l'heure légale minimale
            depart_min = legal_min
            st.info(
                f"🕖 **Aucune contrainte horaire client renseignée.** "
                f"Heure de départ : **{_fmt_min(depart_min)}** "
                f"(heure légale minimale configurée)."
            )

        # Seconde passe 2-opt avec l'heure de départ calculée.
        # Si le mode périphérique a injecté des contraintes intra-rocade,
        # peri_affectes est non vide → has_tw sera True et la passe s'exécutera.
        has_tw = any(
            tw.get("earliest") is not None or tw.get("latest") is not None
            for tw in time_windows[1:]
        )
        if has_tw:
            with st.spinner("⚙️ Affinage de l'ordre selon les contraintes horaires…"):
                trip2 = osrm_trip(coords_list, time_windows=time_windows, depart_min=depart_min)
                if trip2:
                    trip = trip2

        # Heures d'arrivée réelles — on utilise matrix_real (sans malus) pour des
        # ETAs fidèles à la réalité ; la matrice pénalisée n'a servi qu'à l'optimisation.
        # Passe 1 : sans pause → localise le passage de midi
        arrivals_prelim = _compute_arrivals(
            trip["order"], trip["matrix_real"], depart_min, time_windows)
        pause_idx = _find_pause_position(
            [{"departure_min": a.get("departure_min")} for a in arrivals_prelim])
        # Passe 2 : avec pause injectée (si activée) → toutes les ETAs après la pause OK
        arrivals = _compute_arrivals(
            trip["order"], trip["matrix_real"], depart_min, time_windows,
            pause_after_idx=(pause_idx if pause_enabled else None))

        order         = trip["order"]
        stops_ordered = []
        rank          = 1
        arr_idx       = 0
        for orig_idx in order:
            if orig_idx == 0:
                continue
            sri = orig_idx - 1
            if sri >= len(valid_stops):
                continue
            row    = valid_stops.iloc[sri]
            coords = geo.get(row["Adresse"])
            arr    = arrivals[arr_idx] if arr_idx < len(arrivals) else {}
            try:
                dur_stop = int(float(row.get("Durée (min)", 30) or 30))
            except (ValueError, TypeError):
                dur_stop = 30
            stops_ordered.append({
                "order_num":    rank,
                "action":       row["Action"],
                "quantity":     _qty_label(row),
                "produit":      row.get("Produit", ""),
                "option":       row.get("Option",  ""),
                "couleur":      normalize_couleur(row.get("Couleur", "")),
                "article":      _article_key(row.get("Produit", ""), row.get("Option", "")),
                "qty_num":      int(row.get("Quantité", 1) or 1),
                "client":       row["Nom du client"],
                "address":      row["Adresse"],
                "lat":          coords[0] if coords else None,
                "lon":          coords[1] if coords else None,
                "tw_early":     _parse_hhmm(row.get("Pas avant", "")),
                "tw_late":      _parse_hhmm(row.get("Pas après", "")),
                "duration_min": dur_stop,
                "arrival_min":  arr.get("arrival_min"),
                "departure_min":arr.get("departure_min"),
                "wait_min":     arr.get("wait_min", 0),
                "violated":     arr.get("violated", False),
                "observations": str(row.get("Observations", "") or ""),
            })
            rank    += 1
            arr_idx += 1

        dist_km   = trip["distance_km"]
        # Durée = trajet OSRM + pause déjeuner (seulement si activée)
        dur_min   = trip["duration_min"] + (PAUSE_DEJEUNER_MIN if pause_enabled else 0)
        fuel_l    = dist_km * fuel_conso / 100
        fuel_cost = fuel_l * fuel_price
        km_saved  = max(0, (orig["distance_km"]  - dist_km))  if orig else 0
        min_saved = max(0, (orig["duration_min"] - dur_min)) if orig else 0

        # Avertissements fenêtres non respectables
        violated_stops = [s for s in stops_ordered if s["violated"]]
        if violated_stops:
            lines = ["⚠️ Certaines contraintes horaires ne peuvent pas être respectées :"]
            for s in violated_stops:
                lines.append(
                    f"- **Arrêt {s['order_num']}** {s['address']} : "
                    f"arrivée estimée {_fmt_min(s['arrival_min'])} "
                    f"(limite : {_fmt_min(s['tw_late'])})"
                )
            st.warning("  \n".join(lines))

        # Temps de retour au dépôt de retour via OSRM /route
        if stops_ordered:
            last_stop_coords = [s for s in [
                (stops_ordered[-1]["lat"], stops_ordered[-1]["lon"])
            ] if s[0] is not None]
            if last_stop_coords and depot_retour_coords:
                try:
                    ret_coord_str = f"{last_stop_coords[0][1]},{last_stop_coords[0][0]};{depot_retour_coords[1]},{depot_retour_coords[0]}"
                    ret_r = requests.get(f"{OSRM_URL}/route/v1/driving/{ret_coord_str}",
                                         params={"overview": "false"}, timeout=10)
                    ret_data = ret_r.json()
                    travel_back = ret_data["routes"][0]["duration"] / 60 if ret_data.get("code") == "Ok" else 0
                except Exception:
                    travel_back = (trip["matrix"][order[-1]][order[0]] or 0) / 60
            else:
                travel_back = 0
            last_dep = stops_ordered[-1]["departure_min"] or stops_ordered[-1]["arrival_min"] or depart_min
            # La pause est déjà dans last_dep si elle précède le dernier arrêt ;
            # sinon (pause après le dernier arrêt) on l'ajoute ici — mais uniquement
            # si la pause est activée.
            if pause_enabled and pause_idx >= len(stops_ordered) - 1:
                return_min = last_dep + travel_back + PAUSE_DEJEUNER_MIN
            else:
                return_min = last_dep + travel_back
        else:
            return_min = depart_min

        st.session_state.result = {
            "stops_ordered":      stops_ordered,
            "distance_km":        dist_km,
            "duration_min":       dur_min,
            "fuel_liters":        fuel_l,
            "fuel_cost":          fuel_cost,
            "km_saved":           km_saved,
            "time_saved_min":     min_saved,
            "geometry":           trip["geometry"],
            "depot_coords":       depot_coords,
            "depot_retour_coords":depot_retour_coords,
            "depot_depart_addr":  depot_depart_addr,
            "depot_retour_addr":  depot_retour_addr,
            "depart_min":         depart_min,
            "return_min":         return_min,
            # État de la pause déjeuner (utilisé par les exports, l'affichage et le
            # recalcul manuel pour rester cohérent avec le choix de l'utilisateur)
            "pause_dejeuner":     pause_enabled,
            # Mode des fiches état des lieux : True = une par arrêt, False = fiche
            # généraliste unique pour toute la tournée (respecté par les exports)
            "etat_lieux_par_arret": st.session_state.etat_lieux_par_arret,
            # Périphérique : nb d'arrêts intra-rocade dont les TW ont été injectées
            "peri_affectes_nb":   len(peri_affectes),
            # Matrice de temps de trajet pour le recalcul manuel
            # Indices : 0 = dépôt, k+1 = stops_ordered[k]
            "stop_matrix":        [
                [(trip["matrix_real"][trip["order"][i]][trip["order"][j]] or 0)
                 for j in range(len(trip["order"]))]
                for i in range(len(trip["order"]))
            ],
            # Paramètres carburant (nécessaires pour recalcul manuel)
            "fuel_conso":         fuel_conso,
            "fuel_price":         fuel_price,
            # Besoins matériel de la tournée (récap chargement + contrôle stock)
            "besoins_records":    _besoins_records_pour_export(
                valid_stops, st.session_state.df_stock,
                st.session_state.stock_manutentions),
        }
        # Réinitialiser l'ordre manuel : une nouvelle optimisation annule
        # toute personnalisation précédente
        reset_manual_order()
        st.success("✅ Tournée optimisée ! Consultez l'onglet **Tournée optimisée**.")
        st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 2 — RÉSULTATS
# ══════════════════════════════════════════════════════════════════════════════

with tab_optim:
    if st.session_state.result is None:
        st.info("👈 Saisissez vos arrêts puis cliquez sur **Optimiser la tournée**.")
    else:
        r          = st.session_state.result          # référence = ordre optimisé
        # Source de données active : ordre manuel si existant, sinon optimisé.
        # Même fonction que celle utilisée par l'onglet Export → l'affichage et
        # les documents générés ne peuvent plus diverger.
        display_r  = active_result()
        is_manual  = is_manual_order()
        pause_enabled = display_r.get("pause_dejeuner", True)

        hours = int(display_r["duration_min"] // 60)
        mins  = int(display_r["duration_min"] % 60)

        # ── Titre + badge ──
        title_col, badge_col = st.columns([5, 1])
        with title_col:
            st.subheader("📊 Récapitulatif de la tournée")
        with badge_col:
            if is_manual:
                st.markdown(
                    '<div style="background:#fff3cd;border:1px solid #ffc107;'
                    'border-radius:6px;padding:4px 10px;font-size:0.82em;'
                    'color:#856404;text-align:center;margin-top:8px">'
                    '✏️ Ordre manuel</div>',
                    unsafe_allow_html=True)

        c1, c2, c3, c4, c5, c6, c7, c8 = st.columns(8)
        c1.metric("📍 Arrêts",        len(display_r["stops_ordered"]))
        c2.metric("🛣️ Distance",       f"{display_r['distance_km']:.1f} km")
        c3.metric("⏱️ Durée trajet",   f"{hours}h{mins:02d}")
        c4.metric("⛽ Carburant",      f"{display_r['fuel_liters']:.1f} L")
        c5.metric("💶 Coût carburant", f"{display_r['fuel_cost']:.2f} €")
        c6.metric("⏳ Temps gagné",    f"{int(r['time_saved_min'])} min",
                  delta=f"-{r['km_saved']:.1f} km", delta_color="inverse")
        c7.metric("🕖 Départ dépôt",   _fmt_min(display_r.get("depart_min")))
        c8.metric("🏁 Retour dépôt",   _fmt_min(display_r.get("return_min")))

        # Bandeau pause déjeuner (état actif)
        if pause_enabled:
            st.caption(f"🍽️ **Pause déjeuner de {PAUSE_DEJEUNER_MIN} min** comptabilisée dans cette tournée.")
        else:
            st.caption("⛔ **Pause déjeuner désactivée** — aucune pause n'est comptée dans cette tournée.")

        # Bandeau trafic périphérique
        if r.get("peri_affectes_nb", 0) > 0:
            nb = r["peri_affectes_nb"]
            st.info(
                f"🚦 **Mode périphérique actif** — "
                f"**{nb} arrêt{'s' if nb > 1 else ''}** en centre-ville (intra-rocade) "
                f"{'ont été repoussés' if nb > 1 else 'a été repoussé'} après les heures de pointe "
                f"pour éviter les bouchons sur la rocade toulousaine."
            )

        # Durée totale
        if display_r.get("depart_min") is not None and display_r.get("return_min") is not None:
            total_tour_min = display_r["return_min"] - display_r["depart_min"]
            h_tot = int(total_tour_min // 60)
            m_tot = int(total_tour_min % 60)
            interv_total = sum(s.get("duration_min", 0) for s in display_r["stops_ordered"])
            pause_min_effective = PAUSE_DEJEUNER_MIN if pause_enabled else 0
            trajet_min = max(0, int(display_r["duration_min"] - pause_min_effective))
            if pause_enabled:
                st.info(
                    f"🗓️ **Durée totale de la tournée** (trajet + interventions + pause) : "
                    f"**{h_tot}h{m_tot:02d}** "
                    f"— dont {trajet_min} min de trajet, "
                    f"{interv_total} min d'interventions "
                    f"et **{PAUSE_DEJEUNER_MIN} min de pause déjeuner** 🍽️"
                )
            else:
                st.info(
                    f"🗓️ **Durée totale de la tournée** (trajet + interventions, sans pause) : "
                    f"**{h_tot}h{m_tot:02d}** "
                    f"— dont {trajet_min} min de trajet "
                    f"et {interv_total} min d'interventions. "
                    f"⛔ **Aucune pause déjeuner** n'est comptabilisée."
                )

        st.markdown("---")
        col_map, col_list = st.columns([3, 2])

        # ── Carte ──
        with col_map:
            st.subheader("🗺️ Carte de la tournée")
            m_folium = build_map(
                r["depot_coords"], display_r["stops_ordered"], display_r["geometry"],
                depot_depart_addr=r.get("depot_depart_addr", "Dépôt départ"),
                depot_retour_addr=r.get("depot_retour_addr"),
                depot_retour_coords=r.get("depot_retour_coords"),
            )
            st_folium(m_folium, use_container_width=True, height=520, returned_objects=[])

        # ── Liste des arrêts (drag-and-drop) ──
        with col_list:
            st.subheader("📋 Ordre des arrêts")

            base_stops = r["stops_ordered"]   # référence immuable = ordre optimisé
            n_stops    = len(base_stops)

            # Initialiser ou valider manual_order
            if (st.session_state.manual_order is None or
                    len(st.session_state.manual_order) != n_stops):
                st.session_state.manual_order = list(range(n_stops))
            cur_order = st.session_state.manual_order

            # ── Bouton reset ──
            btn_col, info_col = st.columns([1, 2])
            with btn_col:
                if is_manual:
                    if st.button("↩️ Ordre optimisé", use_container_width=True,
                                 help="Rétablir l'ordre calculé automatiquement"):
                        reset_manual_order()
                        st.rerun()
            with info_col:
                if HAS_SORTABLES:
                    st.caption("↕️ Glissez les cartes pour réorganiser")
                else:
                    st.caption(
                        "📦 Installez `streamlit-sortables` pour activer le glisser-déposer"
                    )

            # ── Drag-and-drop via streamlit-sortables ──
            if HAS_SORTABLES:
                # Chaque label encode l'orig_idx entre crochets pour parsing fiable
                item_labels = []
                for pos, orig_idx in enumerate(cur_order):
                    s = base_stops[orig_idx]
                    disp_stop = (display_r["stops_ordered"][pos]
                                 if pos < len(display_r["stops_ordered"]) else s)
                    item_labels.append(_sortable_label(s, orig_idx, disp_stop))

                # La clé porte un nonce : reset_manual_order() l'incrémente pour
                # forcer le remontage du composant. Sans cela, l'ordre glissé
                # resterait mémorisé côté widget et écraserait la remise à zéro
                # au rerun suivant.
                sorted_labels = _sortables_sort_items(
                    item_labels,
                    key=f"manual_tour_sort_{st.session_state.manual_sort_nonce}",
                )

                # Parser le nouvel ordre depuis les labels ("[3] …" → 3)
                new_order = []
                for lbl in sorted_labels:
                    try:
                        new_order.append(int(lbl.split("]")[0].strip("[")))
                    except (ValueError, IndexError):
                        continue

                # Garde-fou : on n'applique un ordre que s'il constitue bien une
                # permutation complète des arrêts existants. Un parsing partiel
                # (label tronqué, composant désynchronisé) perdrait des arrêts
                # dans la tournée et donc dans les exports.
                _valide = (len(new_order) == n_stops and
                           sorted(new_order) == list(range(n_stops)))

                if new_order and not _valide:
                    st.warning(
                        "⚠️ Réorganisation ignorée : la liste reçue était "
                        "incomplète. Rechargez la page si le problème persiste."
                    )
                elif _valide and new_order != cur_order:
                    with st.spinner("🔄 Recalcul de l'itinéraire…"):
                        new_manual_result = _recalc_manual_route(r, new_order)
                    st.session_state.manual_order  = new_order
                    st.session_state.manual_result = new_manual_result
                    st.rerun()

            # ── Cartes détaillées (affichage) ──
            st.markdown(
                f'<div class="stop-card depot-card"><b>🏭 Dépôt — Départ</b><br>'
                f'<small>{r.get("depot_depart_addr","")}</small></div>',
                unsafe_allow_html=True,
            )

            _stops_display = display_r["stops_ordered"]
            _pause_idx     = _find_pause_position(_stops_display)
            _pause_shown   = False

            for _si, stop in enumerate(_stops_display):
                # ── Carte de l'arrêt ──
                bg  = ACTION_COLORS.get(stop["action"], "#f0f0f0")
                brd = ACTION_BORDER_COLORS.get(stop["action"], "#999")
                cli = f" · {stop['client']}" if stop['client'] else ""
                tw_parts = []
                if stop.get("tw_early"):
                    tw_parts.append(f"⏰ Pas avant {_fmt_min(stop['tw_early'])}")
                if stop.get("tw_late"):
                    tw_parts.append(f"⏰ Pas après {_fmt_min(stop['tw_late'])}")
                tw_html   = (f"<br><small style='color:#666'>"
                             + " · ".join(tw_parts) + "</small>") if tw_parts else ""
                arr_str   = _fmt_min(stop.get("arrival_min"))
                wait      = stop.get("wait_min", 0)
                wait_html = (f" <small style='color:#999'>(attente {int(wait)} min)</small>"
                             if wait and wait > 0.5 else "")
                violated  = stop.get("violated", False)
                arr_color = "#dc3545" if violated else "#28a745"
                arr_icon  = "⚠️" if violated else "🕐"
                dep_str   = _fmt_min(stop.get("departure_min"))
                dur_str   = stop.get("duration_min", 0)
                dur_html  = (f" <small style='color:#555'>"
                             f"(intervention : {dur_str} min → départ {dep_str})</small>"
                             if dur_str and dep_str else "")
                arr_html  = (f"<br><small style='color:{arr_color};font-weight:600'>"
                             f"{arr_icon} Arrivée estimée : {arr_str}</small>"
                             f"{wait_html}{dur_html}"
                             if arr_str else "")
                coul_stop = stop.get("couleur", "") or ""
                coul_html = (" " + _pastille_couleur(coul_stop)) if coul_stop else ""
                st.markdown(
                    f'<div class="stop-card" style="background:{bg};'
                    f'border-left:4px solid {brd};">'
                    f'<b>#{stop["order_num"]} {stop["action"]}</b>{cli}<br>'
                    f'<small>📍 {stop["address"]}</small><br>'
                    f'<small>📦 {stop["quantity"]}</small>{coul_html}'
                    f'{tw_html}{arr_html}</div>',
                    unsafe_allow_html=True,
                )
                # ── Carte pause déjeuner (insérée après _pause_idx) — si activée ──
                if pause_enabled and not _pause_shown and _si == _pause_idx:
                    _ps, _pe = _pause_slot(_stops_display, _pause_idx)
                    _slot_txt = f"{_ps} – {_pe}" if _ps else f"{PAUSE_DEJEUNER_MIN} min"
                    st.markdown(
                        f'<div class="stop-card" style="background:#FFF9C4;'
                        f'border-left:4px solid #F9A825;">'
                        f'<b>🍽️ Pause déjeuner — {PAUSE_DEJEUNER_MIN} min</b><br>'
                        f'<small style="color:#795548">🕐 Créneau : <b>{_slot_txt}</b></small>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                    _pause_shown = True

            retour_str  = _fmt_min(display_r.get("return_min"))
            retour_html = (f"<br><small style='color:#1f4e79;font-weight:600'>"
                           f"🏁 Retour estimé : {retour_str}</small>"
                           if retour_str else "")
            st.markdown(
                f'<div class="stop-card depot-card"><b>🏁 Dépôt — Retour</b><br>'
                f'<small>{r.get("depot_retour_addr","")}</small>{retour_html}</div>',
                unsafe_allow_html=True,
            )

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 3 — STOCK & DISPONIBILITÉ
# ══════════════════════════════════════════════════════════════════════════════

with tab_stock:

    def _matrice_html(df, articles, couleurs, metrique, seuil):
        """Construit la matrice HTML Article × Couleur.

        metrique : 'En stock' | 'Installés' | 'Parc total'
        Code couleur appliqué uniquement sur la métrique 'En stock' :
            0 → rupture (rouge) · ≤ seuil → tension (orange) · sinon → OK (vert)
        """
        th = ('padding:8px 10px;text-align:center;font-size:0.82em;'
              'border-bottom:2px solid #1F4E79;white-space:nowrap')
        td = ('padding:7px 10px;text-align:center;font-size:0.95em;'
              'border-bottom:1px solid #ECECEC')
        html = ['<div style="overflow-x:auto"><table style="width:100%;'
                'border-collapse:collapse;background:#fff;border-radius:8px">']
        html.append('<thead><tr>')
        html.append(f'<th style="{th};text-align:left;color:#1F4E79">Article</th>')
        for c in couleurs:
            html.append(f'<th style="{th}">{_pastille_couleur(c)}</th>')
        html.append(f'<th style="{th};color:#1F4E79">Total</th></tr></thead><tbody>')

        totaux_col = {c: 0 for c in couleurs}
        total_gen  = 0
        for a in articles:
            sub = df[df["Article"] == a]
            html.append('<tr>')
            html.append(f'<td style="{td};text-align:left;font-weight:600;'
                        f'color:#263238">{a}</td>')
            ligne_total = 0
            for c in couleurs:
                s = sub[sub["Couleur"] == c]
                if metrique == "Parc total":
                    v = int(s["Installés"].sum() + s["En stock"].sum())
                else:
                    v = int(s[metrique].sum())
                ligne_total += v
                totaux_col[c] += v
                if metrique == "En stock" and len(s) > 0:
                    if v == 0:
                        style = "background:#FDECEA;color:#B00020;font-weight:700"
                    elif v <= seuil:
                        style = "background:#FFF6E5;color:#B26A00;font-weight:700"
                    else:
                        style = "background:#EDF7ED;color:#1B5E20;font-weight:700"
                elif v == 0:
                    style = "color:#C7C7C7"
                else:
                    style = "color:#263238;font-weight:600"
                aff = v if (v != 0 or len(s) > 0) else "·"
                html.append(f'<td style="{td};{style}">{aff}</td>')
            total_gen += ligne_total
            html.append(f'<td style="{td};font-weight:700;color:#1F4E79;'
                        f'background:#F4F7FB">{ligne_total}</td></tr>')

        html.append(f'<tr><td style="{td};text-align:left;font-weight:700;'
                    f'color:#1F4E79;background:#F4F7FB">TOTAL</td>')
        for c in couleurs:
            html.append(f'<td style="{td};font-weight:700;color:#1F4E79;'
                        f'background:#F4F7FB">{totaux_col[c]}</td>')
        html.append(f'<td style="{td};font-weight:800;color:#fff;'
                    f'background:#1F4E79">{total_gen}</td></tr>')
        html.append('</tbody></table></div>')
        return "".join(html)

    df_stock = _clean_stock_df(st.session_state.df_stock)

    # ══════════════════════════════════════════════════════════════════════════
    # BLOC 1 — SOURCE DES DONNÉES ET FICHIERS
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("### 1\ufe0f\u20e3  Source des données")

    if st.session_state.stock_import_ok:
        st.success(
            f"📥 Données issues du fichier **{st.session_state.stock_import_name}**, "
            f"importé dans cette session."
        )
    else:
        st.warning(
            "⚠️ **Aucun fichier importé dans cette session.** Les valeurs affichées "
            "proviennent du relevé de référence de l'exploitation. "
            "L'application ne conserve rien entre deux sessions : importez le fichier "
            "du jour depuis la barre latérale (**📦 Stock du parc**) pour travailler "
            "sur des données à jour."
        )

    col_dl1, col_dl2, col_dl3 = st.columns(3)
    _today_str = datetime.date.today().strftime("%d/%m/%Y")
    _file_str  = datetime.date.today().strftime("%d-%m-%Y")

    with col_dl1:
        st.download_button(
            "⬇️ Exporter le stock actuel",
            data=export_stock_excel(df_stock, maj_date=_today_str),
            file_name=f"Stock_parc_{_file_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary",
            help="Classeur directement réimportable au prochain démarrage "
                 "(onglets Stock + Synthèse).")
    with col_dl2:
        _modele = _stock_default_df().copy()
        _modele[["Installés", "En stock"]] = 0
        st.download_button(
            "📄 Télécharger un modèle vierge",
            data=export_stock_excel(_modele, maj_date=_today_str),
            file_name="Modele_stock_parc.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Structure attendue à l'import : Article · Couleur · Installés · En stock.")
    with col_dl3:
        if st.button("♻️ Restaurer le relevé de référence",
                     use_container_width=True,
                     help="Revient aux valeurs du relevé transmis par l'exploitation."):
            st.session_state.df_stock          = _stock_default_df()
            st.session_state.stock_import_ok   = False
            st.session_state.stock_import_name = None
            st.session_state.stock_source      = "Valeurs de référence (relevé exploitation)"
            if "editor_stock" in st.session_state:
                del st.session_state["editor_stock"]
            st.rerun()

    # ══════════════════════════════════════════════════════════════════════════
    # BLOC 2 — INDICATEURS DE PARC
    # ══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("### 2\ufe0f\u20e3  Vue d'ensemble du parc")

    _tot_ins   = int(df_stock["Installés"].sum())
    _tot_stk   = int(df_stock["En stock"].sum())
    _tot_parc  = _tot_ins + _tot_stk
    _nb_ref    = len(df_stock)
    _nb_rup    = int((df_stock["En stock"] == 0).sum())
    _taux      = (_tot_ins / _tot_parc * 100) if _tot_parc else 0

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("🏗️ Parc total",      _tot_parc)
    k2.metric("📦 En stock",        _tot_stk)
    k3.metric("🏠 Installés",       _tot_ins)
    k4.metric("📈 Taux d'usage",    f"{_taux:.0f} %",
              help="Part du parc actuellement installée chez les clients.")
    k5.metric("🔴 Réf. en rupture", _nb_rup,
              delta=f"sur {_nb_ref} réf.", delta_color="off")

    # ══════════════════════════════════════════════════════════════════════════
    # BLOC 3 — MATRICE ARTICLE × COULEUR
    # ══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("### 3\ufe0f\u20e3  Tableau du parc par produit et couleur")

    f1, f2, f3 = st.columns([2, 2, 2])
    with f1:
        _metrique = st.radio(
            "Donnée affichée",
            ["En stock", "Installés", "Parc total"],
            horizontal=True,
            help="« En stock » = disponible au dépôt · « Installés » = chez les clients.")
    with f2:
        _couleurs_dispo = [c for c in COULEURS_PRODUITS
                           if c in set(df_stock["Couleur"].astype(str))]
        _couleurs_autres = sorted(set(df_stock["Couleur"].astype(str))
                                  - set(COULEURS_PRODUITS) - {""})
        _couleurs_dispo += list(_couleurs_autres)
        _filtre_couleurs = st.multiselect(
            "🎨 Filtrer par couleur", options=_couleurs_dispo,
            default=_couleurs_dispo,
            help="Décochez une couleur pour la masquer du tableau et des totaux.")
    with f3:
        _articles_dispo = ([a for a in ARTICLES_STOCK if a in set(df_stock["Article"])]
                           + sorted(set(df_stock["Article"]) - set(ARTICLES_STOCK)))
        _filtre_articles = st.multiselect(
            "🧱 Filtrer par article", options=_articles_dispo,
            default=_articles_dispo)

    st.slider(
        "Seuil d'alerte (tension de stock)", min_value=0, max_value=5,
        key="stock_seuil_alerte",
        help="Une quantité inférieure ou égale à ce seuil est signalée en orange "
             "dans le tableau. Zéro reste toujours signalé en rouge.")

    _df_f = df_stock[df_stock["Couleur"].isin(_filtre_couleurs) &
                     df_stock["Article"].isin(_filtre_articles)]

    if len(_df_f) == 0:
        st.info("Aucune référence ne correspond aux filtres sélectionnés.")
    else:
        st.markdown(
            _matrice_html(_df_f, _filtre_articles, _filtre_couleurs,
                          _metrique, int(st.session_state.stock_seuil_alerte)),
            unsafe_allow_html=True)
        if _metrique == "En stock":
            st.caption(
                "🟢 disponible · 🟠 tension (≤ seuil d'alerte) · 🔴 rupture · "
                "« · » : couleur non référencée pour cet article.")
        else:
            st.caption("« · » : couleur non référencée pour cet article.")

    # ══════════════════════════════════════════════════════════════════════════
    # BLOC 4 — TABLEAU DÉTAILLÉ MODIFIABLE
    # ══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("### 4\ufe0f\u20e3  Détail modifiable")
    st.caption(
        "Ajustez directement les quantités si le relevé a évolué depuis l'import. "
        "Les modifications ne sont conservées que le temps de la session : "
        "exportez le fichier après correction (bloc 1)."
    )

    _detail = df_stock.copy()
    _detail["Parc total"] = _detail["Installés"] + _detail["En stock"]

    _edit = st.data_editor(
        _detail,
        use_container_width=True,
        num_rows="dynamic",
        column_order=STOCK_COLUMNS + ["Parc total"],
        column_config={
            "Article": st.column_config.SelectboxColumn(
                "Article", required=True, width="large", options=ARTICLES_STOCK),
            "Couleur": st.column_config.SelectboxColumn(
                "🎨 Couleur", width="small", options=[""] + COULEURS_PRODUITS),
            "Installés": st.column_config.NumberColumn(
                "🏠 Installés", min_value=0, max_value=999, step=1, default=0,
                width="small"),
            "En stock": st.column_config.NumberColumn(
                "📦 En stock", min_value=0, max_value=999, step=1, default=0,
                width="small"),
            "Parc total": st.column_config.NumberColumn(
                "Σ Parc total", disabled=True, width="small",
                help="Calculé : Installés + En stock. Recalculé après validation."),
        },
        hide_index=True,
        key="editor_stock",
    )

    c_save, c_info = st.columns([1, 3])
    with c_save:
        if st.button("💾 Appliquer les modifications", use_container_width=True,
                     type="primary"):
            st.session_state.df_stock = _clean_stock_df(
                _edit.drop(columns=[c for c in ("Parc total",) if c in _edit.columns]))
            st.session_state.stock_source = "Saisie manuelle dans l'application"
            del st.session_state["editor_stock"]
            st.rerun()
    with c_info:
        st.caption(
            "💡 Les doublons (même article, même couleur) sont automatiquement "
            "fusionnés à la validation."
        )

    # ══════════════════════════════════════════════════════════════════════════
    # BLOC 5 — CONFRONTATION AVEC LA TOURNÉE EN COURS
    # ══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("### 5\ufe0f\u20e3  Disponibilité pour la tournée en cours")

    st.toggle(
        "Comptabiliser les actions Chargement / Déchargement comme mouvements de stock",
        key="stock_manutentions",
        help=(
            "**Désactivé (recommandé par défaut)** : seules les actions *Déposer* "
            "(sortie de stock) et *Retirer* (retour en stock) génèrent un mouvement. "
            "*Chargement* et *Déchargement* sont considérés comme des manutentions "
            "internes sans impact sur le parc.\n\n"
            "**Activé** : *Chargement* = retour au dépôt (+1 en stock), "
            "*Déchargement* = dépose sur site (−1 en stock). "
            "À activer uniquement si cette convention correspond à votre exploitation."
        ),
    )

    _df_saisie = st.session_state.df_stops.copy()
    _besoins   = besoins_tournee(
        _df_saisie, inclure_manutentions=st.session_state.stock_manutentions)
    _controle  = controle_disponibilite(df_stock, _besoins)

    if len(_besoins) == 0:
        st.info(
            "ℹ️ Aucun mouvement de matériel n'est généré par la saisie actuelle. "
            "Renseignez des arrêts avec une action *Déposer* ou *Retirer* et une "
            "adresse pour obtenir le contrôle de disponibilité."
        )
    else:
        _nb_manque = int((_controle["Manque"] > 0).sum()) if len(_controle) else 0
        if _nb_manque > 0:
            st.error(
                f"🔴 **{_nb_manque} référence(s) en rupture** pour cette tournée. "
                "Le détail ci-dessous indique la quantité manquante."
            )
        elif len(_controle) > 0:
            st.success("🟢 **Toutes les sorties de matériel sont couvertes par le stock.**")

        cbes, cctl = st.columns(2)
        with cbes:
            st.markdown("**📤 Mouvements générés par la tournée**")
            _aff_bes = _besoins.copy()
            _aff_bes["Couleur"] = _aff_bes["Couleur"].replace("", "Indifférente")
            st.dataframe(
                _aff_bes, use_container_width=True, hide_index=True,
                column_config={
                    "Sorties": st.column_config.NumberColumn("📤 À charger"),
                    "Retours": st.column_config.NumberColumn("📥 À rapporter"),
                    "Net":     st.column_config.NumberColumn(
                        "Δ Stock", help="Variation prévisionnelle du stock au dépôt."),
                })
        with cctl:
            st.markdown("**✅ Contrôle de disponibilité**")
            if len(_controle) == 0:
                st.caption("Aucune sortie de matériel à contrôler "
                           "(la tournée ne génère que des retours).")
            else:
                st.dataframe(_controle, use_container_width=True, hide_index=True)

        # ── Stock projeté après exécution de la tournée ──
        st.markdown("**🔮 Parc projeté après exécution de la tournée**")
        st.caption(
            "Simulation appliquant les mouvements ci-dessus au parc actuel. "
            "Les besoins sans couleur imposée sont imputés sur la couleur la mieux "
            "dotée. Exportez ce fichier en fin de journée pour disposer d'un point "
            "de départ à jour lors de la prochaine session."
        )
        _projete = projeter_stock(df_stock, _besoins)
        _comp = _projete.merge(
            df_stock, on=["Article", "Couleur"], how="left",
            suffixes=("", "_avant")).fillna(0)
        _comp["Δ Stock"] = (_comp["En stock"] - _comp["En stock_avant"]).astype(int)
        _comp = _comp[_comp["Δ Stock"] != 0][
            ["Article", "Couleur", "En stock_avant", "En stock", "Δ Stock"]]
        _comp = _comp.rename(columns={"En stock_avant": "Avant", "En stock": "Après"})

        if len(_comp) == 0:
            st.caption("Aucune variation de stock à l'issue de cette tournée.")
        else:
            st.dataframe(_comp, use_container_width=True, hide_index=True)

        st.download_button(
            "⬇️ Exporter le parc projeté (fin de tournée)",
            data=export_stock_excel(df_stock, df_projete=_projete, maj_date=_today_str),
            file_name=f"Stock_parc_projete_{_file_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Contient l'état avant tournée (onglet Stock) et l'état projeté "
                 "après tournée (onglet Stock projeté).")

    with st.expander("📘 Règles de gestion appliquées"):
        st.markdown(
            "**Constitution des références**  \n"
            "La référence article est déduite du couple *Produit* / *Option* saisi "
            "dans le tableau des arrêts : un **WC chimique** avec l'option "
            "*Lave-main* correspond à l'article **WC chimique + Lave-main**. "
            "Le produit **WC client** appartient au client et n'est jamais décompté "
            "du parc.\n\n"
            "**Mouvements de stock**  \n"
            "• *Déposer* : −1 en stock, +1 installé  \n"
            "• *Retirer* : +1 en stock, −1 installé  \n"
            "• *Nettoyer* : aucun mouvement (intervention sur place)  \n"
            "• *Chargement* / *Déchargement* : neutres par défaut, configurables "
            "via l'interrupteur ci-dessus\n\n"
            "**Gestion de la couleur**  \n"
            "La couleur est un critère facultatif. Renseignée, la disponibilité est "
            "contrôlée sur cette couleur précise. Laissée vide, elle est contrôlée "
            "toutes couleurs confondues pour l'article concerné.\n\n"
            "**Persistance**  \n"
            "Streamlit ne conserve aucune donnée entre deux sessions. Le cycle de "
            "travail est : *importer le fichier au démarrage → travailler → exporter "
            "le fichier mis à jour en fin de journée*."
        )

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 4 — EXPORT
# ══════════════════════════════════════════════════════════════════════════════

with tab_export:
    if active_result() is None:
        st.info("👈 Optimisez d'abord une tournée pour pouvoir l'exporter.")
    else:
        # ⚠️ CORRECTIF — On exporte le résultat ACTIF, pas st.session_state.result.
        # Si l'utilisateur a réorganisé sa tournée par glisser-déposer dans
        # l'onglet « Tournée optimisée », c'est son ordre qui doit figurer dans
        # le PDF et le XLSX. Lire st.session_state.result ici réintroduirait le
        # bug (export de l'ordre optimisé malgré la réorganisation manuelle).
        r          = active_result()
        _manual    = is_manual_order()
        st.subheader("📥 Exporter la feuille de tournée")

        # ── Bandeau d'état de l'ordre exporté ──
        if _manual:
            st.success(
                "✏️ **Ordre manuel pris en compte** — les documents générés "
                "reprennent exactement l'ordre des arrêts que vous avez défini "
                "par glisser-déposer dans l'onglet **🗺️ Tournée optimisée**, "
                "avec les heures d'arrivée, la distance et le carburant "
                "recalculés sur cet ordre."
            )
            if not r.get("recalc_ok", True):
                st.warning(
                    "⚠️ Le service de calcul d'itinéraire n'a pas répondu lors "
                    "du dernier glisser-déposer : **l'ordre des arrêts est bien "
                    "le vôtre**, mais la distance et la durée affichées restent "
                    "celles de l'ordre optimisé. Réordonnez un arrêt pour "
                    "relancer le calcul si la connexion est rétablie."
                )
        else:
            st.caption(
                "🤖 Ordre **optimisé automatiquement**. Pour exporter un autre "
                "ordre, réorganisez les arrêts dans l'onglet "
                "**🗺️ Tournée optimisée** : l'export suivra."
            )

        # Contrôle de cohérence : l'ordre exporté doit être celui affiché.
        with st.expander("🔎 Vérifier l'ordre qui sera exporté", expanded=False):
            _ctrl = pd.DataFrame([
                {
                    "Ordre":   s.get("order_num", i + 1),
                    "Action":  s.get("action", ""),
                    "Client":  s.get("client", ""),
                    "Adresse": s.get("address", ""),
                    "Arrivée": _fmt_min(s.get("arrival_min")) or "—",
                }
                for i, s in enumerate(r.get("stops_ordered", []))
            ])
            st.dataframe(_ctrl, use_container_width=True, hide_index=True)
            st.caption(
                "Cette liste est celle qui sera écrite dans le PDF et le XLSX. "
                "Elle doit être identique à l'onglet **Tournée optimisée**."
            )

        # Rappel de l'état de la pause déjeuner tel qu'il sera exporté
        if r.get("pause_dejeuner", True):
            st.caption(f"🍽️ Les documents incluront la **pause déjeuner de {PAUSE_DEJEUNER_MIN} min**.")
        else:
            st.caption("⛔ Les documents seront générés **sans pause déjeuner**.")

        # Rappel du mode des fiches état des lieux tel qu'il sera exporté
        if r.get("etat_lieux_par_arret", True):
            _nb = len(r.get("stops_ordered", []))
            st.caption(
                f"📋 Les exports incluront **{_nb} fiche(s) « État des lieux »** "
                "(une par arrêt), pré-remplies et à signer par l'opérateur."
            )
        else:
            st.caption(
                "🗂️ Les exports incluront **une seule fiche « État des lieux » "
                "généraliste** couvrant toute la tournée, à compléter et signer."
            )

        # Suffixe de nom de fichier : évite de confondre, sur le poste du
        # chauffeur, une feuille de tournée optimisée et une feuille réordonnée.
        _suffix   = "_ordre-manuel" if _manual else ""
        _date_str = st.session_state.tour_date.strftime('%d-%m-%Y')

        col_xl, col_pdf = st.columns(2)

        with col_xl:
            st.markdown("### 📊 Excel (.xlsx)")
            st.write("Feuille de route mise en forme avec codes couleur, "
                     "colonne heure de passage et case à cocher ✓ Fait. "
                     "Feuille **État des lieux** incluse.")
            xl_buf = export_excel(r, st.session_state.tour_date,
                                  st.session_state.driver, fuel_price)
            st.download_button(
                "⬇️ Télécharger Excel", data=xl_buf,
                file_name=f"Tournee_WC_{_date_str}{_suffix}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary")

        with col_pdf:
            st.markdown("### 📄 PDF")
            st.write("Document imprimable prêt pour le chauffeur, "
                     "avec tableau coloré et récapitulatif des indicateurs. "
                     "Page **État des lieux** incluse.")
            pdf_buf = export_pdf(r, st.session_state.tour_date,
                                 st.session_state.driver)
            st.download_button(
                "⬇️ Télécharger PDF", data=pdf_buf,
                file_name=f"Tournee_WC_{_date_str}{_suffix}.pdf",
                mime="application/pdf",
                use_container_width=True, type="primary")

        st.markdown("---")
        st.caption("💡 Le fichier Excel contient une colonne **Heure de passage** "
                   "à renseigner manuellement et une colonne **✓ Fait** pour validation terrain.")
